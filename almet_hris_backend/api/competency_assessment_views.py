# api/competency_assessment_views.py

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Count, Avg
from django.db import transaction
from django.utils import timezone
from django.http import HttpResponse
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

from .competency_assessment_models import (
    CoreCompetencyScale, BehavioralScale, LetterGradeMapping,
    PositionCoreAssessment, PositionBehavioralAssessment,
    EmployeeCoreAssessment, EmployeeBehavioralAssessment
)
from .competency_assessment_serializers import (
    CoreCompetencyScaleSerializer, BehavioralScaleSerializer, LetterGradeMappingSerializer,
    PositionCoreAssessmentSerializer, PositionCoreAssessmentCreateSerializer,
    PositionBehavioralAssessmentSerializer, PositionBehavioralAssessmentCreateSerializer,
    EmployeeCoreAssessmentSerializer, EmployeeCoreAssessmentCreateSerializer,
    EmployeeBehavioralAssessmentSerializer, EmployeeBehavioralAssessmentCreateSerializer,
    
)
from .models import Employee, PositionGroup

import logging

logger = logging.getLogger(__name__)


# SCALE MANAGEMENT VIEWS

class CoreCompetencyScaleViewSet(viewsets.ModelViewSet):
    """Core Competency Scale Management"""
    queryset = CoreCompetencyScale.objects.all()
    serializer_class = CoreCompetencyScaleSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = CoreCompetencyScale.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(scale__icontains=search) | Q(description__icontains=search)
            )
        return queryset.order_by('scale')


class BehavioralScaleViewSet(viewsets.ModelViewSet):
    """Behavioral Scale Management"""
    queryset = BehavioralScale.objects.all()
    serializer_class = BehavioralScaleSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = BehavioralScale.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(scale__icontains=search) | Q(description__icontains=search)
            )
        return queryset.order_by('scale')


class LetterGradeMappingViewSet(viewsets.ModelViewSet):
    """Letter Grade Mapping Management"""
    queryset = LetterGradeMapping.objects.all()
    serializer_class = LetterGradeMappingSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return LetterGradeMapping.objects.all().order_by('-min_percentage')
    
    @swagger_auto_schema(
        method='get',
        operation_description='Get letter grade for given percentage',
        manual_parameters=[
            openapi.Parameter(
                'percentage',
                openapi.IN_QUERY,
                description='The percentage value (0-100)',
                type=openapi.TYPE_NUMBER,
                required=True,
                example=85
            )
        ],
        responses={
            200: openapi.Response(
                description='Letter grade information',
                examples={
                    'application/json': {
                        'percentage': 85,
                        'letter_grade': 'B',
                        'description': 'Good performance'
                    }
                }
            ),
            400: openapi.Response(
                description='Bad Request',
                examples={
                    'application/json': {
                        'error': 'Percentage parameter required'
                    }
                }
            )
        }
    )
    @action(detail=False, methods=['get'])
    def get_grade_for_percentage(self, request):
        """Get letter grade for given percentage"""
        percentage = request.query_params.get('percentage')
        if not percentage:
            return Response({
                'error': 'Percentage parameter required',
                'help': 'Add ?percentage=85 to your request URL'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            pct = float(percentage)
            if pct < 0 or pct > 100:
                return Response({
                    'error': 'Percentage must be between 0 and 100'
                }, status=status.HTTP_400_BAD_REQUEST)
                
            grade = LetterGradeMapping.get_letter_grade(pct)
            grade_obj = LetterGradeMapping.objects.filter(letter_grade=grade).first()
            
            return Response({
                'percentage': pct,
                'letter_grade': grade,
                'description': grade_obj.description if grade_obj else ''
            })
        except ValueError:
            return Response({
                'error': 'Invalid percentage value. Must be a number.'
            }, status=status.HTTP_400_BAD_REQUEST)


# POSITION ASSESSMENT VIEWS

class PositionCoreAssessmentViewSet(viewsets.ModelViewSet):
    """Position Core Competency Assessment Templates"""
    queryset = PositionCoreAssessment.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return PositionCoreAssessmentCreateSerializer
        return PositionCoreAssessmentSerializer
    
    def get_queryset(self):
        queryset = PositionCoreAssessment.objects.select_related(
            'position_group', 'created_by'
        ).prefetch_related('competency_ratings__skill__group')
        
        # Filter by position group
        position_group = self.request.query_params.get('position_group')
        if position_group:
            queryset = queryset.filter(position_group_id=position_group)
        
        # Search by job title
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(job_title__icontains=search)
        
        return queryset.order_by('position_group__hierarchy_level', 'job_title')
    
    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Duplicate position assessment for new job title"""
        original = self.get_object()
        new_job_title = request.data.get('job_title')
        
        if not new_job_title:
            return Response({'error': 'job_title is required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                # Create new position assessment
                new_assessment = PositionCoreAssessment.objects.create(
                    position_group=original.position_group,
                    job_title=new_job_title,
                    created_by=request.user
                )
                
                # Copy all competency ratings
                for rating in original.competency_ratings.all():
                    rating.pk = None
                    rating.position_assessment = new_assessment
                    rating.save()
                
                serializer = PositionCoreAssessmentSerializer(new_assessment)
                return Response({
                    'success': True,
                    'message': f'Position assessment duplicated for {new_job_title}',
                    'assessment': serializer.data
                })
        except Exception as e:
            return Response({
                'error': f'Failed to duplicate assessment: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @swagger_auto_schema(
        method='get',
        operation_description='Get position assessment template for specific employee',
        manual_parameters=[
            openapi.Parameter(
                'employee_id',
                openapi.IN_QUERY,
                description='The employee ID',
                type=openapi.TYPE_INTEGER,
                required=True,
                example=63
            )
        ],
        responses={
            200: openapi.Response(
                description='Employee position assessment template',
                examples={
                    'application/json': {
                        'employee_info': {
                            'id': 63,
                            'name': 'John Doe',
                            'job_title': 'Software Engineer',
                            'position_group': 'Engineering'
                        },
                        'assessment_template': {
                            'id': 'uuid',
                            'job_title': 'Software Engineer',
                            'competency_ratings': []
                        }
                    }
                }
            ),
            400: openapi.Response(
                description='Bad Request',
                examples={
                    'application/json': {
                        'error': 'employee_id is required'
                    }
                }
            ),
            404: openapi.Response(
                description='Not Found',
                examples={
                    'application/json': {
                        'error': 'Employee not found'
                    }
                }
            )
        }
    )
    @action(detail=False, methods=['get'])
    def get_for_employee(self, request):
        """Get position assessment template for specific employee"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({
                'error': 'employee_id is required',
                'help': 'Add ?employee_id=63 to your request URL'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            employee = Employee.objects.get(id=employee_id)
            
            # Find matching position assessment
            assessment = PositionCoreAssessment.objects.filter(
                position_group=employee.position_group,
                job_title=employee.job_title,
                is_active=True
            ).first()
            
            if not assessment:
                return Response({
                    'error': f'No core assessment template found for {employee.job_title}',
                    'employee_info': {
                        'id': employee.id,
                        'name': employee.full_name,
                        'job_title': employee.job_title,
                        'position_group': employee.position_group.get_name_display()
                    }
                }, status=status.HTTP_404_NOT_FOUND)
            
            serializer = PositionCoreAssessmentSerializer(assessment)
            return Response({
                'employee_info': {
                    'id': employee.id,
                    'name': employee.full_name,
                    'job_title': employee.job_title,
                    'position_group': employee.position_group.get_name_display()
                },
                'assessment_template': serializer.data
            })
            
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, 
                          status=status.HTTP_404_NOT_FOUND)
        except ValueError:
            return Response({'error': 'Invalid employee_id format'}, 
                          status=status.HTTP_400_BAD_REQUEST)


class PositionBehavioralAssessmentViewSet(viewsets.ModelViewSet):
    """Position Behavioral Competency Assessment Templates"""
    queryset = PositionBehavioralAssessment.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return PositionBehavioralAssessmentCreateSerializer
        return PositionBehavioralAssessmentSerializer
    
    def get_queryset(self):
        queryset = PositionBehavioralAssessment.objects.select_related(
            'position_group', 'created_by'
        ).prefetch_related('competency_ratings__behavioral_competency__group')
        
        position_group = self.request.query_params.get('position_group')
        if position_group:
            queryset = queryset.filter(position_group_id=position_group)
        
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(job_title__icontains=search)
        
        return queryset.order_by('position_group__hierarchy_level', 'job_title')
    
    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Duplicate behavioral position assessment for new job title"""
        original = self.get_object()
        new_job_title = request.data.get('job_title')
        
        if not new_job_title:
            return Response({'error': 'job_title is required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                # Create new position assessment
                new_assessment = PositionBehavioralAssessment.objects.create(
                    position_group=original.position_group,
                    job_title=new_job_title,
                    created_by=request.user
                )
                
                # Copy all competency ratings
                for rating in original.competency_ratings.all():
                    rating.pk = None
                    rating.position_assessment = new_assessment
                    rating.save()
                
                serializer = PositionBehavioralAssessmentSerializer(new_assessment)
                return Response({
                    'success': True,
                    'message': f'Behavioral position assessment duplicated for {new_job_title}',
                    'assessment': serializer.data
                })
        except Exception as e:
            return Response({
                'error': f'Failed to duplicate behavioral assessment: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @swagger_auto_schema(
        method='get',
        operation_description='Get behavioral assessment template for specific employee',
        manual_parameters=[
            openapi.Parameter(
                'employee_id',
                openapi.IN_QUERY,
                description='The employee ID',
                type=openapi.TYPE_INTEGER,
                required=True,
                example=63
            )
        ],
        responses={
            200: openapi.Response(
                description='Employee behavioral assessment template',
                examples={
                    'application/json': {
                        'employee_info': {
                            'id': 63,
                            'name': 'John Doe',
                            'job_title': 'Software Engineer',
                            'position_group': 'Engineering'
                        },
                        'assessment_template': {
                            'id': 'uuid',
                            'job_title': 'Software Engineer',
                            'competency_ratings': []
                        }
                    }
                }
            ),
            400: openapi.Response(
                description='Bad Request',
                examples={
                    'application/json': {
                        'error': 'employee_id is required'
                    }
                }
            ),
            404: openapi.Response(
                description='Not Found',
                examples={
                    'application/json': {
                        'error': 'Employee not found'
                    }
                }
            )
        }
    )
    @action(detail=False, methods=['get'])
    def get_for_employee(self, request):
        """Get behavioral assessment template for specific employee"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({
                'error': 'employee_id is required',
                'help': 'Add ?employee_id=63 to your request URL'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            employee = Employee.objects.get(id=employee_id)
            
            assessment = PositionBehavioralAssessment.objects.filter(
                position_group=employee.position_group,
                job_title=employee.job_title,
                is_active=True
            ).first()
            
            if not assessment:
                return Response({
                    'error': f'No behavioral assessment template found for {employee.job_title}',
                    'employee_info': {
                        'id': employee.id,
                        'name': employee.full_name,
                        'job_title': employee.job_title,
                        'position_group': employee.position_group.get_name_display()
                    }
                }, status=status.HTTP_404_NOT_FOUND)
            
            serializer = PositionBehavioralAssessmentSerializer(assessment)
            return Response({
                'employee_info': {
                    'id': employee.id,
                    'name': employee.full_name,
                    'job_title': employee.job_title,
                    'position_group': employee.position_group.get_name_display()
                },
                'assessment_template': serializer.data
            })
            
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, 
                          status=status.HTTP_404_NOT_FOUND)
        except ValueError:
            return Response({'error': 'Invalid employee_id format'}, 
                          status=status.HTTP_400_BAD_REQUEST)


# EMPLOYEE ASSESSMENT VIEWS

class EmployeeCoreAssessmentViewSet(viewsets.ModelViewSet):
    """Employee Core Competency Assessments - Simplified"""
    queryset = EmployeeCoreAssessment.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return EmployeeCoreAssessmentCreateSerializer
        return EmployeeCoreAssessmentSerializer
    
    def get_queryset(self):
        queryset = EmployeeCoreAssessment.objects.select_related(
            'employee', 'position_assessment'
        ).prefetch_related('competency_ratings__skill__group')
        
        # Filter by employee
        employee_id = self.request.query_params.get('employee_id')
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        # Filter by status
        assessment_status = self.request.query_params.get('status')
        if assessment_status:
            queryset = queryset.filter(status=assessment_status)
        
        return queryset.order_by('-assessment_date')
    
    def perform_create(self, serializer):
        """Always create as DRAFT"""
        assessment = serializer.save(status='DRAFT')
        return assessment
    
    def perform_update(self, serializer):
        """Keep as DRAFT unless explicitly submitting"""
        assessment = serializer.save()
        # Status only changes through specific actions
        return assessment
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit core assessment as completed (finalizes assessment)"""
        assessment = self.get_object()
        
        # Can only submit DRAFT assessments
        if assessment.status != 'DRAFT':
            return Response({
                'error': 'Only draft assessments can be submitted'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Update assessment data if provided
        if request.data:
            serializer = EmployeeCoreAssessmentCreateSerializer(
                assessment, 
                data=request.data, 
                partial=True,
                context={'request': request}
            )
            
            if serializer.is_valid():
                # Save with COMPLETED status
                updated_assessment = serializer.save()
                updated_assessment.status = 'COMPLETED'  # Explicitly set status
                updated_assessment.save()
                assessment = updated_assessment
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            # No data provided, just change status
            assessment.status = 'COMPLETED'
            assessment.save()
        
        # Calculate final scores
        assessment.calculate_scores()
        
        return Response({
            'success': True,
            'message': 'Core assessment submitted successfully',
            'assessment': EmployeeCoreAssessmentSerializer(assessment).data
        })
    
    @action(detail=True, methods=['post'])
    def reopen(self, request, pk=None):
        """Reopen completed assessment for editing"""
        assessment = self.get_object()
        
        if assessment.status != 'COMPLETED':
            return Response({
                'error': 'Only completed assessments can be reopened'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        assessment.status = 'DRAFT'
        assessment.save()
        
        return Response({
            'success': True,
            'message': 'Assessment reopened for editing',
            'assessment': EmployeeCoreAssessmentSerializer(assessment).data
        })
    
    @action(detail=True, methods=['post'])
    def recalculate_scores(self, request, pk=None):
        """Recalculate assessment scores"""
        assessment = self.get_object()
        assessment.calculate_scores()
        
        serializer = EmployeeCoreAssessmentSerializer(assessment)
        return Response({
            'success': True,
            'message': 'Scores recalculated successfully',
            'assessment': serializer.data
        })
    
    @action(detail=True, methods=['get'])
    def export_document(self, request, pk=None):
        """Export assessment as Excel document"""
        assessment = self.get_object()
        
        try:
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Core Competency Assessment"
            
            # Define styles
            header_font = Font(bold=True, size=16, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            section_font = Font(bold=True, size=12)
            section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            table_header_font = Font(bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Header information
            ws['A1'] = "CORE COMPETENCY ASSESSMENT REPORT"
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws.merge_cells('A1:F1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Employee information
            row = 3
            ws[f'A{row}'] = "Employee Information"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            employee_info = [
                ("Employee ID:", assessment.employee.employee_id),
                ("Name:", assessment.employee.full_name),
                ("Job Title:", assessment.employee.job_title),
                ("Assessment Date:", assessment.assessment_date.strftime('%Y-%m-%d')),
               
                ("Status:", assessment.get_status_display())
            ]
            
            for label, value in employee_info:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            
            # Summary scores
            row += 2
            ws[f'A{row}'] = "ASSESSMENT SUMMARY"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            summary_data = [
                ("Total Position Score:", assessment.total_position_score),
                ("Total Employee Score:", assessment.total_employee_score),
                ("Gap Score:", assessment.gap_score),
                ("Completion Percentage:", f"{assessment.completion_percentage}%")
            ]
            
            for label, value in summary_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            
            # Detailed ratings
            row += 2
            ws[f'A{row}'] = "DETAILED COMPETENCY RATINGS"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Headers
            headers = ['Skill Group', 'Skill Name', 'Required Level', 'Actual Level', 'Gap', 'Notes']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = table_header_font
                cell.fill = section_fill
                cell.border = border
            row += 1
            
            # Rating data
            for rating in assessment.competency_ratings.select_related('skill__group').all():
                ws.cell(row=row, column=1, value=rating.skill.group.name).border = border
                ws.cell(row=row, column=2, value=rating.skill.name).border = border
                ws.cell(row=row, column=3, value=rating.required_level).border = border
                ws.cell(row=row, column=4, value=rating.actual_level).border = border
                
                # Gap with color coding
                gap_cell = ws.cell(row=row, column=5, value=rating.gap)
                gap_cell.border = border
                if rating.gap > 0:
                    gap_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif rating.gap < 0:
                    gap_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                notes_cell = ws.cell(row=row, column=6, value=rating.notes or '')
                notes_cell.border = border
                row += 1
            
            # Auto-adjust column widths
            for col_num in range(1, 7):
                column_letter = get_column_letter(col_num)
                max_length = 0
                
                for row_cells in ws[f'{column_letter}1:{column_letter}{ws.max_row}']:
                    for cell in row_cells:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"core_assessment_{assessment.employee.employee_id}_{assessment.assessment_date}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            logger.error(f"Error exporting core assessment: {str(e)}")
            return Response({
                'error': f'Failed to export assessment: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EmployeeBehavioralAssessmentViewSet(viewsets.ModelViewSet):
    """Employee Behavioral Competency Assessments - Simplified"""
    queryset = EmployeeBehavioralAssessment.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return EmployeeBehavioralAssessmentCreateSerializer
        return EmployeeBehavioralAssessmentSerializer
    
    def get_queryset(self):
        queryset = EmployeeBehavioralAssessment.objects.select_related(
            'employee', 'position_assessment'
        ).prefetch_related('competency_ratings__behavioral_competency__group')
        
        employee_id = self.request.query_params.get('employee_id')
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        assessment_status = self.request.query_params.get('status')
        if assessment_status:
            queryset = queryset.filter(status=assessment_status)
        
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(assessment_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(assessment_date__lte=date_to)
        
        return queryset.order_by('-assessment_date')
    
    def perform_create(self, serializer):
        """Always create as DRAFT"""
        assessment = serializer.save(status='DRAFT')
        return assessment
    
    def perform_update(self, serializer):
        """Keep as DRAFT unless explicitly submitting"""
        assessment = serializer.save()
        # Status only changes through specific actions
        return assessment
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit behavioral assessment as completed (finalizes assessment)"""
        assessment = self.get_object()
        
        # Can only submit DRAFT assessments
        if assessment.status != 'DRAFT':
            return Response({
                'error': 'Only draft assessments can be submitted'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Update assessment data if provided
        if request.data:
            serializer = EmployeeBehavioralAssessmentCreateSerializer(
                assessment, 
                data=request.data, 
                partial=True,
                context={'request': request}
            )
            
            if serializer.is_valid():
                # Save with COMPLETED status
                updated_assessment = serializer.save()
                updated_assessment.status = 'COMPLETED'  # Explicitly set status
                updated_assessment.save()
                assessment = updated_assessment
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            # No data provided, just change status
            assessment.status = 'COMPLETED'
            assessment.save()
        
        # Calculate final scores
        assessment.calculate_scores()
        
        return Response({
            'success': True,
            'message': 'Behavioral assessment submitted successfully',
            'assessment': EmployeeBehavioralAssessmentSerializer(assessment).data
        })
    
    @action(detail=True, methods=['post'])
    def reopen(self, request, pk=None):
        """Reopen completed behavioral assessment for editing"""
        assessment = self.get_object()
        
        if assessment.status != 'COMPLETED':
            return Response({
                'error': 'Only completed assessments can be reopened'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        assessment.status = 'DRAFT'
        assessment.save()
        
        return Response({
            'success': True,
            'message': 'Behavioral assessment reopened for editing',
            'assessment': EmployeeBehavioralAssessmentSerializer(assessment).data
        })
    
    @action(detail=True, methods=['get'])
    def export_document(self, request, pk=None):
        """Export behavioral assessment as Excel document"""
        assessment = self.get_object()
        
        try:
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Behavioral Assessment"
            
            # Define styles
            header_font = Font(bold=True, size=16, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            section_font = Font(bold=True, size=12)
            section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            table_header_font = Font(bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Header information
            ws['A1'] = "BEHAVIORAL COMPETENCY ASSESSMENT REPORT"
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws.merge_cells('A1:F1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Employee information
            row = 3
            ws[f'A{row}'] = "Employee Information"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            employee_info = [
                ("Employee ID:", assessment.employee.employee_id),
                ("Name:", assessment.employee.full_name),
                ("Job Title:", assessment.employee.job_title),
                ("Department:", getattr(assessment.employee.department, 'name', 'N/A')),
                ("Assessment Date:", assessment.assessment_date.strftime('%Y-%m-%d')),
     
                ("Status:", assessment.get_status_display())
            ]
            
            for label, value in employee_info:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            
            # Overall Summary
            row += 2
            ws[f'A{row}'] = "OVERALL PERFORMANCE SUMMARY"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Overall grade with color coding
            ws[f'A{row}'] = "Overall Grade:"
            ws[f'A{row}'].font = Font(bold=True)
            
            grade_cell = ws[f'B{row}']
            grade_cell.value = f"{assessment.overall_letter_grade} ({assessment.overall_percentage}%)"
            grade_cell.font = Font(bold=True, size=14)
            
            # Color code based on grade
            if assessment.overall_letter_grade in ['A+', 'A', 'A-']:
                grade_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif assessment.overall_letter_grade in ['B+', 'B', 'B-']:
                grade_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            elif assessment.overall_letter_grade in ['C+', 'C', 'C-']:
                grade_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            else:
                grade_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            row += 2
            
            # Group Performance Summary
            ws[f'A{row}'] = "COMPETENCY GROUP PERFORMANCE"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Group headers
            group_headers = ['Competency Group', 'Required Score', 'Actual Score', 'Percentage', 'Grade']
            for col, header in enumerate(group_headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = table_header_font
                cell.fill = section_fill
                cell.border = border
            row += 1
            
            # Group scores
            for group_name, scores in assessment.group_scores.items():
                ws.cell(row=row, column=1, value=group_name).border = border
                ws.cell(row=row, column=2, value=scores['position_total']).border = border
                ws.cell(row=row, column=3, value=scores['employee_total']).border = border
                
                # Percentage with color coding
                pct_cell = ws.cell(row=row, column=4, value=f"{scores['percentage']}%")
                pct_cell.border = border
                grade_cell = ws.cell(row=row, column=5, value=scores['letter_grade'])
                grade_cell.border = border
                
                # Color code based on percentage
                if scores['percentage'] >= 90:
                    pct_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    grade_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif scores['percentage'] >= 80:
                    pct_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    grade_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                elif scores['percentage'] >= 70:
                    pct_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    grade_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                else:
                    pct_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    grade_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                row += 1
            
            # Detailed Competency Ratings
            row += 2
            ws[f'A{row}'] = "DETAILED COMPETENCY RATINGS"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Detail headers
            headers = ['Competency Group', 'Competency Name', 'Required Level', 'Actual Level', 'Performance', 'Notes']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = table_header_font
                cell.fill = section_fill
                cell.border = border
            row += 1
            
            # Rating data grouped by competency group
            ratings = assessment.competency_ratings.select_related(
                'behavioral_competency__group'
            ).order_by('behavioral_competency__group__name', 'behavioral_competency__name')
            
            current_group = None
            for rating in ratings:
                # Group separator
                if current_group != rating.behavioral_competency.group.name:
                    if current_group is not None:
                        row += 1  # Add space between groups
                    current_group = rating.behavioral_competency.group.name
                
                ws.cell(row=row, column=1, value=rating.behavioral_competency.group.name).border = border
                ws.cell(row=row, column=2, value=rating.behavioral_competency.name).border = border
                ws.cell(row=row, column=3, value=rating.required_level).border = border
                ws.cell(row=row, column=4, value=rating.actual_level).border = border
                
                # Performance indicator
                performance = "Meets" if rating.actual_level >= rating.required_level else "Below"
                performance_cell = ws.cell(row=row, column=5, value=performance)
                performance_cell.border = border
                
                if performance == "Meets":
                    performance_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    performance_cell.font = Font(color="006400")  # Dark green
                else:
                    performance_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    performance_cell.font = Font(color="8B0000")  # Dark red
                
                notes_cell = ws.cell(row=row, column=6, value=rating.notes or '')
                notes_cell.border = border
                row += 1
            
            # Development recommendations
            row += 2
            ws[f'A{row}'] = "DEVELOPMENT RECOMMENDATIONS"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Find competencies needing improvement
            improvement_areas = []
            for rating in assessment.competency_ratings.select_related('behavioral_competency').all():
                if rating.actual_level < rating.required_level:
                    improvement_areas.append({
                        'competency': rating.behavioral_competency.name,
                        'current': rating.actual_level,
                        'target': rating.required_level,
                        'gap': rating.required_level - rating.actual_level
                    })
            
            if improvement_areas:
                ws[f'A{row}'] = "Priority areas for development:"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for area in sorted(improvement_areas, key=lambda x: x['gap'], reverse=True):
                    ws[f'A{row}'] = f"• {area['competency']}: Current {area['current']} → Target {area['target']} (Gap: {area['gap']})"
                    row += 1
            else:
                ws[f'A{row}'] = "All competencies meet or exceed required levels!"
                ws[f'A{row}'].font = Font(bold=True, color="006400")
            
            # Auto-adjust column widths
            for col_num in range(1, 7):
                column_letter = get_column_letter(col_num)
                max_length = 0
                
                for row_cells in ws[f'{column_letter}1:{column_letter}{ws.max_row}']:
                    for cell in row_cells:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"behavioral_assessment_{assessment.employee.employee_id}_{assessment.assessment_date}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            logger.error(f"Error exporting behavioral assessment: {str(e)}")
            return Response({
                'error': f'Failed to export behavioral assessment: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def recalculate_scores(self, request, pk=None):
        """Recalculate behavioral assessment scores"""
        assessment = self.get_object()
        assessment.calculate_scores()
        
        serializer = EmployeeBehavioralAssessmentSerializer(assessment)
        return Response({
            'success': True,
            'message': 'Behavioral scores recalculated successfully',
            'assessment': serializer.data
        })


# SUMMARY AND REPORTING VIEWS

class AssessmentDashboardViewSet(viewsets.ViewSet):
    """Assessment Dashboard and Summary Statistics"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get assessment summary statistics"""
        # Basic counts
        total_core = EmployeeCoreAssessment.objects.count()
        total_behavioral = EmployeeBehavioralAssessment.objects.count()
        completed_core = EmployeeCoreAssessment.objects.filter(status='COMPLETED').count()
        completed_behavioral = EmployeeBehavioralAssessment.objects.filter(status='COMPLETED').count()
        
        # Recent assessments
        recent_core = EmployeeCoreAssessment.objects.select_related(
            'employee', 'position_assessment'
        ).order_by('-created_at')[:5]
        
        recent_behavioral = EmployeeBehavioralAssessment.objects.select_related(
            'employee', 'position_assessment'
        ).order_by('-created_at')[:5]
        
        # Top performers (core competency)
        top_core = EmployeeCoreAssessment.objects.filter(
            status='COMPLETED'
        ).select_related('employee').order_by('-completion_percentage')[:5]
        
        # Top performers (behavioral)
        top_behavioral = EmployeeBehavioralAssessment.objects.filter(
            status='COMPLETED'
        ).select_related('employee').order_by('-overall_percentage')[:5]
        
        # Serialize data
        core_serializer = EmployeeCoreAssessmentSerializer(recent_core, many=True)
        behavioral_serializer = EmployeeBehavioralAssessmentSerializer(recent_behavioral, many=True)
        
        top_core_data = [
            {
                'employee_name': assessment.employee.full_name,
                'employee_id': assessment.employee.employee_id,
                'completion_percentage': assessment.completion_percentage,
                'assessment_date': assessment.assessment_date
            }
            for assessment in top_core
        ]
        
        top_behavioral_data = [
            {
                'employee_name': assessment.employee.full_name,
                'employee_id': assessment.employee.employee_id,
                'overall_percentage': assessment.overall_percentage,
                'overall_letter_grade': assessment.overall_letter_grade,
                'assessment_date': assessment.assessment_date
            }
            for assessment in top_behavioral
        ]
        
        return Response({
            'summary_statistics': {
                'total_core_assessments': total_core,
                'total_behavioral_assessments': total_behavioral,
                'completed_assessments': completed_core + completed_behavioral,
                'pending_assessments': (total_core - completed_core) + (total_behavioral - completed_behavioral)
            },
            'recent_core_assessments': core_serializer.data,
            'recent_behavioral_assessments': behavioral_serializer.data,
            'top_core_performers': top_core_data,
            'top_behavioral_performers': top_behavioral_data
        })
    
    @swagger_auto_schema(
        method='get',
        operation_description='Get comprehensive assessment overview for specific employee',
        manual_parameters=[
            openapi.Parameter(
                'employee_id',
                openapi.IN_QUERY,
                description='The employee ID',
                type=openapi.TYPE_INTEGER,
                required=True,
                example=63
            )
        ],
        responses={
            200: openapi.Response(
                description='Employee comprehensive assessment overview',
                examples={
                    'application/json': {
                        'employee_info': {
                            'id': 63,
                            'employee_id': 'TRD7',
                            'name': 'John Doe',
                            'job_title': 'Software Engineer',
                            'position_group': 'Engineering',
                            'department': 'IT',
                            'business_function': 'Technology'
                        },
                        'core_assessments': [],
                        'behavioral_assessments': [],
                        'latest_core_assessment': None,
                        'latest_behavioral_assessment': {},
                        'development_areas': [],
                        'strengths': []
                    }
                }
            ),
            400: openapi.Response(
                description='Bad Request',
                examples={
                    'application/json': {
                        'error': 'employee_id is required',
                        'help': 'Add ?employee_id=63 to your request URL'
                    }
                }
            ),
            404: openapi.Response(
                description='Not Found',
                examples={
                    'application/json': {
                        'error': 'Employee not found'
                    }
                }
            )
        }
    )
    @action(detail=False, methods=['get'])
    def employee_overview(self, request):
        """Get comprehensive assessment overview for specific employee"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({
                'error': 'employee_id is required',
                'help': 'Add ?employee_id=63 to your request URL'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            employee = Employee.objects.get(id=employee_id)
            
            # Get all assessments for employee
            core_assessments = EmployeeCoreAssessment.objects.filter(
                employee=employee
            ).select_related('position_assessment').order_by('-assessment_date')
            
            behavioral_assessments = EmployeeBehavioralAssessment.objects.filter(
                employee=employee
            ).select_related('position_assessment').order_by('-assessment_date')
            
            # Get latest assessments
            latest_core = core_assessments.first()
            latest_behavioral = behavioral_assessments.first()
            
            # Development areas (skills with negative gaps)
            development_areas = []
            strengths = []
            
            if latest_core:
                for rating in latest_core.competency_ratings.select_related('skill__group').all():
                    if rating.gap < 0:
                        development_areas.append({
                            'skill_name': rating.skill.name,
                            'skill_group': rating.skill.group.name,
                            'gap': rating.gap,
                            'required_level': rating.required_level,
                            'actual_level': rating.actual_level
                        })
                    elif rating.gap > 0:
                        strengths.append({
                            'skill_name': rating.skill.name,
                            'skill_group': rating.skill.group.name,
                            'gap': rating.gap,
                            'required_level': rating.required_level,
                            'actual_level': rating.actual_level
                        })
            
            # Serialize data
            core_serializer = EmployeeCoreAssessmentSerializer(core_assessments, many=True)
            behavioral_serializer = EmployeeBehavioralAssessmentSerializer(behavioral_assessments, many=True)
            latest_core_serializer = EmployeeCoreAssessmentSerializer(latest_core) if latest_core else None
            latest_behavioral_serializer = EmployeeBehavioralAssessmentSerializer(latest_behavioral) if latest_behavioral else None
            
            return Response({
                'employee_info': {
                    'id': employee.id,
                    'employee_id': employee.employee_id,
                    'name': employee.full_name,
                    'job_title': employee.job_title,
                    'position_group': employee.position_group.get_name_display(),
                    'department': employee.department.name if employee.department else 'N/A',
                    'business_function': employee.business_function.name if employee.business_function else 'N/A'
                },
                'core_assessments': core_serializer.data,
                'behavioral_assessments': behavioral_serializer.data,
                'latest_core_assessment': latest_core_serializer.data if latest_core_serializer else None,
                'latest_behavioral_assessment': latest_behavioral_serializer.data if latest_behavioral_serializer else None,
                'development_areas': development_areas,
                'strengths': strengths
            })
            
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, 
                          status=status.HTTP_404_NOT_FOUND)
        except ValueError:
            return Response({'error': 'Invalid employee_id format'}, 
                          status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def scale_levels(self, request):
        """Get all available scale levels for assessments"""
        # Get all core competency scales
        core_scales = CoreCompetencyScale.objects.filter(is_active=True).order_by('scale')
        core_serializer = CoreCompetencyScaleSerializer(core_scales, many=True)
        
        # Get all behavioral scales
        behavioral_scales = BehavioralScale.objects.filter(is_active=True).order_by('scale')
        behavioral_serializer = BehavioralScaleSerializer(behavioral_scales, many=True)
        
        return Response({
            'core_scales': core_serializer.data,
            'behavioral_scales': behavioral_serializer.data,
            'scale_info': {
                'core_scale_count': core_scales.count(),
                'behavioral_scale_count': behavioral_scales.count()
            }
        })