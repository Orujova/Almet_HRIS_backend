# api/performance_views.py - COMPLETE FIXED VERSION

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.utils import timezone
from django.http import HttpResponse
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .performance_models import *
from .performance_serializers import *
from .models import Employee



from .performance_permissions import (
    has_performance_permission,
    check_performance_permission,
    can_view_performance,
    can_edit_performance,
    filter_viewable_performances,
    is_admin_user
)

logger = logging.getLogger(__name__)


class PerformanceYearViewSet(viewsets.ModelViewSet):
    """Performance Year Configuration"""
    queryset = PerformanceYear.objects.all()
    serializer_class = PerformanceYearSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def active_year(self, request):
        """Get active performance year"""
        active_year = PerformanceYear.objects.filter(is_active=True).first()
        if not active_year:
            return Response({
                'error': 'No active performance year configured'
            }, status=status.HTTP_404_NOT_FOUND)
        
        serializer = PerformanceYearSerializer(active_year)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    @has_performance_permission('performance.settings.manage_years')
    def set_active(self, request, pk=None):
        """Set year as active"""
        year = self.get_object()
        year.is_active = True
        year.save()
        
        return Response({
            'success': True,
            'message': f'Year {year.year} is now active'
        })


class PerformanceWeightConfigViewSet(viewsets.ModelViewSet):
    """Performance Weight Configuration"""
    queryset = PerformanceWeightConfig.objects.all()
    serializer_class = PerformanceWeightConfigSerializer
    permission_classes = [IsAuthenticated]


class GoalLimitConfigViewSet(viewsets.ModelViewSet):
    """Goal Limits Configuration"""
    queryset = GoalLimitConfig.objects.all()
    serializer_class = GoalLimitConfigSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def active_config(self, request):
        """Get active goal limit configuration"""
        config = GoalLimitConfig.get_active_config()
        serializer = GoalLimitConfigSerializer(config)
        return Response(serializer.data)


class DepartmentObjectiveViewSet(viewsets.ModelViewSet):
    """Department Objectives"""
    queryset = DepartmentObjective.objects.all()
    serializer_class = DepartmentObjectiveSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = DepartmentObjective.objects.select_related('department')
        
        department = self.request.query_params.get('department')
        if department:
            queryset = queryset.filter(department_id=department)
        
        return queryset.order_by('department__name', 'title')


class EvaluationScaleViewSet(viewsets.ModelViewSet):
    """Evaluation Scale Management"""
    queryset = EvaluationScale.objects.all()
    serializer_class = EvaluationScaleSerializer
    permission_classes = [IsAuthenticated]


class EvaluationTargetConfigViewSet(viewsets.ModelViewSet):
    """Evaluation Target Configuration"""
    queryset = EvaluationTargetConfig.objects.all()
    serializer_class = EvaluationTargetConfigSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def active_config(self, request):
        """Get active evaluation target configuration"""
        config = EvaluationTargetConfig.get_active_config()
        serializer = EvaluationTargetConfigSerializer(config)
        return Response(serializer.data)


class ObjectiveStatusViewSet(viewsets.ModelViewSet):
    """Objective Status Types"""
    queryset = ObjectiveStatus.objects.all()
    serializer_class = ObjectiveStatusSerializer
    permission_classes = [IsAuthenticated]


class EmployeePerformanceViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return EmployeePerformanceListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return EmployeePerformanceCreateUpdateSerializer
        return EmployeePerformanceDetailSerializer
    
    def get_queryset(self):
        """Filter queryset based on user permissions - FIXED"""
        queryset = EmployeePerformance.objects.select_related(
            'employee',
            'employee__department',
            'employee__line_manager',
            'employee__position_group',  # ✅ ADD
            'performance_year',
            'created_by'
        ).prefetch_related(
            'objectives',
            'competency_ratings',
            'development_needs',
            'comments'
        )
        
        # ✅ Filter viewable performances
        queryset = filter_viewable_performances(self.request.user, queryset)
        
        # ✅ Additional filters
        employee_id = self.request.query_params.get('employee_id')
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        year = self.request.query_params.get('year')
        if year:
            queryset = queryset.filter(performance_year__year=year)
        
        approval_status = self.request.query_params.get('status')
        if approval_status:
            queryset = queryset.filter(approval_status=approval_status)
        
        if self.request.query_params.get('my_team') == 'true':
            try:
                manager_employee = Employee.objects.get(user=self.request.user)
                queryset = queryset.filter(employee__line_manager=manager_employee)
            except Employee.DoesNotExist:
                pass
        
        return queryset.order_by('-performance_year__year', 'employee__employee_id')
    
    def list(self, request, *args, **kwargs):
        """Override list to add error handling"""
        try:
            queryset = self.filter_queryset(self.get_queryset())
            
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        except Exception as e:
            logger.error(f"❌ Error in list view: {e}")
            import traceback
            traceback.print_exc()
            return Response({
                'error': 'Error loading performances',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    def retrieve(self, request, *args, **kwargs):
        """Check access on retrieve"""
        instance = self.get_object()
        
        if not can_view_performance(request.user, instance):
            return Response({
                'error': 'İcazə yoxdur',
                'detail': 'Bu performance record-a baxmaq icazəniz yoxdur'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
    
  

    @action(detail=False, methods=['get'])
    def my_permissions(self, request):
        """✅ FIXED: Get current user's performance permissions"""
        from .performance_permissions import (
            get_user_performance_permissions,
            get_accessible_employees_for_performance,
            is_admin_user
        )
        
        permissions = get_user_performance_permissions(request.user)
        accessible_ids, can_view_all, is_manager = get_accessible_employees_for_performance(request.user)
        
        try:
            employee = Employee.objects.get(user=request.user, is_deleted=False)
            
            # ✅ Get email from user if not in employee
            email = employee.email if hasattr(employee, 'email') and employee.email else request.user.email
            
            employee_info = {
                'id': employee.id,
                'name': employee.full_name,
                'employee_id': employee.employee_id,
                'email': email,  # ✅ Always return email
                # ✅ Add line_manager info for filtering
                'line_manager_id': employee.line_manager.id if employee.line_manager else None,
                'department': employee.department.name if employee.department else None
            }
        except Employee.DoesNotExist:
            employee_info = None
        
        # ✅ Calculate accessible employee count
        if can_view_all:
            accessible_count = Employee.objects.filter(is_deleted=False).count()
        elif isinstance(accessible_ids, list):
            accessible_count = len(accessible_ids)
        else:
            accessible_count = 0
        
        return Response({
            'is_admin': is_admin_user(request.user),
            'can_view_all': can_view_all,
            'is_manager': is_manager,  # ✅ Critical for frontend filtering
            'permissions': permissions,
            'accessible_employee_count': accessible_count,
            'employee': employee_info
        })
    @action(detail=False, methods=['post'])
    @has_performance_permission('performance.initialize')
    def initialize(self, request):
        """Initialize performance record with behavioral competencies"""
        serializer = PerformanceInitializeSerializer(
            data=request.data,
            context={'request': request}
        )
        if serializer.is_valid():
            performance = serializer.save()
            detail_serializer = EmployeePerformanceDetailSerializer(performance)
            return Response(detail_serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # ============ HELPER METHODS ============
    
    def _check_edit_access(self, performance):
        """
        ✅ FIXED: Check if user can edit performance
        Allows editing when clarification is requested even if period ended
        """
        if is_admin_user(self.request.user):
            return True
        
        try:
            employee = Employee.objects.get(user=self.request.user)
            
            # ✅ FIX #1: Check if clarification requested - allow edit
            if performance.approval_status == 'NEED_CLARIFICATION':
                # Manager can edit when clarification requested
                if performance.employee.line_manager == employee:
                    
                    return True
                
                # Employee can edit their own performance if clarification needed
                if performance.employee == employee:
                    return True
            
            # ✅ FIX #2: Employee can edit their own performance if clarification needed
            if performance.employee == employee:
                if performance.approval_status == 'NEED_CLARIFICATION':
                    return True
            
            # ✅ FIX #3: Manager can ONLY edit during MANAGER period AND before submit
            if performance.employee.line_manager == employee:
                # Check if manager period is active
                if performance.performance_year.is_goal_setting_manager_active():
                    if not performance.objectives_employee_submitted:
                        return True
                    else:
                        return Response({
                            'error': 'Objectives already submitted',
                            'message': 'Cannot edit after submission unless employee requests clarification'
                        }, status=status.HTTP_403_FORBIDDEN)
                
                # Manager period ended - only allow if clarification
                # (already handled above)
                return Response({
                    'error': 'Manager goal setting period has ended',
                    'message': f'Manager period was {performance.performance_year.goal_setting_manager_start} to {performance.performance_year.goal_setting_manager_end}',
                    'current_date': timezone.now().date(),
                    'note': 'Can only edit now if employee requests clarification'
                }, status=status.HTTP_403_FORBIDDEN)
        
        except Employee.DoesNotExist:
            pass
        
        # Default permission check
        if not can_edit_performance(self.request.user, performance):
            return Response({
                'error': 'İcazə yoxdur',
                'detail': 'Bu performance record-u düzəltmək icazəniz yoxdur'
            }, status=status.HTTP_403_FORBIDDEN)
        
        return True
    
    def _create_development_needs(self, performance):
        """Create development needs for low-rated behavioral competencies"""
        low_ratings = performance.competency_ratings.filter(
            end_year_rating__value__lte=2
        ).select_related('behavioral_competency')
        
        for rating in low_ratings:
            existing = performance.development_needs.filter(
                competency_gap=rating.behavioral_competency.name
            ).first()
            
            if not existing:
                DevelopmentNeed.objects.create(
                    performance=performance,
                    competency_gap=rating.behavioral_competency.name,
                    development_activity='',
                    progress=0
                )
    
 
   
    @action(detail=True, methods=['post'])
    def approve_objectives_employee(self, request, pk=None):
        """
        ✅ FIXED: Employee approval is FINAL - no manager approval needed
        """
        performance = self.get_object()
        
        # Admin or employee check
        if not is_admin_user(request.user):
            try:
                employee = Employee.objects.get(user=request.user)
                if performance.employee != employee:
                    return Response({
                        'error': 'You can only approve your own objectives'
                    }, status=status.HTTP_403_FORBIDDEN)
            except Employee.DoesNotExist:
                return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Check if EMPLOYEE period is active
        if not performance.performance_year.is_goal_setting_employee_active():
            return Response({
                'error': 'Employee review period is not active',
                'message': f'Employee can only review objectives between {performance.performance_year.goal_setting_employee_start} and {performance.performance_year.goal_setting_employee_end}',
                'current_date': timezone.now().date(),
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if already approved
        if performance.objectives_employee_approved:
            return Response({
                'error': 'Objectives already approved by employee'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if manager submitted first
        if not performance.objectives_employee_submitted:
            return Response({
                'error': 'Manager must submit objectives first',
                'message': 'Waiting for manager to submit objectives'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ FIX: Employee approval is FINAL
        performance.objectives_employee_approved = True
        performance.objectives_employee_approved_date = timezone.now()
        performance.objectives_manager_approved = True  # ✅ Auto-approve for manager
        performance.objectives_manager_approved_date = timezone.now()
        performance.approval_status = 'APPROVED'  # ✅ Status is APPROVED immediately
        performance.save()
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='OBJECTIVES_APPROVED_EMPLOYEE',
            description=f'Employee approved objectives - Goal setting complete',
            performed_by=request.user
        )
        
        return Response({
            'success': True, 
            'message': 'Objectives approved successfully!',
            'next_step': 'Goal setting complete. Wait for mid-year review period.',  # ✅ Fixed
            'approval_status': 'APPROVED'
        })
    
    @action(detail=True, methods=['post'])
    def request_clarification(self, request, pk=None):
        """Employee can request clarification from manager"""
        performance = self.get_object()
        
        if not is_admin_user(request.user):
            try:
                employee = Employee.objects.get(user=request.user)
                if performance.employee != employee:
                    return Response({
                        'error': 'Only employee can request clarification'
                    }, status=status.HTTP_403_FORBIDDEN)
            except Employee.DoesNotExist:
                return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # ✅ FIX: Check if EMPLOYEE period is active
        if not performance.performance_year.is_goal_setting_employee_active():
            return Response({
                'error': 'Employee review period is not active',
                'message': f'Can only request clarification between {performance.performance_year.goal_setting_employee_start} and {performance.performance_year.goal_setting_employee_end}',
                'current_date': timezone.now().date()
            }, status=status.HTTP_400_BAD_REQUEST)
        
        comment_text = request.data.get('comment')
        comment_type = request.data.get('comment_type', 'OBJECTIVE_CLARIFICATION')
        section = request.data.get('section', 'objectives')
        
        if not comment_text:
            return Response({
                'error': 'Comment text required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        comment = PerformanceComment.objects.create(
            performance=performance,
            comment_type=comment_type,
            content=comment_text,
            created_by=request.user
        )
        
        # Reset approval flags
        if section == 'objectives':
            performance.objectives_employee_approved = False
            performance.objectives_manager_approved = False
            performance.objectives_employee_approved_date = None
            performance.objectives_manager_approved_date = None
        elif section == 'end_year':
            performance.final_employee_approved = False
            performance.final_manager_approved = False
            performance.final_employee_approval_date = None
            performance.final_manager_approval_date = None
        
        performance.approval_status = 'NEED_CLARIFICATION'
        performance.save()
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='CLARIFICATION_REQUESTED',
            description=f'Employee requested clarification for {section}',
            performed_by=request.user,
            metadata={
                'comment_id': comment.id,
                'section': section,
                'comment_type': comment_type
            }
        )
        
        return Response({
            'success': True,
            'message': 'Clarification requested - Manager can now edit and resubmit',
            'comment': PerformanceCommentSerializer(comment).data,
            'section': section
        })

    
    @action(detail=True, methods=['post'])
    def cancel_objective(self, request, pk=None):
        """Manager cancels an objective (mid-year only)"""
        performance = self.get_object()
        
        try:
            manager_employee = Employee.objects.get(user=request.user)
            if performance.employee.line_manager != manager_employee:
                has_manage = check_performance_permission(request.user, 'performance.manage_team')[0]
                if not (has_manage or is_admin_user(request.user)):
                    return Response({
                        'error': 'Only manager can cancel objectives'
                    }, status=status.HTTP_403_FORBIDDEN)
        except Employee.DoesNotExist:
            return Response({
                'error': 'Employee profile not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        objective_id = request.data.get('objective_id')
        reason = request.data.get('reason', '')
        
        if not objective_id:
            return Response({
                'error': 'objective_id required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        current_period = performance.performance_year.get_current_period()
        if current_period != 'MID_YEAR_REVIEW':
            return Response({
                'error': 'Objectives can only be cancelled during mid-year review'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            objective = performance.objectives.get(id=objective_id)
            objective.is_cancelled = True
            objective.cancelled_date = timezone.now()
            objective.cancellation_reason = reason
            objective.save()
            
            PerformanceActivityLog.objects.create(
                performance=performance,
                action='OBJECTIVE_CANCELLED',
                description=f'Objective cancelled: {objective.title}',
                performed_by=request.user,
                metadata={'objective_id': str(objective_id), 'reason': reason}
            )
            
            return Response({'success': True, 'message': 'Objective cancelled'})
        except EmployeeObjective.DoesNotExist:
            return Response({
                'error': 'Objective not found'
            }, status=status.HTTP_404_NOT_FOUND)
    

    # ============ MID-YEAR REVIEW SECTION ============

    @action(detail=True, methods=['post'])
    def submit_mid_year_employee(self, request, pk=None):
        """
        ✅ FIXED: Allow adding multiple comments during mid-year period
        """
        performance = self.get_object()
        
        is_own_performance = False
        try:
            employee = Employee.objects.get(user=request.user, is_deleted=False)
            is_own_performance = (performance.employee == employee)
        except Employee.DoesNotExist:
            pass
        
        if not is_admin_user(request.user):
            if not is_own_performance:
                return Response({
                    'error': 'You can only submit your own mid-year review'
                }, status=status.HTTP_403_FORBIDDEN)
            
            has_permission = check_performance_permission(request.user, 'performance.midyear.submit_employee')[0] or \
                            check_performance_permission(request.user, 'performance.edit_own')[0]
            
            if not has_permission:
                return Response({
                    'error': 'İcazə yoxdur',
                    'detail': 'Mid-year self-review submit etmək üçün icazə lazımdır'
                }, status=status.HTTP_403_FORBIDDEN)
        
        if not performance.performance_year.is_mid_year_active():
            return Response({
                'error': 'Mid-year review period is not active',
                'current_period': performance.performance_year.get_current_period()
            }, status=status.HTTP_400_BAD_REQUEST)
        
        comment = request.data.get('comment', '')
        if not comment.strip():
            return Response({
                'error': 'Comment is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ FIX: Append new comment to existing ones
        if performance.mid_year_employee_comment:
            # Add separator and new comment
            performance.mid_year_employee_comment += f"\n\n--- Added on {timezone.now().strftime('%Y-%m-%d %H:%M')} ---\n{comment}"
        else:
            performance.mid_year_employee_comment = comment
        
        objectives_data = request.data.get('objectives', [])
        if objectives_data:
            for obj_data in objectives_data:
                obj_id = obj_data.get('id')
                if obj_id:
                    EmployeeObjective.objects.filter(
                        id=obj_id,
                        performance=performance
                    ).update(
                        status_id=obj_data.get('status')
                    )
        
        # ✅ FIX: Update timestamp but DON'T check if already submitted
        performance.mid_year_employee_submitted = timezone.now()
        performance.save()
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='MID_YEAR_EMPLOYEE_COMMENT_ADDED',
            description='Employee added mid-year self-review comment',
            performed_by=request.user
        )
        
        return Response({
            'success': True, 
            'message': 'Comment added successfully',
            'next_step': 'Waiting for manager to complete mid-year assessment'
        })
    
    
    @action(detail=True, methods=['post'])
    def submit_mid_year_manager(self, request, pk=None):
        """
        ✅ FIXED: Allow adding multiple assessments during mid-year period
        """
        performance = self.get_object()
        
        is_line_manager = False
        try:
            manager_employee = Employee.objects.get(user=request.user, is_deleted=False)
            is_line_manager = (performance.employee.line_manager == manager_employee)
        except Employee.DoesNotExist:
            pass
        
        if not is_admin_user(request.user):
            if not is_line_manager:
                return Response({
                    'error': 'Only the line manager can complete mid-year review'
                }, status=status.HTTP_403_FORBIDDEN)
            
            has_permission = check_performance_permission(request.user, 'performance.midyear.submit_manager')[0] or \
                            check_performance_permission(request.user, 'performance.manage_team')[0] or \
                            check_performance_permission(request.user, 'performance.manage_all')[0]
            
            if not has_permission:
                return Response({
                    'error': 'İcazə yoxdur',
                    'detail': 'Mid-year assessment complete etmək üçün icazə lazımdır'
                }, status=status.HTTP_403_FORBIDDEN)
        
        if not performance.performance_year.is_mid_year_active():
            return Response({
                'error': 'Mid-year review period is not active',
                'current_period': performance.performance_year.get_current_period()
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not performance.mid_year_employee_submitted:
            return Response({
                'error': 'Employee must submit self-review first',
                'message': 'Waiting for employee to submit mid-year self-review'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        comment = request.data.get('comment', '')
        if not comment.strip():
            return Response({
                'error': 'Comment is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ FIX: Append new comment to existing ones
        if performance.mid_year_manager_comment:
            performance.mid_year_manager_comment += f"\n\n--- Added on {timezone.now().strftime('%Y-%m-%d %H:%M')} ---\n{comment}"
        else:
            performance.mid_year_manager_comment = comment
        
        objectives_data = request.data.get('objectives', [])
        if objectives_data:
            for obj_data in objectives_data:
                obj_id = obj_data.get('id')
                if obj_id:
                    EmployeeObjective.objects.filter(
                        id=obj_id,
                        performance=performance
                    ).update(
                        status_id=obj_data.get('status')
                    )
        
        # ✅ FIX: Update timestamp
        performance.mid_year_manager_submitted = timezone.now()
        performance.mid_year_completed = True
        
        if performance.objectives_manager_approved:
            performance.approval_status = 'APPROVED'
        
        performance.save()
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='MID_YEAR_MANAGER_COMMENT_ADDED',
            description='Manager added mid-year assessment comment',
            performed_by=request.user
        )
        
        return Response({
            'success': True, 
            'message': 'Assessment added successfully',
            'next_step': 'Wait for end-year review period'
        })
    
        # ============ END-YEAR REVIEW SECTION ============

 # ============ End-Year REVIEW SECTION ============
    
    @action(detail=True, methods=['post'])
    def submit_end_year_employee(self, request, pk=None):
        """
        ✅ UPDATED: Allow multiple end-year employee comments
        """
        performance = self.get_object()
        comment = request.data.get('comment', '')
        
        try:
            employee = Employee.objects.get(user=request.user)
            if performance.employee != employee and not is_admin_user(request.user):
                return Response({
                    'error': 'You can only submit your own end-year review'
                }, status=status.HTTP_403_FORBIDDEN)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if not performance.performance_year.is_end_year_active():
            return Response({
                'error': 'End-year review period is not active',
                'current_period': performance.performance_year.get_current_period()
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not comment.strip():
            return Response({
                'error': 'Comment is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ Append new comment to existing ones
        if performance.end_year_employee_comment:
            performance.end_year_employee_comment += f"\n\n--- Added on {timezone.now().strftime('%Y-%m-%d %H:%M')} ---\n{comment}"
        else:
            performance.end_year_employee_comment = comment
        
        performance.end_year_employee_submitted = timezone.now()
        performance.save()
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='END_YEAR_EMPLOYEE_COMMENT_ADDED',
            description='Employee added end-year self-review comment',
            performed_by=request.user
        )
        
        return Response({
            'success': True, 
            'message': 'End-year comment added successfully'
        })
    
    
    @action(detail=True, methods=['post'])
    def submit_end_year_manager(self, request, pk=None):
        """
        ✅ NEW: Manager adds end-year assessment comment
        """
        performance = self.get_object()
        
        is_line_manager = False
        try:
            manager_employee = Employee.objects.get(user=request.user, is_deleted=False)
            is_line_manager = (performance.employee.line_manager == manager_employee)
        except Employee.DoesNotExist:
            pass
        
        if not is_admin_user(request.user):
            if not is_line_manager:
                return Response({
                    'error': 'Only the line manager can submit end-year assessment'
                }, status=status.HTTP_403_FORBIDDEN)
            
            has_permission = check_performance_permission(request.user, 'performance.endyear.submit_manager')[0] or \
                            check_performance_permission(request.user, 'performance.manage_team')[0] or \
                            check_performance_permission(request.user, 'performance.manage_all')[0]
            
            if not has_permission:
                return Response({
                    'error': 'İcazə yoxdur',
                    'detail': 'End-year assessment submit etmək üçün icazə lazımdır'
                }, status=status.HTTP_403_FORBIDDEN)
        
        if not performance.performance_year.is_end_year_active():
            return Response({
                'error': 'End-year review period is not active',
                'current_period': performance.performance_year.get_current_period()
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not performance.end_year_employee_submitted:
            return Response({
                'error': 'Employee must submit end-year self-review first',
                'message': 'Waiting for employee to submit end-year self-review'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        comment = request.data.get('comment', '')
        if not comment.strip():
            return Response({
                'error': 'Comment is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ Append new comment to existing ones
        if performance.end_year_manager_comment:
            performance.end_year_manager_comment += f"\n\n--- Added on {timezone.now().strftime('%Y-%m-%d %H:%M')} ---\n{comment}"
        else:
            performance.end_year_manager_comment = comment
        
        performance.end_year_manager_submitted = timezone.now()
        performance.save()
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='END_YEAR_MANAGER_COMMENT_ADDED',
            description='Manager added end-year assessment comment',
            performed_by=request.user
        )
        
        return Response({
            'success': True, 
            'message': 'End-year assessment added successfully'
        })
    
    
    
    @action(detail=True, methods=['post'])
    def complete_end_year(self, request, pk=None):
        """
        ✅ FIXED: STEP 1 - Manager completes end-year review
        NO objectives approval check - only ratings check
        """
        performance = self.get_object()
        self._check_edit_access(performance)
        
        # ✅ ONLY check if ratings exist (NO approval check)
        objectives_without_rating = performance.objectives.filter(
            is_cancelled=False,
            end_year_rating__isnull=True
        ).count()
        
        if objectives_without_rating > 0:
            return Response({
                'error': f'{objectives_without_rating} objectives missing end-year ratings'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        competencies_without_rating = performance.competency_ratings.filter(
            end_year_rating__isnull=True
        ).count()
        
        if competencies_without_rating > 0:
            return Response({
                'error': f'{competencies_without_rating} behavioral competencies missing end-year ratings'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        comment = request.data.get('comment', '')
        if comment:
            performance.end_year_manager_comment = comment
        
        # Calculate final scores
        performance.calculate_scores()
        
        # Auto-create development needs
        self._create_development_needs(performance)
        
        # Update status
        performance.end_year_manager_submitted = timezone.now()
        performance.end_year_completed = True
        performance.approval_status = 'PENDING_EMPLOYEE_APPROVAL'
        performance.save()
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='END_YEAR_COMPLETED',
            description='Manager completed end-year review and scores calculated',
            performed_by=request.user,
            metadata={
                'final_rating': performance.final_rating,
                'overall_percentage': str(performance.overall_weighted_percentage),
                'competencies_letter_grade': performance.competencies_letter_grade,
                'group_scores': performance.group_competency_scores
            }
        )
        
        return Response({
            'success': True,
            'message': 'End-year review completed - Waiting for employee final approval',
            'next_step': 'Employee needs to review and approve final results',
            'final_scores': {
                'objectives_score': str(performance.total_objectives_score),
                'objectives_percentage': str(performance.objectives_percentage),
                'competencies_required_score': performance.total_competencies_required_score,
                'competencies_actual_score': performance.total_competencies_actual_score,
                'competencies_percentage': str(performance.competencies_percentage),
                'competencies_letter_grade': performance.competencies_letter_grade,
                'group_scores': performance.group_competency_scores,
                'overall_percentage': str(performance.overall_weighted_percentage),
                'final_rating': performance.final_rating
            }
        })
    
    @action(detail=True, methods=['post'])
    def approve_final_employee(self, request, pk=None):
        """STEP 2: Employee approves final performance results"""
        performance = self.get_object()
        
        if not is_admin_user(request.user):
            try:
                employee = Employee.objects.get(user=request.user)
                if performance.employee != employee:
                    return Response({
                        'error': 'You can only approve your own performance'
                    }, status=status.HTTP_403_FORBIDDEN)
            except Employee.DoesNotExist:
                return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if not performance.end_year_completed:
            return Response({
                'error': 'End-year review not completed yet',
                'message': 'Manager must complete end-year review first'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if performance.final_employee_approved:
            return Response({
                'error': 'Already approved by employee'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        performance.final_employee_approved = True
        performance.final_employee_approval_date = timezone.now()
        performance.approval_status = 'PENDING_MANAGER_APPROVAL'
        performance.save()
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='FINAL_APPROVED_EMPLOYEE',
            description='Employee approved final performance',
            performed_by=request.user
        )
        
        return Response({
            'success': True, 
            'message': 'Final performance approved by employee',
            'next_step': 'Waiting for final manager approval',
            'approval_status': performance.approval_status
        })
    
    @action(detail=True, methods=['post'])
    def approve_final_manager(self, request, pk=None):
        """STEP 3: Manager final approval (publishes performance)"""
        performance = self.get_object()
        
        if not is_admin_user(request.user):
            try:
                manager_employee = Employee.objects.get(user=request.user)
                if performance.employee.line_manager != manager_employee:
                    has_manage = check_performance_permission(request.user, 'performance.manage_team')[0] or \
                                check_performance_permission(request.user, 'performance.manage_all')[0]
                    if not has_manage:
                        return Response({
                            'error': 'You are not authorized to approve this performance'
                        }, status=status.HTTP_403_FORBIDDEN)
            except Employee.DoesNotExist:
                return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if not performance.end_year_completed:
            return Response({
                'error': 'End-year review not completed yet'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not performance.final_employee_approved:
            return Response({
                'error': 'Employee must approve final performance first'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if performance.final_manager_approved:
            return Response({
                'error': 'Already approved by manager'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        performance.final_manager_approved = True
        performance.final_manager_approval_date = timezone.now()
        performance.approval_status = 'COMPLETED'
        performance.save()
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='FINAL_APPROVED_MANAGER',
            description='Manager approved and published final performance',
            performed_by=request.user
        )
        
        return Response({
            'success': True,
            'message': 'Final performance approved and published',
            'approval_status': 'COMPLETED'
        })
    

    # ============ DEVELOPMENT NEEDS SECTION ============
    
    @action(detail=True, methods=['post'])
    def save_development_needs_draft(self, request, pk=None):
        """Save development needs as draft"""
        performance = self.get_object()
        self._check_edit_access(performance)
        
        needs_data = request.data.get('development_needs', [])
        
        with transaction.atomic():
            updated_ids = []
            for need_data in needs_data:
                need_id = need_data.get('id')
                if need_id:
                    DevelopmentNeed.objects.filter(
                        id=need_id,
                        performance=performance
                    ).update(
                        development_activity=need_data.get('development_activity', ''),
                        progress=need_data.get('progress', 0),
                        comment=need_data.get('comment', '')
                    )
                    updated_ids.append(need_id)
                else:
                    new_need = DevelopmentNeed.objects.create(
                        performance=performance,
                        competency_gap=need_data.get('competency_gap', ''),
                        development_activity=need_data.get('development_activity', ''),
                        progress=need_data.get('progress', 0),
                        comment=need_data.get('comment', '')
                    )
                    updated_ids.append(new_need.id)
            
            performance.development_needs_draft_saved = timezone.now()
            performance.save()
            
            PerformanceActivityLog.objects.create(
                performance=performance,
                action='DEVELOPMENT_NEEDS_DRAFT_SAVED',
                description='Development needs saved as draft',
                performed_by=request.user
            )
        
        return Response({
            'success': True,
            'message': 'Development needs draft saved'
        })
    

    # ============ UTILITIES ============
    
    @action(detail=True, methods=['post'])
    @has_performance_permission('performance.recalculate_scores')
    def recalculate_scores(self, request, pk=None):
        """Recalculate performance scores"""
        performance = self.get_object()
        performance.calculate_scores()
        
        return Response({
            'success': True,
            'message': 'Scores recalculated',
            'scores': {
                'objectives_score': str(performance.total_objectives_score),
                'objectives_percentage': str(performance.objectives_percentage),
                'competencies_required_score': performance.total_competencies_required_score,
                'competencies_actual_score': performance.total_competencies_actual_score,
                'competencies_percentage': str(performance.competencies_percentage),
                'competencies_letter_grade': performance.competencies_letter_grade,
                'group_scores': performance.group_competency_scores,
                'overall_percentage': str(performance.overall_weighted_percentage),
                'final_rating': performance.final_rating
            }
        })
    # ==================== OBJECTIVES ENDPOINTS ====================

   
    @action(detail=True, methods=['post'])
    def save_objectives_draft(self, request, pk=None):
        """
        ✅ FIXED: Save objectives draft with proper end_year_rating handling
        """
        performance = self.get_object()
        
        # ✅ Check edit access
        access_check = self._check_edit_access(performance)
        if access_check is not True:
            return access_check
        
        objectives_data = request.data.get('objectives', [])
        
        
        
        with transaction.atomic():
            updated_ids = []
            for idx, obj_data in enumerate(objectives_data):
                obj_id = obj_data.get('id')
                
                # ✅ CRITICAL: Extract all fields including end_year_rating
                title = obj_data.get('title', '')
                description = obj_data.get('description', '')
                linked_dept_obj_id = obj_data.get('linked_department_objective')
                weight = obj_data.get('weight', 0)
              
                status_id = obj_data.get('status')
                end_year_rating_id = obj_data.get('end_year_rating')  # ✅ GET THIS
                calculated_score = obj_data.get('calculated_score', 0)
                
                if obj_id:
                    # ✅ UPDATE existing objective
                    obj = EmployeeObjective.objects.filter(
                        id=obj_id, 
                        performance=performance
                    ).first()
                    
                    if obj:
                        obj.title = title
                        obj.description = description
                        obj.linked_department_objective_id = linked_dept_obj_id
                        obj.weight = weight
                
                        obj.status_id = status_id
                        obj.display_order = idx
                        obj.calculated_score = calculated_score
                        
                        # ✅ CRITICAL FIX: Set end_year_rating
                        if end_year_rating_id is not None:
                            obj.end_year_rating_id = end_year_rating_id
                        else:
                            obj.end_year_rating = None  # Allow clearing
                        
                        obj.save()
                        updated_ids.append(obj.id)
                        
                     
                else:
                    # ✅ CREATE new objective
                    new_obj = EmployeeObjective.objects.create(
                        performance=performance,
                        title=title,
                        description=description,
                        linked_department_objective_id=linked_dept_obj_id,
                        weight=weight,
                    
                        status_id=status_id,
                        end_year_rating_id=end_year_rating_id,  # ✅ SET THIS
                        calculated_score=calculated_score,
                        display_order=idx
                    )
                    updated_ids.append(new_obj.id)
                  
            # Delete objectives not in the list
            deleted_count = performance.objectives.exclude(id__in=updated_ids).count()
            performance.objectives.exclude(id__in=updated_ids).delete()
            
            if deleted_count > 0:
                logger.info(f"🗑️ Deleted {deleted_count} objectives")
            
            # Update draft timestamp
            performance.objectives_draft_saved_date = timezone.now()
            performance.save()
            
            # ✅ Recalculate scores
            performance.calculate_scores()
            
            PerformanceActivityLog.objects.create(
                performance=performance,
                action='OBJECTIVES_DRAFT_SAVED',
                description=f'Objectives saved: {len(updated_ids)} updated, {deleted_count} deleted',
                performed_by=request.user
            )
            
           
        
        return Response({
            'success': True,
            'message': 'Objectives draft saved successfully',
            'objectives_count': len(updated_ids),
            'total_objectives_score': str(performance.total_objectives_score),
            'objectives_percentage': str(performance.objectives_percentage)
        })
 

    @action(detail=True, methods=['post'])
    def submit_objectives(self, request, pk=None):
        """
        ✅ FIXED: Submit objectives with proper end_year_rating handling
        """
        performance = self.get_object()
        
        # Only manager can submit objectives
        if not is_admin_user(request.user):
            try:
                manager_employee = Employee.objects.get(user=request.user)
                if performance.employee.line_manager != manager_employee:
                    has_manage = check_performance_permission(request.user, 'performance.manage_team')[0] or \
                                check_performance_permission(request.user, 'performance.manage_all')[0]
                    if not has_manage:
                        return Response({
                            'error': 'Only manager can submit objectives'
                        }, status=status.HTTP_403_FORBIDDEN)
            except Employee.DoesNotExist:
                return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # ✅ Check period - allow if clarification requested even if period ended
        is_clarification = performance.approval_status == 'NEED_CLARIFICATION'
        
        if not is_clarification:
            if not performance.performance_year.is_goal_setting_manager_active():
                return Response({
                    'error': 'Manager goal setting period has ended',
                    'message': f'Manager can only submit objectives between {performance.performance_year.goal_setting_manager_start} and {performance.performance_year.goal_setting_manager_end}',
                    'current_date': timezone.now().date(),
                    'manager_period': {
                        'start': performance.performance_year.goal_setting_manager_start,
                        'end': performance.performance_year.goal_setting_manager_end
                    }
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            logger.info(f"✅ Allowing submit during clarification - Period ended but clarification active")
        
        # ✅ SAVE objectives data BEFORE validation (if provided)
        objectives_data = request.data.get('objectives', [])
        
        if objectives_data:
           
            
            with transaction.atomic():
                updated_ids = []
                for idx, obj_data in enumerate(objectives_data):
                    obj_id = obj_data.get('id')
                    
                    # ✅ Extract all fields including end_year_rating
                    title = obj_data.get('title', '')
                    description = obj_data.get('description', '')
                    linked_dept_obj_id = obj_data.get('linked_department_objective')
                    weight = obj_data.get('weight', 0)
                
                    status_id = obj_data.get('status')
                    end_year_rating_id = obj_data.get('end_year_rating')  # ✅ CRITICAL
                    calculated_score = obj_data.get('calculated_score', 0)
                    
                    if obj_id:
                        # ✅ UPDATE existing
                        obj = EmployeeObjective.objects.filter(
                            id=obj_id, 
                            performance=performance
                        ).first()
                        
                        if obj:
                            obj.title = title
                            obj.description = description
                            obj.linked_department_objective_id = linked_dept_obj_id
                            obj.weight = weight
                           
                            obj.status_id = status_id
                            obj.display_order = idx
                            obj.calculated_score = calculated_score
                            
                            # ✅ CRITICAL: Save end_year_rating
                            if end_year_rating_id is not None:
                                obj.end_year_rating_id = end_year_rating_id
                            else:
                                obj.end_year_rating = None
                            
                            obj.save()
                            updated_ids.append(obj.id)
                            
                          
                    else:
                        # ✅ CREATE new
                        new_obj = EmployeeObjective.objects.create(
                            performance=performance,
                            title=title,
                            description=description,
                            linked_department_objective_id=linked_dept_obj_id,
                            weight=weight,
                          
                            status_id=status_id,
                            end_year_rating_id=end_year_rating_id,  # ✅ CRITICAL
                            calculated_score=calculated_score,
                            display_order=idx
                        )
                        updated_ids.append(new_obj.id)
               
                
                # Delete objectives not in the list
                deleted_count = performance.objectives.exclude(id__in=updated_ids).count()
                performance.objectives.exclude(id__in=updated_ids).delete()
                
                if deleted_count > 0:
                    logger.info(f"🗑️ Deleted {deleted_count} objectives")
        
        # ✅ NOW validate objectives AFTER saving
        objectives = performance.objectives.filter(is_cancelled=False)
        goal_config = GoalLimitConfig.get_active_config()
        
        objectives_count = objectives.count()
        
        if objectives_count < goal_config.min_goals:
            return Response({
                'error': f'Minimum {goal_config.min_goals} objectives required (currently {objectives_count})'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if objectives_count > goal_config.max_goals:
            return Response({
                'error': f'Maximum {goal_config.max_goals} objectives allowed (currently {objectives_count})'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check total weight
        total_weight = sum([obj.weight for obj in objectives])
        if total_weight != 100:
            return Response({
                'error': f'Total objective weights must equal 100% (currently {total_weight}%)'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check that all objectives have required fields
        for obj in objectives:
            if not obj.title or not obj.title.strip():
                return Response({
                    'error': 'All objectives must have a title'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if not obj.status:
                return Response({
                    'error': 'All objectives must have a status'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if obj.weight <= 0:
                return Response({
                    'error': 'All objectives must have weight > 0'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Update status
        was_clarification_needed = performance.approval_status == 'NEED_CLARIFICATION'
        
        performance.objectives_employee_submitted = True
        performance.objectives_employee_submitted_date = timezone.now()
        performance.approval_status = 'PENDING_EMPLOYEE_APPROVAL'
        performance.save()
        
        log_msg = 'Manager resubmitted objectives after clarification' if was_clarification_needed else 'Manager submitted objectives for employee approval'
        
        PerformanceActivityLog.objects.create(
            performance=performance,
            action='OBJECTIVES_SUBMITTED',
            description=log_msg,
            performed_by=request.user,
            metadata={
                'was_clarification': was_clarification_needed,
                'objectives_count': objectives_count
            }
        )
        
        return Response({
            'success': True,
            'message': 'Objectives resubmitted successfully - Waiting for employee review' if was_clarification_needed else 'Objectives submitted for employee approval',
            'next_step': 'Waiting for employee to review and approve',
            'objectives_count': objectives_count,
            'was_clarification_response': was_clarification_needed
        })
        
    # ==================== COMPETENCIES ENDPOINTS ====================
    
    def _sync_to_behavioral_assessment(self, employee, competencies_data):
        """
        ✅ FIXED: Sync to BOTH DRAFT and COMPLETED behavioral assessments
        Convert PerformanceEvaluationScale to integer level for behavioral assessment
        """
        from .competency_assessment_models import EmployeeBehavioralAssessment
        from .performance_models import EvaluationScale
        
        try:
            # Find active behavioral assessment
            assessment = EmployeeBehavioralAssessment.objects.filter(
                employee=employee,
                status__in=['DRAFT', 'COMPLETED']
            ).order_by('-assessment_date').first()
            
            if not assessment:
            
                return {
                    'synced': False, 
                    'reason': 'no_assessment',
                    'message': 'No behavioral assessment found'
                }
            
            # Store original status
            original_status = assessment.status
            was_completed = original_status == 'COMPLETED'
            
            synced_count = 0
            updated_count = 0
            
            for comp_data in competencies_data:
                behavioral_competency_id = comp_data['behavioral_competency_id']
                actual_level_scale_id = comp_data['actual_level_id']  # ✅ This is PerformanceEvaluationScale ID
                notes = comp_data['notes']
                
                # Skip if no rating provided
                if not actual_level_scale_id:
                    continue
                
                # ✅ Convert PerformanceEvaluationScale ID to integer level
                try:
                    performance_scale = EvaluationScale.objects.get(id=actual_level_scale_id)
                    actual_level_value = int(performance_scale.value)  # Get the integer value (1-10)
                except EvaluationScale.DoesNotExist:
                  
                    continue
                except (ValueError, TypeError):
                   
                    continue
                
                # Update or create rating in behavioral assessment
                rating, created = assessment.competency_ratings.update_or_create(
                    behavioral_competency_id=behavioral_competency_id,
                    defaults={
                        'actual_level': actual_level_value,  # ✅ Use actual_level (integer)
                        'notes': notes
                    }
                )
                
                if created:
                    synced_count += 1
                else:
                    updated_count += 1
            
            # Recalculate assessment scores
            assessment.calculate_scores()
            assessment.save()
            
            log_message = (
                f"Synced from Performance to Behavioral Assessment: "
                f"{synced_count} created, {updated_count} updated"
            )
            
            if was_completed:
                log_message += " (assessment was COMPLETED - updated anyway)"
           
            
            return {
                'synced': True,
                'assessment_id': str(assessment.id),
                'assessment_status': original_status,
                'was_completed': was_completed,
                'synced_count': synced_count,
                'updated_count': updated_count,
                'message': f'Synced {synced_count + updated_count} ratings to {original_status} behavioral assessment'
            }
            
        except Exception as e:
          
            import traceback
            traceback.print_exc()
            
            return {
                'synced': False,
                'reason': 'error',
                'error': str(e),
                'message': f'Error syncing: {str(e)}'
            }
    
    def _sync_to_leadership_assessment(self, employee, competencies_data):
        """
        ✅ NEW: Sync to leadership assessment instead of behavioral
        Convert PerformanceEvaluationScale to integer level
        """
        from .competency_assessment_models import EmployeeLeadershipAssessment
        from .performance_models import EvaluationScale
        
        try:
            # Find active leadership assessment
            assessment = EmployeeLeadershipAssessment.objects.filter(
                employee=employee,
                status__in=['DRAFT', 'COMPLETED']
            ).order_by('-assessment_date').first()
            
            if not assessment:
               
                return {
                    'synced': False, 
                    'reason': 'no_assessment',
                    'message': 'No leadership assessment found'
                }
            
            # Store original status
            original_status = assessment.status
            was_completed = original_status == 'COMPLETED'
            
            synced_count = 0
            updated_count = 0
            
            for comp_data in competencies_data:
                leadership_item_id = comp_data['leadership_item_id']
                actual_level_scale_id = comp_data['actual_level_id']  # ✅ PerformanceEvaluationScale ID
                notes = comp_data['notes']
                
                # Skip if no rating provided
                if not actual_level_scale_id:
                    continue
                
                # ✅ Convert PerformanceEvaluationScale ID to integer level
                try:
                    performance_scale = EvaluationScale.objects.get(id=actual_level_scale_id)
                    actual_level_value = int(performance_scale.value)
                except EvaluationScale.DoesNotExist:
                
                    continue
                except (ValueError, TypeError):

                    continue
                
                # Update or create rating in leadership assessment
                rating, created = assessment.competency_ratings.update_or_create(
                    leadership_item_id=leadership_item_id,
                    defaults={
                        'actual_level': actual_level_value,  # ✅ Integer value
                        'notes': notes
                    }
                )
                
                if created:
                    synced_count += 1
                else:
                    updated_count += 1
            
            # Recalculate assessment scores
            assessment.calculate_scores()
            assessment.save()
            
            log_message = (
                f"Synced from Performance to Leadership Assessment: "
                f"{synced_count} created, {updated_count} updated"
            )
            
            if was_completed:
                log_message += " (assessment was COMPLETED - updated anyway)"
            
           
            
            return {
                'synced': True,
                'assessment_id': str(assessment.id),
                'assessment_status': original_status,
                'was_completed': was_completed,
                'synced_count': synced_count,
                'updated_count': updated_count,
                'message': f'Synced {synced_count + updated_count} ratings to {original_status} leadership assessment'
            }
            
        except Exception as e:
      
            import traceback
            traceback.print_exc()
            
            return {
                'synced': False,
                'reason': 'error',
                'error': str(e),
                'message': f'Error syncing: {str(e)}'
            }
    
    
    @action(detail=True, methods=['post'])
    def save_competencies_draft(self, request, pk=None):
        """
        ✅ Save competencies AND sync to appropriate assessment (behavioral OR leadership)
        """
        performance = self.get_object()
        
        # Check edit access
        access_check = self._check_edit_access(performance)
        if access_check is not True:
            return access_check
        
        competencies_data = request.data.get('competencies', [])
        
      
        
        with transaction.atomic():
            updated_count = 0
            sync_data = []
            is_leadership = False
            
            for comp_data in competencies_data:
                comp_id = comp_data.get('id')
                end_year_rating = comp_data.get('end_year_rating')
                notes = comp_data.get('notes', '')
                
                if comp_id:
                    competency = EmployeeCompetencyRating.objects.filter(
                        id=comp_id,
                        performance=performance
                    ).first()
                    
                    if competency:
                        # Save rating
                        competency.end_year_rating_id = end_year_rating
                        competency.notes = notes
                        competency.save()
                        updated_count += 1
                        
                        # ✅ Collect for sync - check type
                        if competency.leadership_item:
                            is_leadership = True
                            sync_data.append({
                                'leadership_item_id': competency.leadership_item_id,
                                'actual_level_id': end_year_rating,
                                'notes': notes
                            })
                        elif competency.behavioral_competency:
                            sync_data.append({
                                'behavioral_competency_id': competency.behavioral_competency_id,
                                'actual_level_id': end_year_rating,
                                'notes': notes
                            })
            
            # Update timestamp
            performance.competencies_draft_saved_date = timezone.now()
            performance.save()
            
            # ✅ Sync to appropriate assessment type
            if is_leadership:
                sync_result = self._sync_to_leadership_assessment(
                    performance.employee,
                    sync_data
                )
            else:
                sync_result = self._sync_to_behavioral_assessment(
                    performance.employee,
                    sync_data
                )
            
            # Recalculate scores
            performance.calculate_scores()
            
            PerformanceActivityLog.objects.create(
                performance=performance,
                action='COMPETENCIES_DRAFT_SAVED',
                description=f'Competencies saved: {updated_count} updated',
                performed_by=request.user,
                metadata={
                    'synced_to_assessment': sync_result['synced'],
                    'assessment_type': 'LEADERSHIP' if is_leadership else 'BEHAVIORAL',
                    'assessment_id': sync_result.get('assessment_id')
                }
            )
            
          
        
        return Response({
            'success': True,
            'message': 'Competencies draft saved successfully',
            'updated_count': updated_count,
            'assessment_type': 'LEADERSHIP' if is_leadership else 'BEHAVIORAL',
            'synced_to_assessment': sync_result['synced'],
            'sync_result': sync_result,
            'competencies_percentage': str(performance.competencies_percentage),
            'competencies_letter_grade': performance.competencies_letter_grade,
            'overall_weighted_percentage': str(performance.overall_weighted_percentage),
            'final_rating': performance.final_rating
        })
    
    
    @action(detail=True, methods=['post'])
    def submit_competencies(self, request, pk=None):
        """
        ✅ Submit competencies AND sync to appropriate assessment
        """
        performance = self.get_object()
        
        access_check = self._check_edit_access(performance)
        if access_check is not True:
            return access_check
        
        competencies_data = request.data.get('competencies', [])
        
        with transaction.atomic():
            sync_data = []
            is_leadership = False
            
            # ✅ Save competencies if provided
            if competencies_data:
                for comp_data in competencies_data:
                    comp_id = comp_data.get('id')
                    end_year_rating = comp_data.get('end_year_rating')
                    notes = comp_data.get('notes', '')
                    
                    if comp_id:
                        competency = EmployeeCompetencyRating.objects.filter(
                            id=comp_id,
                            performance=performance
                        ).first()
                        
                        if competency:
                            competency.end_year_rating_id = end_year_rating
                            competency.notes = notes
                            competency.save()
                            
                            # Collect for sync
                            if competency.leadership_item:
                                is_leadership = True
                                sync_data.append({
                                    'leadership_item_id': competency.leadership_item_id,
                                    'actual_level_id': end_year_rating,
                                    'notes': notes
                                })
                            elif competency.behavioral_competency:
                                sync_data.append({
                                    'behavioral_competency_id': competency.behavioral_competency_id,
                                    'actual_level_id': end_year_rating,
                                    'notes': notes
                                })
            
            # Validate all have ratings
            missing = performance.competency_ratings.filter(end_year_rating__isnull=True)
            
            if missing.exists():
                missing_count = missing.count()
                return Response({
                    'error': f'{missing_count} competencies missing ratings',
                    'message': 'Please rate all competencies before submitting'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # ✅ Sync to appropriate assessment
            if is_leadership:
                sync_result = self._sync_to_leadership_assessment(
                    performance.employee,
                    sync_data if sync_data else [
                        {
                            'leadership_item_id': c.leadership_item_id,
                            'actual_level_id': c.end_year_rating_id,
                            'notes': c.notes or ''
                        }
                        for c in performance.competency_ratings.all() if c.leadership_item
                    ]
                )
            else:
                sync_result = self._sync_to_behavioral_assessment(
                    performance.employee,
                    sync_data if sync_data else [
                        {
                            'behavioral_competency_id': c.behavioral_competency_id,
                            'actual_level_id': c.end_year_rating_id,
                            'notes': c.notes or ''
                        }
                        for c in performance.competency_ratings.all() if c.behavioral_competency
                    ]
                )
            
            # Recalculate
            performance.calculate_scores()
            
            # Mark submitted
            performance.competencies_submitted = True
            performance.competencies_submitted_date = timezone.now()
            performance.save()
            
            PerformanceActivityLog.objects.create(
                performance=performance,
                action='COMPETENCIES_SUBMITTED',
                description=f'Competencies submitted - {performance.competencies_letter_grade}',
                performed_by=request.user,
                metadata={
                    'assessment_type': 'LEADERSHIP' if is_leadership else 'BEHAVIORAL',
                    'synced_to_assessment': sync_result['synced'],
                    'assessment_id': sync_result.get('assessment_id')
                }
            )
            
           
        
        return Response({
            'success': True,
            'message': 'Competencies submitted successfully',
            'assessment_type': 'LEADERSHIP' if is_leadership else 'BEHAVIORAL',
            'synced_to_assessment': sync_result['synced'],
            'sync_result': sync_result,
            'scores': {
                'percentage': str(performance.competencies_percentage),
                'letter_grade': performance.competencies_letter_grade,
                'overall_percentage': str(performance.overall_weighted_percentage),
                'final_rating': performance.final_rating
            }
        })
    
    # ==================== CALCULATE SCORES ====================
    
    def calculate_scores(self):
        """
        ✅ Calculate all scores
        """
        from collections import defaultdict
        import logging
        
        logger = logging.getLogger(__name__)
        
       
        
        eval_target = EvaluationTargetConfig.get_active_config()
        weight_config = PerformanceWeightConfig.objects.filter(
            position_group=self.employee.position_group
        ).first()
        
        if not weight_config:
            logger.warning(f"❌ No weight config for position group")
            return
        
        # ========== OBJECTIVES ==========
        objectives = self.objectives.filter(is_cancelled=False)
        obj_score = 0
        
        for obj in objectives:
            if obj.calculated_score and obj.calculated_score > 0:
                obj_score += float(obj.calculated_score)
        
        self.total_objectives_score = round(obj_score, 2)
        self.objectives_percentage = round(
            (self.total_objectives_score / eval_target.objective_score_target) * 100, 2
        ) if eval_target.objective_score_target > 0 else 0
        
    
        
        # ========== COMPETENCIES ==========
        from .competency_assessment_models import LetterGradeMapping
        
        competencies = self.competency_ratings.select_related(
            'behavioral_competency__group',
            'end_year_rating'
        ).all()
        
        group_data = defaultdict(lambda: {'required_total': 0, 'actual_total': 0, 'count': 0})
        total_required = 0
        total_actual = 0
        
        for comp in competencies:
            if not comp.behavioral_competency:
                continue
                
            group_name = comp.behavioral_competency.group.name
            required = comp.required_level or 0
            actual = comp.end_year_rating.value if comp.end_year_rating else 0
            
            group_data[group_name]['required_total'] += required
            group_data[group_name]['actual_total'] += actual
            group_data[group_name]['count'] += 1
            
            total_required += required
            total_actual += actual
        
        # Group scores
        group_scores = {}
        for group_name, data in group_data.items():
            percentage = (data['actual_total'] / data['required_total'] * 100) if data['required_total'] > 0 else 0
            letter_grade = LetterGradeMapping.get_letter_grade(percentage)
            
            group_scores[group_name] = {
                'required_total': data['required_total'],
                'actual_total': data['actual_total'],
                'percentage': round(percentage, 2),
                'letter_grade': letter_grade,
                'count': data['count']
            }
        
        self.group_competency_scores = group_scores
        self.total_competencies_required_score = total_required
        self.total_competencies_actual_score = total_actual
        self.competencies_percentage = round((total_actual / total_required * 100), 2) if total_required > 0 else 0
        self.competencies_letter_grade = LetterGradeMapping.get_letter_grade(self.competencies_percentage)
        
     
        
        # ========== OVERALL ==========
        self.overall_weighted_percentage = round(
            (self.objectives_percentage * weight_config.objectives_weight / 100) +
            (self.competencies_percentage * weight_config.competencies_weight / 100),
            2
        )
        
        rating = EvaluationScale.get_rating_by_percentage(self.overall_weighted_percentage)
        self.final_rating = rating.name if rating else 'N/A'
        
      
        
        self.save()
    @action(detail=True, methods=['get'])
    def activity_log(self, request, pk=None):
        """Get performance activity log"""
        performance = self.get_object()
        
        if not can_view_performance(request.user, performance):
            return Response({
                'error': 'İcazə yoxdur'
            }, status=status.HTTP_403_FORBIDDEN)
        
        logs = performance.activity_logs.all()
        serializer = PerformanceActivityLogSerializer(logs, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def competency_breakdown(self, request, pk=None):
        """Get detailed competency breakdown with gaps"""
        performance = self.get_object()
        
        if not can_view_performance(request.user, performance):
            return Response({
                'error': 'İcazə yoxdur'
            }, status=status.HTTP_403_FORBIDDEN)
        
        overall = {
            'required': performance.total_competencies_required_score,
            'actual': performance.total_competencies_actual_score,
            'percentage': float(performance.competencies_percentage),
            'letter_grade': performance.competencies_letter_grade
        }
        
        by_group = performance.group_competency_scores
        
        gaps = []
        for comp_rating in performance.competency_ratings.select_related('behavioral_competency').all():
            if comp_rating.gap < 0:
                gaps.append({
                    'competency': comp_rating.behavioral_competency.name,
                    'group': comp_rating.behavioral_competency.group.name,
                    'gap': comp_rating.gap,
                    'required': comp_rating.required_level,
                    'actual': comp_rating.actual_value,
                    'rating': comp_rating.end_year_rating.name if comp_rating.end_year_rating else 'N/A'
                })
        
        strengths = []
        for comp_rating in performance.competency_ratings.select_related('behavioral_competency').all():
            if comp_rating.gap > 0:
                strengths.append({
                    'competency': comp_rating.behavioral_competency.name,
                    'group': comp_rating.behavioral_competency.group.name,
                    'gap': comp_rating.gap,
                    'required': comp_rating.required_level,
                    'actual': comp_rating.actual_value,
                    'rating': comp_rating.end_year_rating.name if comp_rating.end_year_rating else 'N/A'
                })
        
        return Response({
            'overall': overall,
            'by_group': by_group,
            'gaps': sorted(gaps, key=lambda x: x['gap']),
            'strengths': sorted(strengths, key=lambda x: x['gap'], reverse=True)
        })
    
    @action(detail=True, methods=['get'])
    @has_performance_permission('performance.export')
    def export_excel(self, request, pk=None):
        """Export performance to Excel"""
        performance = self.get_object()
        
        if not can_view_performance(request.user, performance):
            return Response({
                'error': 'İcazə yoxdur'
            }, status=status.HTTP_403_FORBIDDEN)
        
        wb = openpyxl.Workbook()
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        
        ws_summary['A1'] = 'PERFORMANCE REVIEW SUMMARY'
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        row = 3
        info_data = [
            ['Employee Name:', performance.employee.full_name],
            ['Employee ID:', performance.employee.employee_id],
            ['Department:', performance.employee.department.name if performance.employee.department else 'N/A'],
            # ✅ FIX: Convert position_group to string
            ['Position:', str(performance.employee.position_group) if performance.employee.position_group else 'N/A'],
            ['Manager:', performance.employee.line_manager.full_name if performance.employee.line_manager else 'N/A'],
            ['Performance Year:', str(performance.performance_year.year)],
            ['Status:', performance.get_approval_status_display()],
        ]
        
        for label, value in info_data:
            ws_summary[f'A{row}'] = label
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = value
            row += 1
        
        row += 2
        ws_summary[f'A{row}'] = 'PERFORMANCE SCORES'
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        scores_data = [
            ['Objectives Score:', f"{performance.total_objectives_score}"],
            ['Objectives Percentage:', f"{performance.objectives_percentage}%"],
            ['Competencies Score:', f"{performance.total_competencies_actual_score} / {performance.total_competencies_required_score}"],
            ['Competencies Percentage:', f"{performance.competencies_percentage}%"],
            ['Competencies Letter Grade:', performance.competencies_letter_grade or 'N/A'],
            ['Overall Percentage:', f"{performance.overall_weighted_percentage}%"],
            ['Final Rating:', performance.final_rating or 'N/A'],
        ]
        
        for label, value in scores_data:
            ws_summary[f'A{row}'] = label
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = value
            row += 1
        
        # ========== OBJECTIVES SHEET ==========
        ws_obj = wb.create_sheet('Objectives')
        headers = ['#', 'Title', 'Description', 'Weight %',  'Status', 'End-Year Rating', 'Score']
        for col, header in enumerate(headers, 1):
            cell = ws_obj.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        objectives = performance.objectives.filter(is_cancelled=False).order_by('display_order')
        for idx, obj in enumerate(objectives, 1):
            ws_obj.cell(row=idx+1, column=1, value=idx)
            ws_obj.cell(row=idx+1, column=2, value=obj.title)
            ws_obj.cell(row=idx+1, column=3, value=obj.description)
            ws_obj.cell(row=idx+1, column=4, value=obj.weight)
      
            ws_obj.cell(row=idx+1, column=6, value=obj.status.label if obj.status else 'N/A')
            ws_obj.cell(row=idx+1, column=7, value=obj.end_year_rating.name if obj.end_year_rating else 'N/A')
            ws_obj.cell(row=idx+1, column=8, value=float(obj.calculated_score))
        
        # ========== COMPETENCIES SHEET ==========
        ws_comp = wb.create_sheet('Competencies')
        headers = ['Group', 'Competency', 'Required Level', 'End-Year Rating', 'Actual Value', 'Gap', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws_comp.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        competencies = performance.competency_ratings.select_related(
            'behavioral_competency', 
            'behavioral_competency__group',
            'leadership_item',
            'end_year_rating'
        ).all()
        
        for idx, comp in enumerate(competencies, 1):
            # ✅ Handle both behavioral and leadership competencies
            if comp.behavioral_competency:
                group_name = comp.behavioral_competency.group.name
                comp_name = comp.behavioral_competency.name
            elif comp.leadership_item:
                group_name = comp.leadership_item.child_group.main_group.name if comp.leadership_item.child_group else 'Leadership'
                comp_name = comp.leadership_item.name
            else:
                group_name = 'N/A'
                comp_name = 'N/A'
            
            ws_comp.cell(row=idx+1, column=1, value=group_name)
            ws_comp.cell(row=idx+1, column=2, value=comp_name)
            ws_comp.cell(row=idx+1, column=3, value=comp.required_level or 0)
            ws_comp.cell(row=idx+1, column=4, value=comp.end_year_rating.name if comp.end_year_rating else 'N/A')
            ws_comp.cell(row=idx+1, column=5, value=comp.actual_value)
            ws_comp.cell(row=idx+1, column=6, value=comp.gap)
            ws_comp.cell(row=idx+1, column=7, value=comp.notes or '')
        
        # ========== DEVELOPMENT NEEDS SHEET ==========
        ws_dev = wb.create_sheet('Development Needs')
        headers = ['Competency Gap', 'Development Activity', 'Progress %', 'Comment']
        for col, header in enumerate(headers, 1):
            cell = ws_dev.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        dev_needs = performance.development_needs.all()
        for idx, need in enumerate(dev_needs, 1):
            ws_dev.cell(row=idx+1, column=1, value=need.competency_gap)
            ws_dev.cell(row=idx+1, column=2, value=need.development_activity)
            ws_dev.cell(row=idx+1, column=3, value=need.progress)
            ws_dev.cell(row=idx+1, column=4, value=need.comment or '')
        
        # ========== AUTO-ADJUST COLUMN WIDTHS ==========
        for ws in [ws_summary, ws_obj, ws_comp, ws_dev]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # ========== SAVE AND RETURN ==========
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Performance_{performance.employee.employee_id}_{performance.performance_year.year}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class PerformanceDashboardViewSet(viewsets.ViewSet):
    """Performance Dashboard Statistics with Access Control"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """✅ FIXED: Get dashboard statistics - filtered by access"""
        year = request.query_params.get('year')
        
        if not year:
            active_year = PerformanceYear.objects.filter(is_active=True).first()
            if not active_year:
                return Response({'error': 'No active performance year'}, status=status.HTTP_404_NOT_FOUND)
            year = active_year.year
        else:
            year = int(year)
        
        try:
            perf_year = PerformanceYear.objects.get(year=year)
        except PerformanceYear.DoesNotExist:
            return Response({'error': f'Performance year {year} not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # ✅ Get accessible employees based on permissions
        from .performance_permissions import get_accessible_employees_for_analytics
        
        accessible_employees, can_view_all, is_manager = get_accessible_employees_for_analytics(request.user)
        
        # Filter performances by accessible employees
        performances = EmployeePerformance.objects.filter(
            performance_year=perf_year,
            employee__in=accessible_employees
        ).select_related('employee', 'employee__department')
        
        total_employees = performances.count()
        
        objectives_completed = performances.filter(
            objectives_employee_approved=True,
            objectives_manager_approved=True
        ).count()
        
        mid_year_completed = performances.filter(mid_year_completed=True).count()
        
        # Count only truly completed performances
        end_year_completed = 0
        
        for perf in performances:
            if perf.approval_status != 'COMPLETED':
                continue
            
            if not perf.competencies_submitted:
                continue
            
            objectives = perf.objectives.filter(is_cancelled=False)
            if objectives.exists():
                all_rated = all(obj.end_year_rating is not None for obj in objectives)
                if not all_rated:
                    continue
            
            if not perf.final_rating or perf.final_rating == 'N/A':
                continue
            
            end_year_completed += 1
        
        pending_employee_approval = performances.filter(
            approval_status='PENDING_EMPLOYEE_APPROVAL'
        ).count()
        
        pending_manager_approval = performances.filter(
            approval_status='PENDING_MANAGER_APPROVAL'
        ).count()
        
        need_clarification = performances.filter(
            approval_status='NEED_CLARIFICATION'
        ).count()
        
        # Department stats - only for accessible employees
        by_department = []
        accessible_dept_names = performances.values_list(
            'employee__department__name', 
            flat=True
        ).distinct()
        
        for dept_name in accessible_dept_names:
            if not dept_name:
                continue
            dept_performances = performances.filter(employee__department__name=dept_name)
            by_department.append({
                'department': dept_name,
                'total': dept_performances.count(),
                'objectives_complete': dept_performances.filter(
                    objectives_employee_approved=True,
                    objectives_manager_approved=True
                ).count(),
                'mid_year_complete': dept_performances.filter(mid_year_completed=True).count(),
                'end_year_complete': dept_performances.filter(end_year_completed=True).count()
            })
        
        # Recent activities - only for accessible employees
        performance_ids = performances.values_list('id', flat=True)
        recent_logs = PerformanceActivityLog.objects.filter(
            performance_id__in=performance_ids
        ).select_related('performance__employee').order_by('-created_at')[:10]
        
        # recent_activities = PerformanceActivityLogSerializer(recent_logs, many=True).data
        
        # Competency grade distribution
        competency_grade_distribution = self._get_grade_distribution(performances)
        
        # ✅ Return access level info
        return Response({
            'total_employees': total_employees,
            'objectives_completed': objectives_completed,
            'mid_year_completed': mid_year_completed,
            'end_year_completed': end_year_completed,
            'pending_employee_approval': pending_employee_approval,
            'pending_manager_approval': pending_manager_approval,
            'need_clarification': need_clarification,
            'current_period': perf_year.get_current_period(),
            'year': year,
            'can_view_all': can_view_all,
            'is_manager': is_manager,
            'viewing_scope': 'all employees' if can_view_all else f'{total_employees} employees (you + team)',
            'timeline': {
                'goal_setting': {
                    'employee_start': perf_year.goal_setting_employee_start,
                    'employee_end': perf_year.goal_setting_employee_end,
                    'manager_start': perf_year.goal_setting_manager_start,
                    'manager_end': perf_year.goal_setting_manager_end
                },
                'mid_year': {
                    'start': perf_year.mid_year_review_start,
                    'end': perf_year.mid_year_review_end
                },
                'end_year': {
                    'start': perf_year.end_year_review_start,
                    'end': perf_year.end_year_review_end
                }
            },
            'by_department': by_department,
            # 'recent_activities': recent_activities,
            'competency_grade_distribution': competency_grade_distribution
        })
    
    def _get_grade_distribution(self, performances):
        """Get distribution of competency letter grades"""
        from collections import Counter
        
        completed = performances.filter(end_year_completed=True)
        grades = completed.values_list('competencies_letter_grade', flat=True)
        distribution = Counter(grades)
        
        return {
            'total': completed.count(),
            'grades': dict(distribution)
        }

class PerformanceNotificationTemplateViewSet(viewsets.ModelViewSet):
    """Performance Notification Templates"""
    queryset = PerformanceNotificationTemplate.objects.all()
    serializer_class = PerformanceNotificationTemplateSerializer
    permission_classes = [IsAuthenticated]