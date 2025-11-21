# api/vacation_views.py - Reorganized Tags for Clean Documentation

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.db.models import Q, Count, Sum
from datetime import date, datetime, timedelta

from .vacation_models import *
from .vacation_serializers import *
from .models import Employee
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import pandas as pd
from .role_models import EmployeeRole
from django.http import HttpResponse
from openpyxl.styles import PatternFill
from .vacation_permissions import (
    has_vacation_permission, 
    has_any_vacation_permission, 
    check_vacation_permission,
    get_user_vacation_permissions
)
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q, Sum, Count
from .vacation_models import EmployeeVacationBalance
from .vacation_serializers import EmployeeVacationBalanceSerializer
from .vacation_permissions import has_vacation_permission, is_admin_user, check_vacation_permission
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

import logging
from django.shortcuts import get_object_or_404
logger = logging.getLogger(__name__)
good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
from rest_framework import status as rest_status
from .vacation_notifications import notification_manager
from .models import UserGraphToken
from .token_helpers import get_user_tokens

def get_graph_access_token(user):
    """
    Get Microsoft Graph access token for the authenticated user
    Used for sending emails via Microsoft Graph API.
    """
    try:
        token = UserGraphToken.get_valid_token(user)
        if token:
            logger.info(f"✅ Valid Graph token found for user {user.username}")
            return token
        else:
            logger.warning(f"⚠️ No valid Graph token found for user {user.username}")
            logger.warning("   Email notifications will be skipped")
            return None
    except Exception as e:
        logger.error(f"❌ Error getting Graph token: {e}")
        return None

def get_notification_context(request):
    """Get notification context with Graph token status"""
    graph_token = get_graph_access_token(request.user)
    
    return {
        'can_send_emails': bool(graph_token),
        'graph_token': graph_token,
        'reason': 'Graph token available' if graph_token else 'No Microsoft Graph token. Login again to enable email notifications.',
        'user': request.user
    }

@swagger_auto_schema(
    method='get',
    operation_description="İstifadəçinin bütün vacation icazələrini əldə et",
    operation_summary="Get My Permissions",
    tags=['Vacation'],
    responses={
        200: openapi.Response(
            description='User permissions',
            examples={
                'application/json': {
                    'is_admin': False,
                    'permissions': [
                        'vacation.dashboard.view_own',
                        'vacation.request.create_own'
                    ],
                    'roles': ['Employee - Vacation']
                }
            }
        )
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_vacation_permissions(request):
    """İstifadəçinin vacation permissions-larını göstər"""
    from .vacation_permissions import is_admin_user
    
    is_admin = is_admin_user(request.user)
    permissions = get_user_vacation_permissions(request.user)
    
    # User roles
    try:
        emp = Employee.objects.get(user=request.user, is_deleted=False)
        roles = list(EmployeeRole.objects.filter(
            employee=emp,
            is_active=True
        ).values_list('role__name', flat=True))
    except Employee.DoesNotExist:
        roles = []
    
    return Response({
        'is_admin': is_admin,
        'permissions_count': len(permissions),
        'permissions': permissions,
        'roles_count': len(roles),
        'roles': roles
    })

# ==================== DASHBOARD ====================
# vacation_views.py

@swagger_auto_schema(
    method='get',
    operation_description="Dashboard məlumatları - 6 stat card və əsas statistika",
    operation_summary="Dashboard",
    tags=['Vacation'],
    responses={
        200: openapi.Response(
            description='Dashboard data',
            examples={
                'application/json': {
                    'balance': {
                        'total_balance': 28.0,
                        'yearly_balance': 28.0,
                        'used_days': 5.0,
                        'remaining_balance': 23.0,
                        'scheduled_days': 3.0,
                        'should_be_planned': 25.0
                    }
                }
            }
        )
    }
)
@api_view(['GET'])
@has_vacation_permission('vacation.dashboard.view_own')
@permission_classes([IsAuthenticated])
def vacation_dashboard(request):
    """Dashboard - 6 stat card"""
    try:
        emp = Employee.objects.get(user=request.user, is_deleted=False)
        year = date.today().year
        
        # ✅ DÜZƏLİŞ: get_or_create SILINDI - yalnız mövcud balansı tap
        try:
            balance = EmployeeVacationBalance.objects.get(
                employee=emp, 
                year=year,
                is_deleted=False
            )
        except EmployeeVacationBalance.DoesNotExist:
            # ✅ Balance yoxdursa, 0 göstər (avtomatik yaratma)
            return Response({
                'balance': {
                    'total_balance': 0,
                    'yearly_balance': 0,
                    'used_days': 0,
                    'remaining_balance': 0,
                    'scheduled_days': 0,
                    'should_be_planned': 0
                }
            })
        
        # Balance varsa, refresh et
        balance.refresh_from_db()
        
        return Response({
            'balance': {
                'total_balance': float(balance.total_balance),
                'yearly_balance': float(balance.yearly_balance),
                'used_days': float(balance.used_days),
                'remaining_balance': float(balance.remaining_balance),
                'scheduled_days': float(balance.scheduled_days),
                'should_be_planned': float(balance.should_be_planned)
            }
        })
    except Employee.DoesNotExist:
        return Response({
            'error': 'Employee profili tapılmadı'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Dashboard error: {e}")
        return Response({
            'balance': {
                'total_balance': 0, 
                'yearly_balance': 0, 
                'used_days': 0, 
                'remaining_balance': 0, 
                'scheduled_days': 0, 
                'should_be_planned': 0
            }
        })

# ==================== PRODUCTION CALENDAR SETTINGS ====================
# vacation_views.py

from collections import defaultdict
from datetime import datetime, timedelta

@swagger_auto_schema(
    method='get',
    operation_description="Calendar view - holidays və vacation events",
    operation_summary="Get Calendar Events",
    tags=['Vacation'],
    manual_parameters=[
        openapi.Parameter(
            'month',
            openapi.IN_QUERY,
            description="Ay (1-12)",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        openapi.Parameter(
            'year',
            openapi.IN_QUERY,
            description="İl (məsələn: 2025)",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        openapi.Parameter(
            'employee_id',
            openapi.IN_QUERY,
            description="Employee ID (filter üçün)",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        openapi.Parameter(
            'department_id',
            openapi.IN_QUERY,
            description="Department ID (filter üçün)",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        openapi.Parameter(
            'business_function_id',
            openapi.IN_QUERY,
            description="Business Function ID (filter üçün)",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
    ],
    responses={
        200: openapi.Response(
            description='Calendar events',
            examples={
                'application/json': {
                    'holidays': [
                        {
                            'date': '2025-01-01',
                            'name': 'New Year',
                            'type': 'holiday'
                        }
                    ],
                    'vacations': [
                        {
                            'id': 1,
                            'type': 'request',
                            'employee_name': 'John Doe',
                            'employee_id': 'EMP001',
                            'vacation_type': 'Annual Leave',
                            'start_date': '2025-01-10',
                            'end_date': '2025-01-15',
                            'status': 'Approved',
                            'days': 4
                        }
                    ],
                    'summary': {
                        'total_holidays': 10,
                        'total_vacations': 25,
                        'employees_on_vacation': 15
                    }
                }
            }
        )
    }
)
@api_view(['GET'])
@has_any_vacation_permission([
    'vacation.calendar.view_all',
    'vacation.calendar.view_team',
    'vacation.calendar.view_own'
])
@permission_classes([IsAuthenticated])
def get_calendar_events(request):
    """Calendar view - holidays və vacation events"""
    from .vacation_permissions import is_admin_user
    
    try:
        # Get filters
        month = request.GET.get('month')
        year = request.GET.get('year')
        employee_id = request.GET.get('employee_id')
        department_id = request.GET.get('department_id')
        business_function_id = request.GET.get('business_function_id')
        
        # Default to current month/year
        if not month or not year:
            today = date.today()
            month = month or today.month
            year = year or today.year
        
        month = int(month)
        year = int(year)
        
        # Calculate date range
        start_date = date(year, month, 1)
        # Last day of month
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # ✅ Get holidays from settings
        settings = VacationSetting.get_active()
        holidays = []
        
        if settings and settings.non_working_days:
            for holiday in settings.non_working_days:
                if isinstance(holiday, dict):
                    holiday_date = datetime.strptime(holiday['date'], '%Y-%m-%d').date()
                    if start_date <= holiday_date <= end_date:
                        holidays.append({
                            'date': holiday['date'],
                            'name': holiday.get('name', 'Holiday'),
                            'type': 'holiday'
                        })
        
        # ✅ Get vacation requests
        requests_qs = VacationRequest.objects.filter(
            is_deleted=False,
            status__in=['PENDING_LINE_MANAGER', 'PENDING_HR', 'APPROVED']
        ).filter(
            Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
        ).select_related(
            'employee', 
            'employee__department', 
            'employee__business_function',
            'vacation_type'
        )
        
        # ✅ Get vacation schedules
        schedules_qs = VacationSchedule.objects.filter(
            is_deleted=False,
            status__in=['SCHEDULED', 'REGISTERED']
        ).filter(
            Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
        ).select_related(
            'employee', 
            'employee__department', 
            'employee__business_function',
            'vacation_type'
        )
        
        # ✅ Permission-based filtering
        is_admin = is_admin_user(request.user)
        has_view_all, _ = check_vacation_permission(request.user, 'vacation.calendar.view_all')
        has_view_team, _ = check_vacation_permission(request.user, 'vacation.calendar.view_team')
        
        if not (is_admin or has_view_all):
            if has_view_team:
                # Team member-ləri göstər
                try:
                    emp = Employee.objects.get(user=request.user, is_deleted=False)
                    team_employees = Employee.objects.filter(
                        line_manager=emp,
                        is_deleted=False
                    ).values_list('id', flat=True)
                    
                    allowed_employee_ids = list(team_employees) + [emp.id]
                    
                    requests_qs = requests_qs.filter(employee_id__in=allowed_employee_ids)
                    schedules_qs = schedules_qs.filter(employee_id__in=allowed_employee_ids)
                except Employee.DoesNotExist:
                    # Yalnız özü
                    requests_qs = requests_qs.filter(employee__user=request.user)
                    schedules_qs = schedules_qs.filter(employee__user=request.user)
            else:
                # Yalnız özünü göstər
                requests_qs = requests_qs.filter(employee__user=request.user)
                schedules_qs = schedules_qs.filter(employee__user=request.user)
        
        # ✅ Apply additional filters
        if employee_id:
            requests_qs = requests_qs.filter(employee_id=employee_id)
            schedules_qs = schedules_qs.filter(employee_id=employee_id)
        
        if department_id:
            requests_qs = requests_qs.filter(employee__department_id=department_id)
            schedules_qs = schedules_qs.filter(employee__department_id=department_id)
        
        if business_function_id:
            requests_qs = requests_qs.filter(employee__business_function_id=business_function_id)
            schedules_qs = schedules_qs.filter(employee__business_function_id=business_function_id)
        
        # ✅ Build vacation events
        vacations = []
        employee_ids_on_vacation = set()
        
        for req in requests_qs:
            vacations.append({
                'id': req.id,
                'type': 'request',
                'request_id': req.request_id,
                'employee_id': req.employee.id,
                'employee_name': req.employee.full_name,
                'employee_code': getattr(req.employee, 'employee_id', ''),
                'department': req.employee.department.name if req.employee.department else '',
                'business_function': req.employee.business_function.name if req.employee.business_function else '',
                'vacation_type': req.vacation_type.name,
                'vacation_type_id': req.vacation_type.id,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'status': req.get_status_display(),
                'status_code': req.status,
                'days': float(req.number_of_days),
                'comment': req.comment
            })
            employee_ids_on_vacation.add(req.employee.id)
        
        for sch in schedules_qs:
            vacations.append({
                'id': sch.id,
                'type': 'schedule',
                'request_id': f'SCH{sch.id}',
                'employee_id': sch.employee.id,
                'employee_name': sch.employee.full_name,
                'employee_code': getattr(sch.employee, 'employee_id', ''),
                'department': sch.employee.department.name if sch.employee.department else '',
                'business_function': sch.employee.business_function.name if sch.employee.business_function else '',
                'vacation_type': sch.vacation_type.name,
                'vacation_type_id': sch.vacation_type.id,
                'start_date': sch.start_date.strftime('%Y-%m-%d'),
                'end_date': sch.end_date.strftime('%Y-%m-%d'),
                'status': sch.get_status_display(),
                'status_code': sch.status,
                'days': float(sch.number_of_days),
                'comment': sch.comment
            })
            employee_ids_on_vacation.add(sch.employee.id)
        
        # ✅ Summary
        summary = {
            'total_holidays': len(holidays),
            'total_vacations': len(vacations),
            'employees_on_vacation': len(employee_ids_on_vacation),
            'month': month,
            'year': year
        }
        
        return Response({
            'holidays': holidays,
            'vacations': vacations,
            'summary': summary,
            'filters_applied': {
                'month': month,
                'year': year,
                'employee_id': employee_id,
                'department_id': department_id,
                'business_function_id': business_function_id
            }
        })
        
    except Exception as e:
        logger.error(f"Calendar events error: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='put',
    operation_description="Production Calendar - qeyri-iş günlərini yenilə",
    operation_summary="Update Non-Working Days",
    tags=['Vacation - Settings'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['non_working_days'],
        properties={
            'non_working_days': openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'date': openapi.Schema(type=openapi.TYPE_STRING, format='date'),
                        'name': openapi.Schema(type=openapi.TYPE_STRING, description='Holiday name')
                    }
                ),
                description='Qeyri-iş günləri massivi',
                example=[
                    {'date': '2025-01-01', 'name': 'New Year'},
                    {'date': '2025-01-20', 'name': 'Martyrs Day'},
                    {'date': '2025-03-08', 'name': 'Women Day'},
                    {'date': '2025-03-20', 'name': 'Novruz Bayram'}
                ]
            ),
        }
    ),
    responses={
        200: openapi.Response(
            description='Production calendar yeniləndi',
            examples={
                'application/json': {
                    'message': 'Production calendar uğurla yeniləndi',
                    'non_working_days': [
                        {'date': '2025-01-01', 'name': 'New Year'},
                        {'date': '2025-01-20', 'name': 'Martyrs Day'}
                    ]
                }
            }
        )
    }
)
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@has_vacation_permission('vacation.settings.update_production_calendar')
def update_production_calendar(request):
    """Production Calendar - qeyri-iş günlərini yenilə"""
    try:
        serializer = ProductionCalendarSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        non_working_days = serializer.validated_data['non_working_days']
        
        # Active settings tap və ya yarat
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        # Dictionary formatında saxla
        settings.non_working_days = non_working_days
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Production calendar uğurla yeniləndi',
            'non_working_days': settings.non_working_days,
            'updated_at': settings.updated_at,
            'updated_by': request.user.get_full_name() or request.user.username
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='post',
    operation_description="Production Calendar - qeyri-iş günlərini təyin et",
    operation_summary="Set Non-Working Days",
    tags=['Vacation - Settings'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['non_working_days'],
        properties={
            'non_working_days': openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'date': openapi.Schema(type=openapi.TYPE_STRING, format='date'),
                        'name': openapi.Schema(type=openapi.TYPE_STRING, description='Holiday name')
                    }
                ),
                description='Qeyri-iş günləri massivi',
                example=[
                    {'date': '2025-01-01', 'name': 'New Year'},
                    {'date': '2025-01-20', 'name': 'Martyrs Day'},
                    {'date': '2025-03-08', 'name': 'Women Day'},
                    {'date': '2025-03-20', 'name': 'Novruz Bayram'}
                ]
            ),
        }
    ),
    responses={
        200: openapi.Response(
            description='Production calendar yeniləndi',
            examples={
                'application/json': {
                    'message': 'Production calendar uğurla yeniləndi',
                    'non_working_days': [
                        {'date': '2025-01-01', 'name': 'New Year'},
                        {'date': '2025-01-20', 'name': 'Martyrs Day'}
                    ]
                }
            }
        )
    }
)
@api_view(['POST'])
@has_vacation_permission('vacation.settings.update_production_calendar')
@permission_classes([IsAuthenticated])
def set_production_calendar(request):
    """Production Calendar - qeyri-iş günlərini təyin et"""
    try:
        serializer = ProductionCalendarSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        non_working_days = serializer.validated_data['non_working_days']
        
        # Active settings tap və ya yarat
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        # Dictionary formatında saxla
        settings.non_working_days = non_working_days
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Production calendar uğurla yeniləndi',
            'non_working_days': settings.non_working_days
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='get',
    operation_description="Production Calendar məlumatlarını əldə et",
    operation_summary="Get Production Calendar",
    tags=['Vacation - Settings'],
    responses={
        200: openapi.Response(
            description='Production calendar məlumatları',
            examples={
                'application/json': {
                    'non_working_days': [
                        {'date': '2025-01-01', 'name': 'New Year'},
                        {'date': '2025-01-20', 'name': 'Martyrs Day'}
                    ]
                }
            }
        )
    }
)
@api_view(['GET'])
@has_vacation_permission('vacation.settings.view')
@permission_classes([IsAuthenticated])
def get_production_calendar(request):
    """Production Calendar məlumatlarını əldə et"""
    try:
        settings = VacationSetting.get_active()
        
        return Response({
            'non_working_days': settings.non_working_days if settings else []
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== GENERAL VACATION SETTINGS  ====================
@swagger_auto_schema(
    method='put',
    operation_description="Vacation ümumi parametrləri - balans, edit limiti, bildirişlər yenilə",
    operation_summary="Update General Vacation Settings",
    tags=['Vacation - Settings'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'allow_negative_balance': openapi.Schema(
                type=openapi.TYPE_BOOLEAN,
                description='Balans 0 olduqda request yaratmağa icazə ver',
                example=False
            ),
            'max_schedule_edits': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Schedule neçə dəfə edit oluna bilər',
                example=3
            ),
            'notification_days_before': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Məzuniyyət başlamazdan neçə gün əvvəl bildiriş göndər',
                example=7
            ),
            'notification_frequency': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Bildirişi neçə dəfə göndər',
                example=2
            ),
        }
    ),
    responses={
        200: openapi.Response(
            description='Ümumi parametrlər yeniləndi',
            examples={
                'application/json': {
                    'message': 'Vacation parametrləri uğurla yeniləndi',
                    'settings': {
                        'allow_negative_balance': False,
                        'max_schedule_edits': 3,
                        'notification_days_before': 7,
                        'notification_frequency': 2
                    },
                    'updated_at': '2025-09-25T10:30:00Z',
                    'updated_by': 'John Doe'
                }
            }
        )
    }
)
@api_view(['PUT'])
@has_vacation_permission('vacation.settings.update_general')
@permission_classes([IsAuthenticated])
def update_general_vacation_settings(request):
    """Vacation ümumi parametrlərini yenilə"""
    try:
        serializer = GeneralVacationSettingsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        
        # Active settings tap və ya yarat
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        # Parametrləri yenilə
        if 'allow_negative_balance' in data:
            settings.allow_negative_balance = data['allow_negative_balance']
        
        if 'max_schedule_edits' in data:
            settings.max_schedule_edits = data['max_schedule_edits']
        
        if 'notification_days_before' in data:
            settings.notification_days_before = data['notification_days_before']
        
        if 'notification_frequency' in data:
            settings.notification_frequency = data['notification_frequency']
        
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Vacation parametrləri uğurla yeniləndi',
            'settings': {
                'allow_negative_balance': settings.allow_negative_balance,
                'max_schedule_edits': settings.max_schedule_edits,
                'notification_days_before': settings.notification_days_before,
                'notification_frequency': settings.notification_frequency
            },
            'updated_at': settings.updated_at,
            'updated_by': request.user.get_full_name() or request.user.username
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='post',
    operation_description="Vacation ümumi parametrləri - balans, edit limiti, bildirişlər",
    operation_summary="Set General Vacation Settings",
    tags=['Vacation - Settings'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'allow_negative_balance': openapi.Schema(
                type=openapi.TYPE_BOOLEAN,
                description='Balans 0 olduqda request yaratmağa icazə ver',
                example=False
            ),
            'max_schedule_edits': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Schedule neçə dəfə edit oluna bilər',
                example=3
            ),
            'notification_days_before': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Məzuniyyət başlamazdan neçə gün əvvəl bildiriş göndər',
                example=7
            ),
            'notification_frequency': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Bildirişi neçə dəfə göndər',
                example=2
            ),
        }
    ),
    responses={
        200: openapi.Response(
            description='Ümumi parametrlər yeniləndi',
            examples={
                'application/json': {
                    'message': 'Vacation parametrləri uğurla yeniləndi',
                    'settings': {
                        'allow_negative_balance': False,
                        'max_schedule_edits': 3,
                        'notification_days_before': 7,
                        'notification_frequency': 2
                    }
                }
            }
        )
    }
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@has_vacation_permission('vacation.settings.update_general')
def set_general_vacation_settings(request):
    """Vacation ümumi parametrləri"""
    try:
        serializer = GeneralVacationSettingsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        
        # Active settings tap və ya yarat
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        # Parametrləri yenilə
        if 'allow_negative_balance' in data:
            settings.allow_negative_balance = data['allow_negative_balance']
        
        if 'max_schedule_edits' in data:
            settings.max_schedule_edits = data['max_schedule_edits']
        
        if 'notification_days_before' in data:
            settings.notification_days_before = data['notification_days_before']
        
        if 'notification_frequency' in data:
            settings.notification_frequency = data['notification_frequency']
        
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Vacation parametrləri uğurla yeniləndi',
            'settings': {
                'allow_negative_balance': settings.allow_negative_balance,
                'max_schedule_edits': settings.max_schedule_edits,
                'notification_days_before': settings.notification_days_before,
                'notification_frequency': settings.notification_frequency
            }
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='get',
    operation_description="Vacation ümumi parametrlərini əldə et",
    operation_summary="Get General Vacation Settings",
    tags=['Vacation - Settings'],
    responses={
        200: openapi.Response(
            description='Ümumi parametrlər',
            examples={
                'application/json': {
                    'allow_negative_balance': False,
                    'max_schedule_edits': 3,
                    'notification_days_before': 7,
                    'notification_frequency': 2
                }
            }
        )
    }
)
@api_view(['GET'])
@has_vacation_permission('vacation.settings.view')
@permission_classes([IsAuthenticated])
def get_general_vacation_settings(request):
    """Vacation ümumi parametrlərini əldə et"""
    try:
        settings = VacationSetting.get_active()
        
        if not settings:
            return Response({
                'allow_negative_balance': False,
                'max_schedule_edits': 3,
                'notification_days_before': 7,
                'notification_frequency': 2
            })
        
        return Response({
            'allow_negative_balance': settings.allow_negative_balance,
            'max_schedule_edits': settings.max_schedule_edits,
            'notification_days_before': settings.notification_days_before,
            'notification_frequency': settings.notification_frequency
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ==================== HR REPRESENTATIVE SETTINGS  ====================
@swagger_auto_schema(
    method='put',
    operation_description="Default HR nümayəndəsini yenilə",
    operation_summary="Update Default HR Representative",
    tags=['Vacation - Settings'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['default_hr_representative_id'],
        properties={
            'default_hr_representative_id': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Default HR nümayəndəsi Employee ID',
                example=5
            ),
        }
    ),
    responses={
        200: openapi.Response(
            description='Default HR yeniləndi',
            examples={
                'application/json': {
                    'message': 'Default HR nümayəndəsi uğurla yeniləndi',
                    'previous_hr': {
                        'id': 3,
                        'name': 'Previous HR Rep',
                        'department': 'HR'
                    },
                    'current_hr': {
                        'id': 5,
                        'name': 'Sarah Johnson',
                        'department': 'HR'
                    },
                    'updated_at': '2025-09-25T10:30:00Z',
                    'updated_by': 'Admin User'
                }
            }
        )
    }
)
@api_view(['PUT'])
@has_vacation_permission('vacation.settings.update_hr_representative')
@permission_classes([IsAuthenticated])
def update_default_hr_representative(request):
    """Default HR nümayəndəsini yenilə"""
    try:
        serializer = HRRepresentativeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        hr_id = serializer.validated_data['default_hr_representative_id']
        hr_employee = Employee.objects.get(id=hr_id, is_deleted=False)
        
        # Active settings tap və ya yarat
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        # Əvvəlki HR-ı saxla
        previous_hr = settings.default_hr_representative
        previous_hr_info = None
        if previous_hr:
            previous_hr_info = {
                'id': previous_hr.id,
                'name': previous_hr.full_name,
                'department': previous_hr.department.name if previous_hr.department else ''
            }
        
        # Yeni HR təyin et
        settings.default_hr_representative = hr_employee
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Default HR nümayəndəsi uğurla yeniləndi',
            'previous_hr': previous_hr_info,
            'current_hr': {
                'id': hr_employee.id,
                'name': hr_employee.full_name,
                'department': hr_employee.department.name if hr_employee.department else ''
            },
            'updated_at': settings.updated_at,
            'updated_by': request.user.get_full_name() or request.user.username
        })
        
    except Employee.DoesNotExist:
        return Response({
            'error': 'HR nümayəndəsi tapılmadı'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='post',
    operation_description="Default HR nümayəndəsini təyin et",
    operation_summary="Set Default HR Representative",
    tags=['Vacation - Settings'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['default_hr_representative_id'],
        properties={
            'default_hr_representative_id': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Default HR nümayəndəsi Employee ID',
                example=5
            ),
        }
    ),
    responses={
        200: openapi.Response(
            description='Default HR təyin edildi',
            examples={
                'application/json': {
                    'message': 'Default HR nümayəndəsi uğurla təyin edildi',
                    'hr_representative': {
                        'id': 5,
                        'name': 'Sarah Johnson',
                        'department': 'HR'
                    }
                }
            }
        )
    }
)
@api_view(['POST'])
@has_vacation_permission('vacation.settings.update_hr_representative')
@permission_classes([IsAuthenticated])
def set_default_hr_representative(request):
    """Default HR nümayəndəsini təyin et"""
    try:
        serializer = HRRepresentativeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        hr_id = serializer.validated_data['default_hr_representative_id']
        hr_employee = Employee.objects.get(id=hr_id, is_deleted=False)
        
        # Active settings tap və ya yarat
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        settings.default_hr_representative = hr_employee
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Default HR nümayəndəsi uğurla təyin edildi',
            'hr_representative': {
                'id': hr_employee.id,
                'name': hr_employee.full_name,
                'department': hr_employee.department.name if hr_employee.department else ''
            }
        })
        
    except Employee.DoesNotExist:
        return Response({
            'error': 'HR nümayəndəsi tapılmadı'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='get',
    operation_description="Mövcud HR nümayəndələrini əldə et",
    operation_summary="Get HR Representatives",
    tags=['Vacation - Settings'],
    responses={
        200: openapi.Response(
            description='HR nümayəndələri siyahısı'
        )
    }
)
@api_view(['GET'])
@has_vacation_permission('vacation.settings.view')
@permission_classes([IsAuthenticated])
def get_hr_representatives(request):
    """HR nümayəndələrini əldə et"""
    try:
        settings = VacationSetting.get_active()
        current_default = settings.default_hr_representative if settings else None
        
        # HR departamentindəki işçilər
        hr_employees = Employee.objects.filter(
            department__name__icontains='HR',
            is_deleted=False
        )
        
        hr_list = []
        for emp in hr_employees:
            hr_list.append({
                'id': emp.id,
                'name': emp.full_name,
                'email': emp.user.email if emp.user else '',
                'phone': emp.phone,
                'department': emp.department.name if emp.department else '',
                'is_default': current_default and current_default.id == emp.id
            })
        
        return Response({
            'current_default': {
                'id': current_default.id,
                'name': current_default.full_name,
                'department': current_default.department.name if current_default.department else ''
            } if current_default else None,
            'hr_representatives': hr_list
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)




# ==================== BALANCE MANAGEMENT ====================
@swagger_auto_schema(
    method='post',
    operation_description="Excel faylı ilə vacation balanslarını toplu yükle",
    operation_summary="Bulk Upload Balances",
    tags=['Vacation - Settings'],
    consumes=['multipart/form-data'],
    manual_parameters=[
        openapi.Parameter(
            'file',
            openapi.IN_FORM,
            description="Excel faylı (.xlsx)",
            type=openapi.TYPE_FILE,
            required=True
        ),
        openapi.Parameter(
            'year',
            openapi.IN_FORM,
            description="İl (məsələn: 2025)",
            type=openapi.TYPE_INTEGER,
            required=True
        )
    ],
    responses={
        200: openapi.Response(
            description='Upload successful',
            examples={
                'application/json': {
                    'message': '10 uğurlu, 0 səhv',
                    'results': {
                        'successful': 10,
                        'failed': 0,
                        'errors': []
                    }
                }
            }
        )
    }
)
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([IsAuthenticated])
@has_vacation_permission('vacation.balance.bulk_upload')
def bulk_upload_balances(request):
    """Excel ilə balance upload et"""
    if 'file' not in request.FILES:
        return Response({'error': 'File yoxdur'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    year = int(request.data.get('year', date.today().year))
    
    try:
        # Read Excel - header at row 5 (index 4), skip row 6 (descriptions)
        # Only read up to row 100 to avoid footer text
        df = pd.read_excel(file, header=4, skiprows=[5], nrows=100)
        
        # Strip whitespace and normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Required columns check
        required_cols = ['employee_id', 'start_balance', 'yearly_balance']
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            return Response({
                'error': f'Missing columns: {", ".join(missing_cols)}',
                'found_columns': list(df.columns),
                'hint': 'Please use the downloaded template without modifications'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        results = {'successful': 0, 'failed': 0, 'errors': [], 'skipped': 0}
        
        for idx, row in df.iterrows():
            try:
                # Get employee_id and check if it's valid
                emp_id_raw = row.get('employee_id')
                
                # Skip if empty, NaN, or not a valid employee ID format
                if pd.isna(emp_id_raw):
                    results['skipped'] += 1
                    continue
                
                emp_id = str(emp_id_raw).strip()
                
                # Skip empty strings or very long strings (likely footer text)
                if not emp_id or len(emp_id) > 20:
                    results['skipped'] += 1
                    continue
                
                # Find employee
                emp = Employee.objects.get(
                    employee_id=emp_id, 
                    is_deleted=False
                )
                
                start_bal = float(row['start_balance']) if pd.notna(row['start_balance']) else 0
                yearly_bal = float(row['yearly_balance']) if pd.notna(row['yearly_balance']) else 0
                
                EmployeeVacationBalance.objects.update_or_create(
                    employee=emp,
                    year=year,
                    defaults={
                        'start_balance': start_bal,
                        'yearly_balance': yearly_bal,
                        'updated_by': request.user
                    }
                )
                
                results['successful'] += 1
                
            except Employee.DoesNotExist:
                results['errors'].append(f"Employee ID '{row['employee_id']}' sistemdə tapılmadı")
                results['failed'] += 1
            except ValueError as e:
                results['errors'].append(f"Employee ID '{row['employee_id']}': Rəqəm formatı səhvdir")
                results['failed'] += 1
            except Exception as e:
                results['errors'].append(f"Employee ID '{row['employee_id']}': {str(e)}")
                results['failed'] += 1
        
        # Return success only if at least one row was processed
        if results['successful'] == 0 and results['failed'] == 0:
            return Response({
                'error': 'Faylda heç bir məlumat tapılmadı',
                'hint': 'Please add employee data starting from row 7'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'message': f"{results['successful']} uğurlu, {results['failed']} səhv",
            'results': results
        })
    
    except Exception as e:
        return Response({
            'error': f'File processing error: {str(e)}',
            'hint': 'Make sure you are using the latest template file'
        }, status=status.HTTP_400_BAD_REQUEST)
        
@swagger_auto_schema(
    method='get',
    operation_description="Excel template endir",
    operation_summary="Download Balance Template",
    tags=['Vacation - Settings'],
    responses={
        200: openapi.Response(
            description='Excel file',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@has_vacation_permission('vacation.balance.view_all')
def download_balance_template(request):
    """Enhanced Excel template with improved design"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Template"

    ws.sheet_view.showGridLines = False

    # Define styles
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(size=16, bold=True, color="305496")
    desc_font = Font(size=9, italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:C2')
    ws['A1'] = 'VACATION BALANCE TEMPLATE'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Date info
    ws['A3'] = f'Generated on: {datetime.now().strftime("%B %d, %Y - %H:%M")}'
    ws['A3'].font = Font(size=10, italic=True, color="808080")

    # Column headers
    headers = ['employee_id', 'start_balance', 'yearly_balance']
    descriptions = [
        'Employee ID from system',
        'Remaining balance from previous year',
        'Annual vacation allocation',
    ]

    start_row = 5
    for col, (header, desc) in enumerate(zip(headers, descriptions), 1):
        # Header
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Description
        desc_cell = ws.cell(row=start_row + 1, column=col, value=desc)
        desc_cell.font = desc_font
        desc_cell.alignment = Alignment(horizontal='center', wrap_text=True)
        desc_cell.border = border

    # Add instruction in a separate area (far from data area)
    instruction_row = 15  # Much further down
    ws.merge_cells(f'A{instruction_row}:C{instruction_row}')
    ws[f'A{instruction_row}'] = "Instructions: Fill employee data starting from row 7. Do not modify this template structure."
    ws[f'A{instruction_row}'].font = Font(size=9, italic=True, color="808080")
    ws[f'A{instruction_row}'].alignment = Alignment(horizontal='left')

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=vacation_balance_template.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@has_vacation_permission('vacation.balance.view_all')
def get_all_balances(request):
    """
    GET /vacation/balances/
    Get all employee vacation balances with filters
    Permission: vacation.balance.view_all or admin
    """
    user = request.user
    
    # Get filters
    year = request.GET.get('year', datetime.now().year)
    department_id = request.GET.get('department_id', '')  # ✅ Changed from department name to ID
    business_function_id = request.GET.get('business_function_id', '')  # ✅ YENİ
    min_remaining = request.GET.get('min_remaining', '')
    max_remaining = request.GET.get('max_remaining', '')
    
    # Build queryset
    queryset = EmployeeVacationBalance.objects.filter(
        is_deleted=False,
        year=year
    ).select_related('employee', 'employee__department', 'employee__business_function')
    
    # Apply filters
    if department_id:
        queryset = queryset.filter(employee__department_id=department_id)
    
    # ✅ YENİ: Business Function filter
    if business_function_id:
        queryset = queryset.filter(employee__business_function_id=business_function_id)
    
    # Calculate filtered balances
    balances_list = []
    for balance in queryset:
        # Apply remaining balance filters
        if min_remaining and balance.remaining_balance < float(min_remaining):
            continue
        if max_remaining and balance.remaining_balance > float(max_remaining):
            continue
        balances_list.append(balance)
    
    # Serialize
    serializer = EmployeeVacationBalanceSerializer(balances_list, many=True)
    balances = serializer.data
    
    # Calculate summary
    total_allocated = sum(float(b.total_balance) for b in balances_list)
    total_used = sum(float(b.used_days) for b in balances_list)
    total_scheduled = sum(float(b.scheduled_days) for b in balances_list)
    total_remaining = sum(float(b.remaining_balance) for b in balances_list)
    employee_count = len(balances_list)
    
    summary = {
        'total_employees': employee_count,
        'total_allocated': round(total_allocated, 1),
        'total_used': round(total_used, 1),
        'total_scheduled': round(total_scheduled, 1),
        'total_remaining': round(total_remaining, 1)
    }
    
    return Response({
        'balances': balances,
        'summary': summary,
        'filters_applied': {
            'year': year,
            'department_id': department_id,
            'business_function_id': business_function_id,  # ✅ YENİ
            'min_remaining': min_remaining,
            'max_remaining': max_remaining
        }
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@has_vacation_permission('vacation.balance.export')
def export_all_balances(request):
    """
    GET /vacation/balances/export/
    Export all balances to Excel
    Permission: vacation.balance.export or admin
    """
    user = request.user
    
    # Get filters
    year = request.GET.get('year', datetime.now().year)
    department_id = request.GET.get('department_id', '')
    min_remaining = request.GET.get('min_remaining', '')
    max_remaining = request.GET.get('max_remaining', '')
    business_function_id = request.GET.get('business_function_id', '')
    queryset = EmployeeVacationBalance.objects.filter(
        is_deleted=False,
        year=year
    ).select_related('employee', 'employee__department').order_by(
        'employee__department__name', 'employee__full_name'
    )
    
    if department_id:
        queryset = queryset.filter(employee__department_id=department_id)
    
    # ✅ YENİ: Business Function filter
    if business_function_id:
        queryset = queryset.filter(employee__business_function_id=business_function_id)
    
    # Apply remaining balance filters
    balances_list = []
    for balance in queryset:
        if min_remaining and balance.remaining_balance < float(min_remaining):
            continue
        if max_remaining and balance.remaining_balance > float(max_remaining):
            continue
        balances_list.append(balance)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Vacation Balances {year}'
    
    # Header style
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = [
        'Employee Name', 'Employee ID', 'Department', 'Year',
        'Start Balance', 'Yearly Balance', 'Total Balance',
        'Used Days', 'Scheduled Days', 'Remaining Balance', 'To Plan'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, balance in enumerate(balances_list, 2):
        ws.cell(row=row_num, column=1, value=balance.employee.full_name)
        ws.cell(row=row_num, column=2, value=getattr(balance.employee, 'employee_id', ''))
        ws.cell(row=row_num, column=3, value=balance.employee.department.name if balance.employee.department else '')
        ws.cell(row=row_num, column=4, value=balance.year)
        ws.cell(row=row_num, column=5, value=float(balance.start_balance))
        ws.cell(row=row_num, column=6, value=float(balance.yearly_balance))
        ws.cell(row=row_num, column=7, value=float(balance.total_balance))
        ws.cell(row=row_num, column=8, value=float(balance.used_days))
        ws.cell(row=row_num, column=9, value=float(balance.scheduled_days))
        ws.cell(row=row_num, column=10, value=float(balance.remaining_balance))
        ws.cell(row=row_num, column=11, value=float(balance.should_be_planned))
        
        # Center align numeric columns
        for col in range(4, 12):
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=vacation_balances_{year}.xlsx'
    wb.save(response)
    
    return response


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@has_vacation_permission('vacation.balance.update')
def update_employee_balance(request):
    """
    PUT /vacation/balances/update/
    Update individual employee balance
    Permission: vacation.balance.update or admin
    """
    employee_id = request.data.get('employee_id')
    year = request.data.get('year', datetime.now().year)
    
    if not employee_id:
        return Response(
            {'error': 'employee_id is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        from .models import Employee
        employee = Employee.objects.get(id=employee_id, is_deleted=False)
    except Employee.DoesNotExist:
        return Response(
            {'error': 'Employee not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Get or create balance
    balance, created = EmployeeVacationBalance.objects.get_or_create(
        employee=employee,
        year=year,
        defaults={
            'start_balance': request.data.get('start_balance', 0),
            'yearly_balance': request.data.get('yearly_balance', 28),
            'used_days': request.data.get('used_days', 0),
            'scheduled_days': request.data.get('scheduled_days', 0),
            'updated_by': request.user
        }
    )
    
    if not created:
        # Update existing balance
        if 'start_balance' in request.data:
            balance.start_balance = request.data['start_balance']
        if 'yearly_balance' in request.data:
            balance.yearly_balance = request.data['yearly_balance']
        if 'used_days' in request.data:
            balance.used_days = request.data['used_days']
        if 'scheduled_days' in request.data:
            balance.scheduled_days = request.data['scheduled_days']
        
        balance.updated_by = request.user
        balance.save()
    
    serializer = EmployeeVacationBalanceSerializer(balance)
    return Response({
        'message': 'Balance updated successfully',
        'balance': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@has_vacation_permission('vacation.balance.reset')
def reset_balances(request):
    """
    POST /vacation/balances/reset/
    Reset balances for a specific year
    Permission: vacation.balance.reset or admin
    """
    year = request.data.get('year', datetime.now().year)
    department_id = request.data.get('department_id', None)
    
    queryset = EmployeeVacationBalance.objects.filter(
        is_deleted=False,
        year=year
    )
    
    if department_id:
        queryset = queryset.filter(employee__department_id=department_id)
    
    # Reset used_days and scheduled_days to 0
    updated_count = queryset.update(
        used_days=0,
        scheduled_days=0,
        updated_by=request.user
    )
    
    return Response({
        'message': f'Reset {updated_count} balances for year {year}',
        'count': updated_count
    })

# ==================== REQUEST IMMEDIATE ====================
@swagger_auto_schema(
    method='post',
    operation_description="Create vacation request immediately with optional file attachments",
    operation_summary="Create Immediate Request with Files",
    tags=['Vacation'],
    manual_parameters=[
        openapi.Parameter('requester_type', openapi.IN_FORM, type=openapi.TYPE_STRING, required=True),
        openapi.Parameter('employee_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False),
        openapi.Parameter('vacation_type_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=True),
        openapi.Parameter('start_date', openapi.IN_FORM, type=openapi.TYPE_STRING, format='date', required=True),
        openapi.Parameter('end_date', openapi.IN_FORM, type=openapi.TYPE_STRING, format='date', required=True),
        openapi.Parameter('comment', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('hr_representative_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False),
        openapi.Parameter('files', openapi.IN_FORM, type=openapi.TYPE_ARRAY, 
                         items=openapi.Items(type=openapi.TYPE_FILE), required=False,
                         description='Multiple files (Max 10MB each, PDF/JPG/PNG/DOC/DOCX/XLS/XLSX)'),
    ],
    responses={201: openapi.Response(description='Request created')}
)
@api_view(['POST'])
@has_vacation_permission('vacation.request.create_own')
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def create_immediate_request(request):
    """Create vacation request immediately with file attachments"""
    import json
    
    try:
        # Parse JSON fields from form data
        data = request.data.dict()
        
        # Parse employee_manual if exists
        if 'employee_manual' in data:
            try:
                data['employee_manual'] = json.loads(data['employee_manual'])
            except json.JSONDecodeError:
                return Response({
                    'error': 'Invalid employee_manual format. Must be valid JSON object.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get uploaded files
        uploaded_files = request.FILES.getlist('files')
        
        # Validate request data
        serializer = VacationRequestCreateSerializer(data=data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        validated_data = serializer.validated_data
        
        # ✅ FIX: Get requester employee first
        try:
            requester_emp = Employee.objects.get(user=request.user, is_deleted=False)
        except Employee.DoesNotExist:
            return Response({
                'error': 'Employee profile not found for current user'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # ✅ FIX: Determine target employee BEFORE using it
        if validated_data['requester_type'] == 'for_me':
            employee = requester_emp
        else:
            if validated_data.get('employee_id'):
                try:
                    employee = Employee.objects.get(id=validated_data['employee_id'], is_deleted=False)
                    # Check if requester is line manager
                    if employee.line_manager != requester_emp:
                        return Response({
                            'error': 'This employee is not in your team'
                        }, status=status.HTTP_403_FORBIDDEN)
                except Employee.DoesNotExist:
                    return Response({
                        'error': 'Employee not found'
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # Manual employee
                manual_data = validated_data.get('employee_manual', {})
                if not manual_data.get('name'):
                    return Response({
                        'error': 'Employee name is required'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                employee = Employee.objects.create(
                    full_name=manual_data.get('name', ''),
                    phone=manual_data.get('phone', ''),
                    line_manager=requester_emp,
                    created_by=request.user
                )
        
        # ✅ NOW employee is defined - check conflicts
        temp_request = VacationRequest(
            employee=employee,
            start_date=validated_data['start_date'],
            end_date=validated_data['end_date']
        )
        
        has_conflict, conflicts = temp_request.check_date_conflicts()
        if has_conflict:
            return Response({
                'error': 'Bu tarixlərdə artıq vacation mövcuddur',
                'conflicts': conflicts,
                'message': 'Zəhmət olmasa başqa tarix seçin'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Balance check
        settings = VacationSetting.get_active()
        year = date.today().year
        
        # Get or create balance
        balance, created = EmployeeVacationBalance.objects.get_or_create(
            employee=employee,
            year=year,
            defaults={
                'start_balance': 0,
                'yearly_balance': 28,
                'updated_by': request.user
            }
        )
        
        # Calculate working days
        working_days = 0
        if settings:
            working_days = settings.calculate_working_days(
                validated_data['start_date'], 
                validated_data['end_date']
            )
        
        # Check negative balance
        if settings and not settings.allow_negative_balance:
            if working_days > balance.remaining_balance:
                return Response({
                    'error': f'Insufficient balance. You have {balance.remaining_balance} days remaining. Negative balance not allowed.',
                    'available_balance': float(balance.remaining_balance),
                    'requested_days': working_days
                }, status=status.HTTP_400_BAD_REQUEST)
        
        with transaction.atomic():
            # Create vacation request
            vac_req = VacationRequest.objects.create(
                employee=employee,
                requester=request.user,
                request_type='IMMEDIATE',
                vacation_type_id=validated_data['vacation_type_id'],
                start_date=validated_data['start_date'],
                end_date=validated_data['end_date'],
                comment=validated_data.get('comment', ''),
                hr_representative_id=validated_data.get('hr_representative_id')
            )
            
            # Upload files if provided
            uploaded_attachments = []
            file_errors = []
            
            for file in uploaded_files:
                try:
                    # Validate file
                    from .business_trip_serializers import TripAttachmentUploadSerializer
                    upload_serializer = TripAttachmentUploadSerializer(data={'file': file})
                    if not upload_serializer.is_valid():
                        file_errors.append({
                            'filename': file.name,
                            'errors': upload_serializer.errors
                        })
                        continue
                    
                    # Create attachment using VacationAttachment model
                    from .vacation_models import VacationAttachment
                    attachment = VacationAttachment.objects.create(
                        vacation_request=vac_req,
                        file=file,
                        original_filename=file.name,
                        file_size=file.size,
                        file_type=file.content_type,
                        uploaded_by=request.user
                    )
                    uploaded_attachments.append(attachment)
                    
                except Exception as e:
                    file_errors.append({
                        'filename': file.name,
                        'error': str(e)
                    })
            
            # Submit request
            vac_req.submit_request(request.user)
            
            # ✅ Send notification
            graph_token = get_graph_access_token(request.user)
            notification_sent = False
            if graph_token:
                notification_sent = notification_manager.notify_request_created(vac_req, graph_token)
                if notification_sent:
                    logger.info("✅ Notification sent to Line Manager")
                else:
                    logger.warning("⚠️ Failed to send notification")
            else:
                logger.warning("⚠️ Graph token not available - notification skipped")
            
            # Refresh balance
            balance.refresh_from_db()
            
            # Prepare response
            response_data = {
                'message': 'Vacation request created and submitted successfully.',
                'notification_sent': notification_sent,
                'request': VacationRequestDetailSerializer(vac_req).data,
                'files_uploaded': len(uploaded_attachments),
                'files_failed': len(file_errors),
                'balance': {
                    'total_balance': float(balance.total_balance),
                    'yearly_balance': float(balance.yearly_balance),
                    'used_days': float(balance.used_days),
                    'remaining_balance': float(balance.remaining_balance),
                    'scheduled_days': float(balance.scheduled_days),
                    'should_be_planned': float(balance.should_be_planned)
                }
            }
            
            if uploaded_attachments:
                from .vacation_serializers import VacationAttachmentSerializer
                response_data['attachments'] = VacationAttachmentSerializer(
                    uploaded_attachments,
                    many=True,
                    context={'request': request}
                ).data
            
            if file_errors:
                response_data['file_errors'] = file_errors
            
            return Response(response_data, status=status.HTTP_201_CREATED)
    
    except Employee.DoesNotExist:
        return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error creating vacation request: {e}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== CREATE SCHEDULE ====================
@swagger_auto_schema(
    method='post',
    operation_description="Vacation Schedule yaratmaq (təsdiq tələb etmir)",
    operation_summary="Create Schedule",
    tags=['Vacation'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['requester_type', 'vacation_type_id', 'start_date', 'end_date'],
        properties={
            'requester_type': openapi.Schema(
                type=openapi.TYPE_STRING,
                enum=['for_me', 'for_my_employee'],
                description='Kimə görə schedule yaradırsan'
            ),
            'employee_id': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='İşçi ID (yalnız for_my_employee üçün)',
                example=123
            ),
            'employee_manual': openapi.Schema(
                type=openapi.TYPE_OBJECT,
                description='Manual employee məlumatları',
                properties={
                    'name': openapi.Schema(type=openapi.TYPE_STRING, example='John Doe'),
                    'phone': openapi.Schema(type=openapi.TYPE_STRING, example='+994501234567'),
                    'department': openapi.Schema(type=openapi.TYPE_STRING, example='IT')
                }
            ),
            'vacation_type_id': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Vacation növü ID',
                example=1
            ),
            'start_date': openapi.Schema(
                type=openapi.TYPE_STRING,
                format='date',
                description='Başlama tarixi',
                example='2025-11-10'
            ),
            'end_date': openapi.Schema(
                type=openapi.TYPE_STRING,
                format='date',
                description='Bitmə tarixi',
                example='2025-11-15'
            ),
            'comment': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Şərh',
                example='Planlı məzuniyyət'
            ),
        }
    ),
    responses={
        201: openapi.Response(
            description='Schedule yaradıldı',
            examples={
                'application/json': {
                    'message': 'Schedule yaradıldı',
                    'schedule': {
                        'id': 1,
                        'employee_name': 'John Doe',
                        'start_date': '2025-11-10',
                        'end_date': '2025-11-15',
                        'number_of_days': 4,
                        'status': 'SCHEDULED'
                    },
                    'conflicts': []
                }
            }
        )
    }
)
@api_view(['POST'])
@has_vacation_permission('vacation.schedule.create_own')
@permission_classes([IsAuthenticated])
def create_schedule(request):
    """Schedule yarat"""
    serializer = VacationScheduleCreateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    
    try:
        requester_emp = Employee.objects.get(user=request.user)
        year = date.today().year
        
        if data['requester_type'] == 'for_me':
            employee = requester_emp
        else:
            if data.get('employee_id'):
                employee = Employee.objects.get(id=data['employee_id'])
                if employee.line_manager != requester_emp:
                    return Response({
                        'error': 'Bu işçi sizin tabeliyinizdə deyil'
                    }, status=status.HTTP_403_FORBIDDEN)
            else:
                manual_data = data.get('employee_manual', {})
                if not manual_data.get('name'):
                    return Response({
                        'error': 'Employee adı mütləqdir'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                employee = Employee.objects.create(
                    full_name=manual_data.get('name', ''),
                    phone=manual_data.get('phone', ''),
                    line_manager=requester_emp,
                    created_by=request.user
                )
        temp_schedule = VacationSchedule(
            employee=employee,
            start_date=data['start_date'],
            end_date=data['end_date']
        )
        
        has_conflict, conflicts = temp_schedule.check_date_conflicts()
        if has_conflict:
            return Response({
                'error': 'Bu tarixlərdə artıq vacation mövcuddur',
                'conflicts': conflicts,
                'message': 'Zəhmət olmasa başqa tarix seçin'
            }, status=status.HTTP_400_BAD_REQUEST)
        # Balance tap və ya yarat
        balance, created = EmployeeVacationBalance.objects.get_or_create(
            employee=employee,
            year=year,
            defaults={
                'start_balance': 0,
                'yearly_balance': 28,
                'updated_by': request.user
            }
        )
        
        # Calculate working days
        settings = VacationSetting.get_active()
        working_days = 0
        if settings:
            working_days = settings.calculate_working_days(
                data['start_date'], 
                data['end_date']
            )
        
        # Balance yoxla (schedule üçün scheduled_days artır)
        if settings and not settings.allow_negative_balance:
            total_planned = balance.scheduled_days + working_days
            if total_planned > balance.total_balance:
                return Response({
                    'error': f'Qeyri-kafi balans. Toplam planlaşdırılmış ({total_planned}) günlər balansı ({balance.total_balance}) keçir.',
                    'available_balance': float(balance.remaining_balance),
                    'requested_days': working_days,
                    'current_scheduled': float(balance.scheduled_days)
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Conflicting schedules tap
        conflicts = VacationSchedule.objects.filter(
            employee__in=Employee.objects.filter(
                Q(department=employee.department) | Q(line_manager=employee.line_manager)
            ).exclude(id=employee.id) if employee.id else Employee.objects.none(),
            start_date__lte=data['end_date'],
            end_date__gte=data['start_date'],
            status='SCHEDULED',
            is_deleted=False
        )
        
        with transaction.atomic():
            schedule = VacationSchedule.objects.create(
                employee=employee,
                vacation_type_id=data['vacation_type_id'],
                start_date=data['start_date'],
                end_date=data['end_date'],
                comment=data.get('comment', ''),
                created_by=request.user
            )
            
            # Refresh balance
            balance.refresh_from_db()
            
            return Response({
                'message': 'Schedule yaradıldı',
                'schedule': VacationScheduleSerializer(schedule).data,
                'conflicts': VacationScheduleSerializer(conflicts, many=True).data,
                'balance': {
                    'total_balance': float(balance.total_balance),
                    'yearly_balance': float(balance.yearly_balance),
                    'used_days': float(balance.used_days),
                    'remaining_balance': float(balance.remaining_balance),
                    'scheduled_days': float(balance.scheduled_days),
                    'should_be_planned': float(balance.should_be_planned)
                }
            }, status=status.HTTP_201_CREATED)
    
    except Employee.DoesNotExist:
        return Response({
            'error': 'Employee tapılmadı'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)

# ==================== MY SCHEDULE TABS ====================
@swagger_auto_schema(
    method='get',
    operation_description="Schedule tabları - upcoming, peers, all",
    operation_summary="My Schedule Tabs",
    tags=['Vacation'],
    responses={
        200: openapi.Response(
            description='Schedule tabs data',
            examples={
                'application/json': {
                    'upcoming': [],
                    'peers': [],
                    'all': []
                }
            }
        )
    }
)
@api_view(['GET'])
@has_vacation_permission('vacation.schedule.view_own')
@permission_classes([IsAuthenticated])
def my_schedule_tabs(request):
    """Schedule tabları - upcoming, peers, all"""
    try:
        emp = Employee.objects.get(user=request.user)
        
        # Upcoming
        upcoming = VacationSchedule.objects.filter(
            employee=emp,
            start_date__gte=date.today(),
            status='SCHEDULED',
            is_deleted=False
        )
        
        # Peers
        peers = Employee.objects.filter(
            Q(department=emp.department) | Q(line_manager=emp.line_manager),
            is_deleted=False
        ).exclude(id=emp.id)
        
        peers_schedules = VacationSchedule.objects.filter(
            employee__in=peers,
            start_date__gte=date.today(),
            status='SCHEDULED',
            is_deleted=False
        )
        
        # All
        all_schedules = VacationSchedule.objects.filter(employee=emp, is_deleted=False)
        
        return Response({
            'upcoming': VacationScheduleSerializer(upcoming, many=True).data,
            'peers': VacationScheduleSerializer(peers_schedules, many=True).data,
            'all': VacationScheduleSerializer(all_schedules, many=True).data
        })
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profili tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ==================== REGISTER SCHEDULE ====================
@swagger_auto_schema(
    method='post',
    operation_description="Register schedule as taken with notification",
    operation_summary="Register Schedule",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Schedule registered')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@has_vacation_permission('vacation.schedule.register')
def register_schedule(request, pk):
    """Register schedule as taken with notification"""
    try:
        schedule = VacationSchedule.objects.get(pk=pk, is_deleted=False)
        
        # Register schedule
        schedule.register_as_taken(request.user)
        
        # ✅ Send notification
        graph_token = get_graph_access_token(request.user)
        notification_sent = False
        if graph_token:
            notification_sent = notification_manager.notify_schedule_registered(schedule, graph_token)
            if notification_sent:
                logger.info("✅ Schedule registered notification sent")
        
        # Refresh balance
        balance = EmployeeVacationBalance.objects.get(
            employee=schedule.employee,
            year=schedule.start_date.year
        )
        
        return Response({
            'message': 'Schedule registered successfully',
            'notification_sent': notification_sent,
            'schedule': VacationScheduleSerializer(schedule).data,
            'updated_balance': {
                'total_balance': float(balance.total_balance),
                'yearly_balance': float(balance.yearly_balance),
                'used_days': float(balance.used_days),
                'scheduled_days': float(balance.scheduled_days),
                'remaining_balance': float(balance.remaining_balance),
                'should_be_planned': float(balance.should_be_planned)
            }
        })
    except VacationSchedule.DoesNotExist:
        return Response({'error': 'Schedule not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
# ==================== EDIT SCHEDULE ====================
@swagger_auto_schema(
    method='put',
    operation_description="Schedule-i edit et",
    operation_summary="Edit Schedule",
    tags=['Vacation'],
    manual_parameters=[
        openapi.Parameter(
            'id',
            openapi.IN_PATH,
            description="Schedule ID",
            type=openapi.TYPE_INTEGER,
            required=True
        )
    ],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'vacation_type_id': openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description='Vacation növü ID',
                example=1
            ),
            'start_date': openapi.Schema(
                type=openapi.TYPE_STRING,
                format='date',
                description='Başlama tarixi (YYYY-MM-DD)',
                example='2025-11-10'
            ),
            'end_date': openapi.Schema(
                type=openapi.TYPE_STRING,
                format='date',
                description='Bitmə tarixi (YYYY-MM-DD)',
                example='2025-11-15'
            ),
            'comment': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Şərh (optional)',
                example='Dəyişiklik edildi'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Schedule yeniləndi',
            examples={
                'application/json': {
                    'message': 'Schedule yeniləndi',
                    'schedule': {
                        'id': 1,
                        'vacation_type_name': 'Annual Leave',
                        'start_date': '2025-11-10',
                        'end_date': '2025-11-15',
                        'number_of_days': 4,
                        'edit_count': 1,
                        'can_edit': True
                    }
                }
            }
        ),
        400: openapi.Response(description='Validation error'),
        403: openapi.Response(description='Permission denied'),
        404: openapi.Response(description='Schedule tapılmadı')
    }
)
@api_view(['PUT'])
@has_vacation_permission('vacation.schedule.update_own')
@permission_classes([IsAuthenticated])
def edit_schedule(request, pk):
    """Schedule-i edit et"""
    try:
        schedule = VacationSchedule.objects.get(pk=pk, is_deleted=False)
        emp = Employee.objects.get(user=request.user)
        
        # Yalnız schedule-in sahibi edit edə bilər
        if schedule.employee != emp:
            return Response({
                'error': 'Bu schedule-i edit etmək hüququnuz yoxdur'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Edit limiti yoxla
        if not schedule.can_edit():
            return Response({
                'error': 'Bu schedule-i daha edit edə bilməzsiniz'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Update fields
        if 'vacation_type_id' in request.data:
            schedule.vacation_type_id = request.data['vacation_type_id']
        if 'start_date' in request.data:
            schedule.start_date = request.data['start_date']
        if 'end_date' in request.data:
            schedule.end_date = request.data['end_date']
        if 'comment' in request.data:
            schedule.comment = request.data['comment']
        
        has_conflict, conflicts = schedule.check_date_conflicts()
        if has_conflict:
            return Response({
                'error': 'Bu tarixlərdə artıq vacation mövcuddur',
                'conflicts': conflicts,
                'message': 'Zəhmət olmasa başqa tarix seçin'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        schedule.edit_count += 1
        schedule.last_edited_at = timezone.now()
        schedule.last_edited_by = request.user
        schedule.save()
        
        return Response({
            'message': 'Schedule yeniləndi',
            'schedule': VacationScheduleSerializer(schedule).data
        })
        
    except VacationSchedule.DoesNotExist:
        return Response({'error': 'Schedule tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profili tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== DELETE SCHEDULE ====================
@swagger_auto_schema(
    method='delete',
    operation_description="Schedule-i sil",
    operation_summary="Delete Schedule",
    tags=['Vacation'],
    manual_parameters=[
        openapi.Parameter(
            'id',
            openapi.IN_PATH,
            description="Schedule ID",
            type=openapi.TYPE_INTEGER,
            required=True
        )
    ],
    responses={
        200: openapi.Response(description='Schedule silindi')
    }
)
@api_view(['DELETE'])
@has_vacation_permission('vacation.schedule.delete_own')
@permission_classes([IsAuthenticated])
def delete_schedule(request, pk):
    """Schedule-i sil"""
    try:
        schedule = VacationSchedule.objects.get(pk=pk, is_deleted=False)
        emp = Employee.objects.get(user=request.user)
        
        # Yalnız schedule-in sahibi və ya line manager silə bilər
        if schedule.employee != emp and schedule.employee.line_manager != emp:
            return Response({'error': 'Bu schedule-i silmək hüququnuz yoxdur'}, status=status.HTTP_403_FORBIDDEN)
        
        # Registered schedule-i silmək olmaz
        if schedule.status == 'REGISTERED':
            return Response({'error': 'Registered schedule-i silmək olmaz'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Soft delete
        schedule.is_deleted = True
        schedule.deleted_by = request.user
        schedule.deleted_at = timezone.now()
        schedule.save()
        
        # Scheduled days balansını azalt
        schedule._update_scheduled_balance(add=False)
        
        return Response({'message': 'Schedule silindi'})
        
    except VacationSchedule.DoesNotExist:
        return Response({'error': 'Schedule tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profili tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ==================== APPROVAL PENDING ====================
@swagger_auto_schema(
    method='get',
    operation_description="Approval - Pending requestlər (Admin bütün pending-ləri görür)",
    operation_summary="Pending Requests",
    tags=['Vacation'],
    responses={
        200: openapi.Response(
            description='Pending requests',
            examples={
                'application/json': {
                    'line_manager_requests': [],
                    'hr_requests': [],
                    'total_pending': 0
                }
            }
        )
    }
)
@api_view(['GET'])
@has_any_vacation_permission([
    'vacation.request.approve_as_line_manager',
    'vacation.request.approve_as_hr'
])
@permission_classes([IsAuthenticated])
def approval_pending_requests(request):
    """Approval - Pending requestlər"""
    from .vacation_permissions import is_admin_user
    
    try:
        emp = Employee.objects.get(user=request.user)
        
        # Admin isə, BÜTÜN pending request-ləri göstər
        if is_admin_user(request.user):
            lm_requests = VacationRequest.objects.filter(
                status='PENDING_LINE_MANAGER',
                is_deleted=False
            ).order_by('-created_at')
            
            hr_requests = VacationRequest.objects.filter(
                status='PENDING_HR',
                is_deleted=False
            ).order_by('-created_at')
        else:
            # Normal user - yalnız onun line manager və ya HR olduğu requestlər
            lm_requests = VacationRequest.objects.filter(
                line_manager=emp,
                status='PENDING_LINE_MANAGER',
                is_deleted=False
            ).order_by('-created_at')
            
            hr_requests = VacationRequest.objects.filter(
                hr_representative=emp,
                status='PENDING_HR',
                is_deleted=False
            ).order_by('-created_at')
        
        return Response({
            'line_manager_requests': VacationRequestListSerializer(lm_requests, many=True).data,
            'hr_requests': VacationRequestListSerializer(hr_requests, many=True).data,
            'total_pending': lm_requests.count() + hr_requests.count(),
            'is_admin': is_admin_user(request.user)  # Frontend üçün
        })
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profili tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== APPROVAL HISTORY ====================
@swagger_auto_schema(
    method='get',
    operation_description="Approval History",
    operation_summary="Approval History",
    tags=['Vacation'],
    responses={
        200: openapi.Response(
            description='Approval history',
            examples={
                'application/json': {
                    'history': []
                }
            }
        )
    }
)
@api_view(['GET'])
@has_any_vacation_permission([
    'vacation.request.approve_as_line_manager',
    'vacation.request.approve_as_hr'
])
@permission_classes([IsAuthenticated])
def approval_history(request):
    """Approval History"""
    try:
        # Line Manager kimi təsdiq etdiklərim
        lm_approved = VacationRequest.objects.filter(
            line_manager_approved_by=request.user,
            is_deleted=False
        ).order_by('-line_manager_approved_at')[:20]
        
        # HR kimi təsdiq etdiklərim
        hr_approved = VacationRequest.objects.filter(
            hr_approved_by=request.user,
            is_deleted=False
        ).order_by('-hr_approved_at')[:20]
        
        # Reject etdiklərim
        rejected = VacationRequest.objects.filter(
            rejected_by=request.user,
            is_deleted=False
        ).order_by('-rejected_at')[:20]
        
        history = []
        
        for req in lm_approved:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'days': float(req.number_of_days),
                'status': 'Approved (Line Manager)',
                'action': 'Approved',
                'comment': req.line_manager_comment,
                'date': req.line_manager_approved_at
            })
        
        for req in hr_approved:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'days': float(req.number_of_days),
                'status': 'Approved (HR)',
                'action': 'Approved',
                'comment': req.hr_comment,
                'date': req.hr_approved_at
            })
        
        for req in rejected:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'days': float(req.number_of_days),
                'status': req.get_status_display(),
                'action': 'Rejected',
                'comment': req.rejection_reason,
                'date': req.rejected_at
            })
        
        history.sort(key=lambda x: x['date'] if x['date'] else datetime.min, reverse=True)
        
        return Response({'history': history[:20]})
    except Exception as e:
        return Response({'error': f'History yüklənmədi: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


# ==================== APPROVE/REJECT REQUEST ====================
@swagger_auto_schema(
    method='post',
    operation_description="Approve/Reject vacation request with email notifications",
    tags=['Vacation'],
    request_body=VacationApprovalSerializer,
    responses={200: openapi.Response(description='Action completed')}
)
@api_view(['POST'])
@has_any_vacation_permission([
    'vacation.request.approve_as_line_manager',
    'vacation.request.approve_as_hr'
])
@permission_classes([IsAuthenticated])
def approve_reject_request(request, pk):
    """Approve/Reject vacation request with email notifications"""
    serializer = VacationApprovalSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    
    # ✅ Get notification context
    notification_ctx = get_notification_context(request)
    graph_token = notification_ctx['graph_token']
    
    try:
        vac_req = VacationRequest.objects.get(pk=pk, is_deleted=False)
        notification_sent = False
        
        # LINE MANAGER APPROVAL/REJECTION
        if vac_req.status == 'PENDING_LINE_MANAGER':
            if data['action'] == 'approve':
                vac_req.approve_by_line_manager(request.user, data.get('comment', ''))
                msg = 'Approved by Line Manager'
                
                # ✅ Send notification to HR
                if graph_token:
                    try:
                        notification_sent = notification_manager.notify_line_manager_approved(
                            vacation_request=vac_req,
                            access_token=graph_token
                        )
                        if notification_sent:
                            logger.info("✅ Notification sent to HR")
                    except Exception as e:
                        logger.error(f"❌ Notification error: {e}")
            else:
                vac_req.reject_by_line_manager(request.user, data.get('reason', ''))
                msg = 'Rejected by Line Manager'
                
                # ✅ Send rejection notification to Employee
                if graph_token:
                    try:
                        notification_sent = notification_manager.notify_request_rejected(
                            vacation_request=vac_req,
                            access_token=graph_token
                        )
                        if notification_sent:
                            logger.info("✅ Rejection notification sent to Employee")
                    except Exception as e:
                        logger.error(f"❌ Notification error: {e}")
        
        # HR APPROVAL/REJECTION
        elif vac_req.status == 'PENDING_HR':
            if data['action'] == 'approve':
                vac_req.approve_by_hr(request.user, data.get('comment', ''))
                msg = 'Approved by HR - Request is now APPROVED'
                
                # ✅ Send final approval notification to Employee
                if graph_token:
                    try:
                        notification_sent = notification_manager.notify_hr_approved(
                            vacation_request=vac_req,
                            access_token=graph_token
                        )
                        if notification_sent:
                            logger.info("✅ Final approval notification sent to Employee")
                    except Exception as e:
                        logger.error(f"❌ Notification error: {e}")
            else:
                vac_req.reject_by_hr(request.user, data.get('reason', ''))
                msg = 'Rejected by HR'
                
                # ✅ Send rejection notification to Employee
                if graph_token:
                    try:
                        notification_sent = notification_manager.notify_request_rejected(
                            vacation_request=vac_req,
                            access_token=graph_token
                        )
                        if notification_sent:
                            logger.info("✅ Rejection notification sent to Employee")
                    except Exception as e:
                        logger.error(f"❌ Notification error: {e}")
        else:
            return Response({
                'error': 'Request is not pending approval'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'message': msg,
            'notification_sent': notification_sent,
            'notification_available': notification_ctx['can_send_emails'],
            'notification_reason': notification_ctx['reason'],
            'request': VacationRequestDetailSerializer(vac_req).data
        })
    
    except VacationRequest.DoesNotExist:
        return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error in approve/reject: {e}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== MY ALL REQUESTS & SCHEDULES ====================
@swagger_auto_schema(
    method='get',
    operation_description="İstifadəçinin bütün vacation request və schedule-lərini göstər",
    operation_summary="My All Requests & Schedules",
    tags=['Vacation'],
    responses={
        200: openapi.Response(
            description='Bütün records',
            examples={
                'application/json': {
                    'records': [
                        {
                            'id': 1,
                            'type': 'request',
                            'request_id': 'VR20250001',
                            'vacation_type': 'Annual Leave',
                            'start_date': '2025-10-15',
                            'end_date': '2025-10-20',
                            'days': 4.0,
                            'status': 'Approved',
                            'attachments_count': 2,  # ✅ Added to example
                            'has_attachments': True,  # ✅ Added to example
                            'created_at': '2025-09-01T10:00:00Z'
                        }
                    ]
                }
            }
        )
    }
)
@api_view(['GET'])
@has_vacation_permission('vacation.request.view_own')
@permission_classes([IsAuthenticated])
def my_all_requests_schedules(request):
    """My All Requests & Schedules"""
    from rest_framework import status as http_status
    
    try:
        emp = Employee.objects.get(user=request.user, is_deleted=False)
        
        # All requests - ✅ Add prefetch for attachments
        requests = VacationRequest.objects.filter(
            employee=emp, 
            is_deleted=False
        ).select_related(
            'vacation_type',
            'line_manager',
            'hr_representative',
            'line_manager_approved_by',
            'hr_approved_by',
            'rejected_by'
        ).prefetch_related('attachments').order_by('-created_at')  # ✅ Added prefetch
        
        # All schedules
        schedules = VacationSchedule.objects.filter(
            employee=emp, 
            is_deleted=False
        ).select_related(
            'vacation_type',
            'created_by',
            'last_edited_by'
        ).order_by('-start_date')
        
        # Combine
        combined = []
        
        for req in requests:
            # ✅ Count attachments
            attachments_count = req.attachments.filter(is_deleted=False).count()
            
            combined.append({
                'id': req.id,
                'type': 'request',
                'request_id': req.request_id,
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'return_date': req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                'days': float(req.number_of_days),
                'status': req.get_status_display(),
                'status_code': req.status,
                'comment': req.comment,
                'attachments_count': attachments_count,  # ✅ Added
                'has_attachments': attachments_count > 0,  # ✅ Added
                'line_manager': req.line_manager.full_name if req.line_manager else '',
                'hr_representative': req.hr_representative.full_name if req.hr_representative else '',
                'line_manager_comment': req.line_manager_comment,
                'hr_comment': req.hr_comment,
                'rejection_reason': req.rejection_reason,
                'created_at': req.created_at.isoformat() if req.created_at else None
            })
        
        for sch in schedules:
            combined.append({
                'id': sch.id,
                'type': 'schedule',
                'request_id': f'SCH{sch.id}',
                'vacation_type': sch.vacation_type.name,
                'start_date': sch.start_date.strftime('%Y-%m-%d'),
                'end_date': sch.end_date.strftime('%Y-%m-%d'),
                'return_date': sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                'days': float(sch.number_of_days),
                'status': sch.get_status_display(),
                'status_code': sch.status,
                'comment': sch.comment,
                'attachments_count': 0,  # ✅ Schedules don't have attachments
                'has_attachments': False,  # ✅ Added
                'created_at': sch.created_at.isoformat() if sch.created_at else None,
                'can_edit': sch.can_edit(),
                'edit_count': sch.edit_count,
                'created_by': sch.created_by.get_full_name() if sch.created_by else '',
                'last_edited_by': sch.last_edited_by.get_full_name() if sch.last_edited_by else '',
                'last_edited_at': sch.last_edited_at.isoformat() if sch.last_edited_at else None
            })
        
        combined.sort(key=lambda x: x['created_at'] if x['created_at'] else '', reverse=True)
        
        return Response({
            'records': combined,
            'total_count': len(combined),
            'requests_count': requests.count(),
            'schedules_count': schedules.count()
        })
        
    except Employee.DoesNotExist:
        return Response({
            'error': 'Employee profili tapılmadı'
        }, status=http_status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error in my_all_requests_schedules: {e}")
        return Response({
            'error': str(e)
        }, status=http_status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='get',
    operation_description="Get detailed information of a vacation request including attachments and approval history",
    operation_summary="Get Vacation Request Detail",
    tags=['Vacation'],
    responses={
        200: openapi.Response(
            description='Vacation request details',
            schema=VacationRequestDetailSerializer
        ),
        403: 'Permission denied',
        404: 'Request not found'
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_vacation_request_detail(request, pk):
    """
    Get detailed information of a vacation request
    
    Shows:
    - Basic vacation information
    - Employee details
    - Vacation configuration (type)
    - File attachments
    - Approval workflow status
    - Comments from all approvers
    - Rejection reasons (if applicable)
    """
    try:
        from .vacation_permissions import is_admin_user
        
        # Get the vacation request
        vac_req = VacationRequest.objects.select_related(
            'employee', 
            'employee__department',
            'employee__business_function',
            'employee__unit',
            'employee__job_function',
            'vacation_type',
            'line_manager',
            'hr_representative',
            'requester'
        ).prefetch_related(
            'attachments'
        ).get(pk=pk, is_deleted=False)
        
        # Check access permission
        emp = None
        try:
            emp = Employee.objects.get(user=request.user, is_deleted=False)
        except Employee.DoesNotExist:
            pass
        
        # Determine if user can view this request
        can_view = False
        
        # Admin can view all
        if is_admin_user(request.user):
            can_view = True
        
        # Employee can view their own requests
        elif emp and vac_req.employee == emp:
            can_view = True
        
        # Requester can view requests they created
        elif vac_req.requester == request.user:
            can_view = True
        
        # Approvers can view requests assigned to them
        elif emp and (
            vac_req.line_manager == emp or 
            vac_req.hr_representative == emp
        ):
            can_view = True
        
        # Check if user has view_all permission
        elif check_vacation_permission(request.user, 'vacation.request.view_all')[0]:
            can_view = True
        
        if not can_view:
            return Response({
                'error': 'Permission denied',
                'detail': 'You do not have permission to view this vacation request'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Serialize the data
        serializer = VacationRequestDetailSerializer(
            vac_req, 
            context={'request': request}
        )
        
        # Add extra context information
        response_data = serializer.data
        
        # Add approval workflow status
        response_data['workflow'] = {
            'current_step': vac_req.status,
            'steps': [
                {
                    'name': 'Line Manager Approval',
                    'status': 'completed' if vac_req.line_manager_approved_at else (
                        'rejected' if vac_req.status == 'REJECTED_LINE_MANAGER' else (
                            'pending' if vac_req.status == 'PENDING_LINE_MANAGER' else 'not_started'
                        )
                    ),
                    'approver': vac_req.line_manager.full_name if vac_req.line_manager else None,
                    'approved_at': vac_req.line_manager_approved_at,
                    'comment': vac_req.line_manager_comment
                },
                {
                    'name': 'HR Processing',
                    'status': 'completed' if vac_req.hr_approved_at else (
                        'rejected' if vac_req.status == 'REJECTED_HR' else (
                            'pending' if vac_req.status == 'PENDING_HR' else 'not_started'
                        )
                    ),
                    'approver': vac_req.hr_representative.full_name if vac_req.hr_representative else None,
                    'approved_at': vac_req.hr_approved_at,
                    'comment': vac_req.hr_comment
                }
            ]
        }
        
        # Add requester information
        response_data['requester_info'] = {
            'type': vac_req.get_request_type_display(),
            'name': vac_req.requester.get_full_name() if vac_req.requester else None,
            'email': vac_req.requester.email if vac_req.requester else None
        }
        
        # Add permission flags for frontend
        response_data['permissions'] = {
            'can_approve': (
                (vac_req.status == 'PENDING_LINE_MANAGER' and emp and vac_req.line_manager == emp) or
                (vac_req.status == 'PENDING_HR' and emp and vac_req.hr_representative == emp) or
                is_admin_user(request.user)
            ),
            'is_admin': is_admin_user(request.user)
        }
        
        # Add attachments
        from .vacation_models import VacationAttachment
        from .vacation_serializers import VacationAttachmentSerializer
        attachments = vac_req.attachments.filter(is_deleted=False).order_by('-uploaded_at')
        response_data['attachments'] = VacationAttachmentSerializer(
            attachments,
            many=True,
            context={'request': request}
        ).data
        
        # Add summary statistics
        response_data['summary'] = {
            'total_attachments': attachments.count(),
        }
        
        return Response(response_data)
        
    except VacationRequest.DoesNotExist:
        return Response({
            'error': 'Vacation request not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error fetching vacation request detail: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)
                
@swagger_auto_schema(
    method='get',
    operation_description="Bütün vacation request və schedule-ları göstər (hamı görə bilər)",
    operation_summary="All Vacation Records",
    tags=['Vacation'],
    manual_parameters=[
        openapi.Parameter(
            'status',
            openapi.IN_QUERY,
            description="Status filteri",
            type=openapi.TYPE_STRING,
            required=False
        ),
        openapi.Parameter(
            'vacation_type_id',
            openapi.IN_QUERY,
            description="Vacation type filteri",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        openapi.Parameter(
            'department_id',
            openapi.IN_QUERY,
            description="Department filteri",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        # ✅ YENİ: Business Function filter
        openapi.Parameter(
            'business_function_id',
            openapi.IN_QUERY,
            description="Business Function (Company) filteri",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        openapi.Parameter(
            'start_date',
            openapi.IN_QUERY,
            description="Başlama tarixi filteri (YYYY-MM-DD)",
            type=openapi.TYPE_STRING,
            required=False
        ),
        openapi.Parameter(
            'end_date',
            openapi.IN_QUERY,
            description="Bitmə tarixi filteri (YYYY-MM-DD)",
            type=openapi.TYPE_STRING,
            required=False
        ),
        openapi.Parameter(
            'employee_name',
            openapi.IN_QUERY,
            description="Employee adı filteri",
            type=openapi.TYPE_STRING,
            required=False
        ),
        openapi.Parameter(
            'year',
            openapi.IN_QUERY,
            description="İl filteri (məsələn: 2025)",
            type=openapi.TYPE_INTEGER,
            required=False
        )
    ],
    responses={
        200: openapi.Response(
            description='Bütün vacation records',
            examples={
                'application/json': {
                    'records': [
                        {
                            'id': 1,
                            'type': 'request',
                            'request_id': 'VR20250001',
                            'employee_name': 'John Doe',
                            'employee_id': 'EMP001',
                            'department': 'IT',
                            'business_function': 'Almet Trading',
                            'vacation_type': 'Annual Leave',
                            'start_date': '2025-10-15',
                            'end_date': '2025-10-20',
                            'days': 4.0,
                            'status': 'Approved',
                            'created_at': '2025-09-01T10:00:00Z'
                        }
                    ],
                    'total_count': 150,
                    'requests_count': 75,
                    'schedules_count': 75
                }
            }
        )
    }
)
@api_view(['GET'])
@has_vacation_permission('vacation.request.view_all')
@permission_classes([IsAuthenticated])
def all_vacation_records(request):
    """Bütün vacation records-u JSON formatında list qaytarır"""
    from rest_framework import status as http_status
    
    try:
        # Filter parameters
        status_filter = request.GET.get('status')
        vacation_type_id = request.GET.get('vacation_type_id')
        department_id = request.GET.get('department_id')
        business_function_id = request.GET.get('business_function_id')  # ✅ YENİ
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        employee_name = request.GET.get('employee_name')
        year = request.GET.get('year')
        
        # All requests - ✅ Add prefetch for attachments
        requests_qs = VacationRequest.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'employee__business_function', 
            'vacation_type', 'line_manager', 'hr_representative',
            'line_manager_approved_by', 'hr_approved_by', 'rejected_by'
        ).prefetch_related('attachments')
        
        # All schedules  
        schedules_qs = VacationSchedule.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'employee__business_function', 
            'vacation_type', 'created_by', 'last_edited_by'
        )
        
        # Apply filters (same as before)
        if status_filter:
            requests_qs = requests_qs.filter(status=status_filter)
            schedules_qs = schedules_qs.filter(status=status_filter)
        
        if vacation_type_id:
            requests_qs = requests_qs.filter(vacation_type_id=vacation_type_id)
            schedules_qs = schedules_qs.filter(vacation_type_id=vacation_type_id)
        
        if department_id:
            requests_qs = requests_qs.filter(employee__department_id=department_id)
            schedules_qs = schedules_qs.filter(employee__department_id=department_id)
        
        # ✅ YENİ: Business Function filter
        if business_function_id:
            requests_qs = requests_qs.filter(employee__business_function_id=business_function_id)
            schedules_qs = schedules_qs.filter(employee__business_function_id=business_function_id)
        
        if start_date:
            requests_qs = requests_qs.filter(start_date__gte=start_date)
            schedules_qs = schedules_qs.filter(start_date__gte=start_date)
        
        if end_date:
            requests_qs = requests_qs.filter(end_date__lte=end_date)
            schedules_qs = schedules_qs.filter(end_date__lte=end_date)
        
        if employee_name:
            requests_qs = requests_qs.filter(employee__full_name__icontains=employee_name)
            schedules_qs = schedules_qs.filter(employee__full_name__icontains=employee_name)
        
        if year:
            requests_qs = requests_qs.filter(start_date__year=year)
            schedules_qs = schedules_qs.filter(start_date__year=year)
        
        # Get data
        requests = requests_qs.order_by('-created_at')
        schedules = schedules_qs.order_by('-created_at')
        
        # Combine all records
        all_records = []
        
        # Add requests
        for req in requests:
            attachments_count = req.attachments.filter(is_deleted=False).count()
            
            all_records.append({
                'id': req.id,
                'type': 'request',
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'employee_id': getattr(req.employee, 'employee_id', ''),
                'department': req.employee.department.name if req.employee.department else '',
                'business_function': req.employee.business_function.name if req.employee.business_function else '',
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'return_date': req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                'days': float(req.number_of_days),
                'status': req.get_status_display(),
                'status_code': req.status,
                'comment': req.comment,
                'line_manager': req.line_manager.full_name if req.line_manager else '',
                'hr_representative': req.hr_representative.full_name if req.hr_representative else '',
                'attachments_count': attachments_count,
                'has_attachments': attachments_count > 0,
                'created_at': req.created_at.isoformat() if req.created_at else None,
                'updated_at': req.updated_at.isoformat() if req.updated_at else None
            })
        
        # Add schedules
        for sch in schedules:
            all_records.append({
                'id': sch.id,
                'type': 'schedule',
                'request_id': f'SCH{sch.id}',
                'employee_name': sch.employee.full_name,
                'employee_id': getattr(sch.employee, 'employee_id', ''),
                'department': sch.employee.department.name if sch.employee.department else '',
                'business_function': sch.employee.business_function.name if sch.employee.business_function else '',
                'vacation_type': sch.vacation_type.name,
                'start_date': sch.start_date.strftime('%Y-%m-%d'),
                'end_date': sch.end_date.strftime('%Y-%m-%d'),
                'return_date': sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                'days': float(sch.number_of_days),
                'status': sch.get_status_display(),
                'status_code': sch.status,
                'comment': sch.comment,
                'can_edit': sch.can_edit(),
                'edit_count': sch.edit_count,
                'created_by': sch.created_by.get_full_name() if sch.created_by else '',
                'attachments_count': 0,
                'has_attachments': False,
                'created_at': sch.created_at.isoformat() if sch.created_at else None,
                'updated_at': sch.updated_at.isoformat() if sch.updated_at else None
            })
        
        # Sort by created_at desc
        all_records.sort(key=lambda x: x['created_at'] if x['created_at'] else '', reverse=True)
        
        return Response({
            'records': all_records,
            'total_count': len(all_records),
            'requests_count': requests.count(),
            'schedules_count': schedules.count(),
            'filters_applied': {
                'status': status_filter,
                'vacation_type_id': vacation_type_id,
                'department_id': department_id,
                'business_function_id': business_function_id,  # ✅ YENİ
                'start_date': start_date,
                'end_date': end_date,
                'employee_name': employee_name,
                'year': year
            }
        })
        
    except Exception as e:
        logger.error(f"Error in all_vacation_records: {e}")
        return Response({
            'error': f'Error retrieving records: {str(e)}'
        }, status=http_status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='get',
    operation_description="Bütün vacation records-u Excel formatında export et (filterlər dəstəklənir)",
    operation_summary="Export All Vacation Records",
    tags=['Vacation'],
    manual_parameters=[
        openapi.Parameter(
            'status',
            openapi.IN_QUERY,
            description="Status filteri",
            type=openapi.TYPE_STRING,
            required=False
        ),
        openapi.Parameter(
            'vacation_type_id',
            openapi.IN_QUERY,
            description="Vacation type filteri",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        openapi.Parameter(
            'department_id',
            openapi.IN_QUERY,
            description="Department filteri",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        openapi.Parameter(
            'start_date',
            openapi.IN_QUERY,
            description="Başlama tarixi filteri (YYYY-MM-DD)",
            type=openapi.TYPE_STRING,
            required=False
        ),
        openapi.Parameter(
            'end_date',
            openapi.IN_QUERY,
            description="Bitmə tarixi filteri (YYYY-MM-DD)",
            type=openapi.TYPE_STRING,
            required=False
        ),
        openapi.Parameter(
            'employee_name',
            openapi.IN_QUERY,
            description="Employee adı filteri",
            type=openapi.TYPE_STRING,
            required=False
        ),
        openapi.Parameter(
            'year',
            openapi.IN_QUERY,
            description="İl filteri (məsələn: 2025)",
            type=openapi.TYPE_INTEGER,
            required=False
        ),
        openapi.Parameter(
            'format',
            openapi.IN_QUERY,
            description="Export formatı: 'combined' (hər şey bir sheet-də) və ya 'separated' (ayrı sheet-lər)",
            type=openapi.TYPE_STRING,
            enum=['combined', 'separated'],
            required=False
        )
    ],
    responses={
        200: openapi.Response(
            description='Excel file',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    }
)
@api_view(['GET'])
@has_any_vacation_permission([
    'vacation.request.export_all',  # HR və Admin
    'vacation.request.export_team'   # Line Manager (öz team-i)
])
@permission_classes([IsAuthenticated])
def export_all_vacation_records(request):
    """Bütün vacation records-u enhanced formatda export et"""
    from .vacation_permissions import is_admin_user
    
    try:
        # Filter parameters
        status_filter = request.GET.get('status')
        vacation_type_id = request.GET.get('vacation_type_id')
        department_id = request.GET.get('department_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        employee_name = request.GET.get('employee_name')
        year = request.GET.get('year')
        export_format = request.GET.get('format', 'dashboard')
        include_charts = request.GET.get('include_charts', 'true').lower() == 'true'
        business_function_id = request.GET.get('business_function_id') 
        # ✅ Permission-based filtering
        is_admin = is_admin_user(request.user)
        has_hr_permission, _ = check_vacation_permission(request.user, 'vacation.request.export_all')
        has_team_permission, _ = check_vacation_permission(request.user, 'vacation.request.export_team')
        
        # All requests base query
        requests_qs = VacationRequest.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'employee__business_function', 
            'vacation_type', 'line_manager', 'hr_representative',
            'line_manager_approved_by', 'hr_approved_by', 'rejected_by'
        )
        
        # All schedules base query
        schedules_qs = VacationSchedule.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'employee__business_function', 
            'vacation_type', 'created_by', 'last_edited_by'
        )
        
        # ✅ PERMISSION FILTERING
        if is_admin or has_hr_permission:
            # Admin və HR - bütün data
            pass
        elif has_team_permission:
            # Line Manager - yalnız öz team-i və özü
            try:
                emp = Employee.objects.get(user=request.user, is_deleted=False)
                
                # Öz team-indəki işçilər
                team_employees = Employee.objects.filter(
                    line_manager=emp,
                    is_deleted=False
                ).values_list('id', flat=True)
                
                # Özü + team
                allowed_employee_ids = list(team_employees) + [emp.id]
                
                requests_qs = requests_qs.filter(employee_id__in=allowed_employee_ids)
                schedules_qs = schedules_qs.filter(employee_id__in=allowed_employee_ids)
                
            except Employee.DoesNotExist:
                # Employee yoxdursa, yalnız özünü
                requests_qs = requests_qs.filter(employee__user=request.user)
                schedules_qs = schedules_qs.filter(employee__user=request.user)
        else:
            # Başqa istifadəçilər - yalnız özü (bu hal baş verməməlidir, amma safety)
            requests_qs = requests_qs.filter(employee__user=request.user)
            schedules_qs = schedules_qs.filter(employee__user=request.user)
        
        # Apply other filters (existing code)
        if status_filter:  # ✅ Changed from 'status'
            requests_qs = requests_qs.filter(status=status_filter)
            schedules_qs = schedules_qs.filter(status=status_filter)
        
        if vacation_type_id:
            requests_qs = requests_qs.filter(vacation_type_id=vacation_type_id)
            schedules_qs = schedules_qs.filter(vacation_type_id=vacation_type_id)
        
        if department_id:
            requests_qs = requests_qs.filter(employee__department_id=department_id)
            schedules_qs = schedules_qs.filter(employee__department_id=department_id)
        
        if start_date:
            requests_qs = requests_qs.filter(start_date__gte=start_date)
            schedules_qs = schedules_qs.filter(start_date__gte=start_date)
        
        if end_date:
            requests_qs = requests_qs.filter(end_date__lte=end_date)
            schedules_qs = schedules_qs.filter(end_date__lte=end_date)
        if business_function_id:
            requests_qs = requests_qs.filter(employee__business_function_id=business_function_id)
            schedules_qs = schedules_qs.filter(employee__business_function_id=business_function_id)
        if employee_name:
            requests_qs = requests_qs.filter(employee__full_name__icontains=employee_name)
            schedules_qs = schedules_qs.filter(employee__full_name__icontains=employee_name)
        
        if year:
            requests_qs = requests_qs.filter(start_date__year=year)
            schedules_qs = schedules_qs.filter(start_date__year=year)
        
        # Get data
        requests = requests_qs.order_by('-created_at')
        schedules = schedules_qs.order_by('-created_at')
        
      
        
        wb = Workbook()
        
        if export_format == 'separated':
            # Ayrı sheet-lər
            
            # Requests sheet
            ws_req = wb.active
            ws_req.title = "Vacation Requests"
            
            req_headers = [
                'Request ID', 'Employee Name', 'Employee ID', 'Department', 'Business Function',
                'Vacation Type', 'Start Date', 'End Date', 'Return Date', 'Working Days',
                'Status', 'Comment', 'Request Type',
                'Line Manager', 'LM Comment', 'LM Approved At', 'LM Approved By',
                'HR Representative', 'HR Comment', 'HR Approved At', 'HR Approved By',
                'Rejected By', 'Rejection Reason', 'Rejected At',
                'Created At', 'Updated At'
            ]
            ws_req.append(req_headers)
            
            for req in requests:
                ws_req.append([
                    req.request_id,
                    req.employee.full_name,
                    getattr(req.employee, 'employee_id', ''),
                    req.employee.department.name if req.employee.department else '',
                    req.employee.business_function.name if req.employee.business_function else '',
                    req.vacation_type.name,
                    req.start_date.strftime('%Y-%m-%d'),
                    req.end_date.strftime('%Y-%m-%d'),
                    req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                    float(req.number_of_days),
                    req.get_status_display(),
                    req.comment,
                    req.get_request_type_display(),
                    req.line_manager.full_name if req.line_manager else '',
                    req.line_manager_comment,
                    req.line_manager_approved_at.strftime('%Y-%m-%d %H:%M') if req.line_manager_approved_at else '',
                    req.line_manager_approved_by.get_full_name() if req.line_manager_approved_by else '',
                    req.hr_representative.full_name if req.hr_representative else '',
                    req.hr_comment,
                    req.hr_approved_at.strftime('%Y-%m-%d %H:%M') if req.hr_approved_at else '',
                    req.hr_approved_by.get_full_name() if req.hr_approved_by else '',
                    req.rejected_by.get_full_name() if req.rejected_by else '',
                    req.rejection_reason,
                    req.rejected_at.strftime('%Y-%m-%d %H:%M') if req.rejected_at else '',
                    req.created_at.strftime('%Y-%m-%d %H:%M') if req.created_at else '',
                    req.updated_at.strftime('%Y-%m-%d %H:%M') if req.updated_at else ''
                ])
            
            # Schedules sheet
            ws_sch = wb.create_sheet("Vacation Schedules")
            
            sch_headers = [
                'Schedule ID', 'Employee Name', 'Employee ID', 'Department', 'Business Function',
                'Vacation Type', 'Start Date', 'End Date', 'Return Date', 'Working Days',
                'Status', 'Comment', 
                'Edit Count', 'Can Edit', 'Last Edited By', 'Last Edited At',
                'Created By', 'Created At', 'Updated At'
            ]
            ws_sch.append(sch_headers)
            
            for sch in schedules:
                ws_sch.append([
                    f'SCH{sch.id}',
                    sch.employee.full_name,
                    getattr(sch.employee, 'employee_id', ''),
                    sch.employee.department.name if sch.employee.department else '',
                    sch.employee.business_function.name if sch.employee.business_function else '',
                    sch.vacation_type.name,
                    sch.start_date.strftime('%Y-%m-%d'),
                    sch.end_date.strftime('%Y-%m-%d'),
                    sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                    float(sch.number_of_days),
                    sch.get_status_display(),
                    sch.comment,
                    sch.edit_count,
                    'Yes' if sch.can_edit() else 'No',
                    sch.last_edited_by.get_full_name() if sch.last_edited_by else '',
                    sch.last_edited_at.strftime('%Y-%m-%d %H:%M') if sch.last_edited_at else '',
                    sch.created_by.get_full_name() if sch.created_by else '',
                    sch.created_at.strftime('%Y-%m-%d %H:%M') if sch.created_at else '',
                    sch.updated_at.strftime('%Y-%m-%d %H:%M') if sch.updated_at else ''
                ])
            
        else:
            # Combined sheet (default)
            ws = wb.active
            ws.title = "All Vacation Records"
            
            headers = [
                'Type', 'ID', 'Employee Name', 'Employee ID', 'Department', 'Business Function',
                'Vacation Type', 'Start Date', 'End Date', 'Return Date', 'Working Days',
                'Status', 'Comment', 
                'Line Manager/Created By', 'HR Representative', 'Approval Status',
                'Edit Count', 'Created At', 'Updated At'
            ]
            ws.append(headers)
            
            # Combine and sort all records
            all_records = []
            
            # Add requests
            for req in requests:
                approval_status = []
                if req.line_manager_approved_at:
                    approval_status.append('LM ✓')
                elif req.status == 'PENDING_LINE_MANAGER':
                    approval_status.append('LM ⏳')
                elif req.status == 'REJECTED_LINE_MANAGER':
                    approval_status.append('LM ❌')
                
                if req.hr_approved_at:
                    approval_status.append('HR ✓')
                elif req.status == 'PENDING_HR':
                    approval_status.append('HR ⏳')
                elif req.status == 'REJECTED_HR':
                    approval_status.append('HR ❌')
                
                all_records.append({
                    'type': 'Request',
                    'id': req.request_id,
                    'employee_name': req.employee.full_name,
                    'employee_id': getattr(req.employee, 'employee_id', ''),
                    'department': req.employee.department.name if req.employee.department else '',
                    'business_function': req.employee.business_function.name if req.employee.business_function else '',
                    'vacation_type': req.vacation_type.name,
                    'start_date': req.start_date.strftime('%Y-%m-%d'),
                    'end_date': req.end_date.strftime('%Y-%m-%d'),
                    'return_date': req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                    'working_days': float(req.number_of_days),
                    'status': req.get_status_display(),
                    'comment': req.comment,
                    'manager_created_by': req.line_manager.full_name if req.line_manager else '',
                    'hr_representative': req.hr_representative.full_name if req.hr_representative else '',
                    'approval_status': ' | '.join(approval_status),
                    'edit_count': '',
                    'created_at': req.created_at,
                    'updated_at': req.updated_at
                })
            
            # Add schedules
            for sch in schedules:
                all_records.append({
                    'type': 'Schedule',
                    'id': f'SCH{sch.id}',
                    'employee_name': sch.employee.full_name,
                    'employee_id': getattr(sch.employee, 'employee_id', ''),
                    'department': sch.employee.department.name if sch.employee.department else '',
                    'business_function': sch.employee.business_function.name if sch.employee.business_function else '',
                    'vacation_type': sch.vacation_type.name,
                    'start_date': sch.start_date.strftime('%Y-%m-%d'),
                    'end_date': sch.end_date.strftime('%Y-%m-%d'),
                    'return_date': sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                    'working_days': float(sch.number_of_days),
                    'status': sch.get_status_display(),
                    'comment': sch.comment,
                    'manager_created_by': sch.created_by.get_full_name() if sch.created_by else '',
                    'hr_representative': '',
                    'approval_status': 'No Approval Needed',
                    'edit_count': sch.edit_count,
                    'created_at': sch.created_at,
                    'updated_at': sch.updated_at
                })
            
            # Sort by created_at desc
            all_records.sort(key=lambda x: x['created_at'] if x['created_at'] else datetime.min, reverse=True)
            
            # Add data to sheet
            for record in all_records:
                ws.append([
                    record['type'],
                    record['id'],
                    record['employee_name'],
                    record['employee_id'],
                    record['department'],
                    record['business_function'],
                    record['vacation_type'],
                    record['start_date'],
                    record['end_date'],
                    record['return_date'],
                    record['working_days'],
                    record['status'],
                    record['comment'],
                    record['manager_created_by'],
                    record['hr_representative'],
                    record['approval_status'],
                    record['edit_count'],
                    record['created_at'].strftime('%Y-%m-%d %H:%M') if record['created_at'] else '',
                    record['updated_at'].strftime('%Y-%m-%d %H:%M') if record['updated_at'] else ''
                ])
        
        # Auto-adjust column widths for all sheets
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        # Generate filename
        filter_parts = []
        if status_filter:
            filter_parts.append(f"status_{status_filter}")
        if department_id:
            filter_parts.append(f"dept_{department_id}")
        if year:
            filter_parts.append(f"year_{year}")
        

   
        
        permission_indicator = 'admin' if is_admin else ('hr' if has_hr_permission else 'team')
        filename = f'vacation_records_{permission_indicator}_{export_format}_{date.today().strftime("%Y%m%d")}.xlsx'
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        
        return response
        
    except Exception as e:
        return Response({'error': f'Export error: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

# ==================== EXPORT REQUESTS & SCHEDULES ====================

@swagger_auto_schema(
    method='get',
    operation_description="İstifadəçinin bütün vacation məlumatlarını Excel formatında export et",
    operation_summary="Export My Vacations",
    tags=['Vacation'],
    responses={
        200: openapi.Response(
            description='Excel file',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    }
)
@api_view(['GET'])
@has_vacation_permission('vacation.request.export_own')
@permission_classes([IsAuthenticated])
def export_my_vacations(request):
    """İstifadəçinin vacation məlumatlarını enhanced formatda export et"""
    try:
        emp = Employee.objects.get(user=request.user)
        
        # All requests and schedules
        requests = VacationRequest.objects.filter(employee=emp, is_deleted=False).order_by('-created_at')
        schedules = VacationSchedule.objects.filter(employee=emp, is_deleted=False).order_by('-start_date')
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Personal Summary"
        ws_summary.sheet_view.showGridLines = False
        
        # Styles
        title_font = Font(size=16, bold=True, color="2B4C7E")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        stat_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        
        # Personal info header
        ws_summary['A1'] = f'VACATION SUMMARY - {emp.full_name}'
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:E1')
        
        ws_summary['A2'] = f'Employee ID: {getattr(emp, "employee_id", "N/A")}'
        ws_summary['A3'] = f'Department: {emp.department.name if emp.department else "N/A"}'
        ws_summary['A4'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        
        # ✅ DÜZƏLTMƏ: Yalnız data olan illəri tap
        years_with_data = set()
        
        # Requests-dən illəri tap
        for req in requests:
            years_with_data.add(req.start_date.year)
        
        # Schedules-dən illəri tap
        for sch in schedules:
            years_with_data.add(sch.start_date.year)
        
        # Balansı olan illəri tap
        balances_with_data = EmployeeVacationBalance.objects.filter(
            employee=emp,
            is_deleted=False
        ).exclude(
            yearly_balance=0,
            used_days=0,
            scheduled_days=0
        ).values_list('year', flat=True)
        
        years_with_data.update(balances_with_data)
        
        # ✅ Əgər heç bir data yoxdursa, cari ili əlavə et
        if not years_with_data:
            years_with_data.add(date.today().year)
        
        # Sort years
        sorted_years = sorted(years_with_data, reverse=True)
        
        ws_summary['A6'] = 'YEARLY STATISTICS'
        ws_summary['A6'].font = Font(size=14, bold=True, color="2B4C7E")
        
        year_headers = ['Year', 'Total Balance', 'Used Days', 'Scheduled Days', 'Remaining', 'Requests', 'Schedules']
        for col, header in enumerate(year_headers, 1):
            cell = ws_summary.cell(row=8, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for i, year in enumerate(sorted_years, 9):
            balance = EmployeeVacationBalance.objects.filter(employee=emp, year=year).first()
            year_requests = requests.filter(start_date__year=year).count()
            year_schedules = schedules.filter(start_date__year=year).count()
            
            if balance:
                data = [
                    year,
                    float(balance.total_balance),
                    float(balance.used_days),
                    float(balance.scheduled_days),
                    float(balance.remaining_balance),
                    year_requests,
                    year_schedules
                ]
            else:
                data = [year, 0, 0, 0, 0, year_requests, year_schedules]
            
            for col, value in enumerate(data, 1):
                cell = ws_summary.cell(row=i, column=col, value=value)
                if col > 1:
                    cell.fill = stat_fill
                cell.alignment = Alignment(horizontal='center' if col == 1 else 'right')
        
        # Requests sheet
        ws_requests = wb.create_sheet("Vacation Requests")
        
        req_headers = [
            'Request ID', 'Type', 'Vacation Type', 'Start Date', 'End Date', 'Return Date',
            'Days', 'Status', 'Comment', 'Line Manager', 'LM Comment', 'HR Comment', 
            'Created At', 'Timeline'
        ]
        
        for col, header in enumerate(req_headers, 1):
            cell = ws_requests.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Status colors
        status_colors = {
            'APPROVED': good_fill,
            'PENDING_LINE_MANAGER': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            'PENDING_HR': PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),
            'REJECTED_LINE_MANAGER': warning_fill,
            'REJECTED_HR': warning_fill,
        }
        
        for row, req in enumerate(requests, 2):
            timeline = f"Created: {req.created_at.strftime('%Y-%m-%d')}"
            if req.line_manager_approved_at:
                timeline += f" → LM: {req.line_manager_approved_at.strftime('%Y-%m-%d')}"
            if req.hr_approved_at:
                timeline += f" → HR: {req.hr_approved_at.strftime('%Y-%m-%d')}"
            if req.rejected_at:
                timeline += f" → Rejected: {req.rejected_at.strftime('%Y-%m-%d')}"
            
            data = [
                req.request_id,
                req.get_request_type_display(),
                req.vacation_type.name,
                req.start_date.strftime('%Y-%m-%d'),
                req.end_date.strftime('%Y-%m-%d'),
                req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                float(req.number_of_days),
                req.get_status_display(),
                req.comment,
                req.line_manager.full_name if req.line_manager else '',
                req.line_manager_comment,
                req.hr_comment,
                req.created_at.strftime('%Y-%m-%d %H:%M'),
                timeline
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_requests.cell(row=row, column=col, value=value)
                if col == 8:  # Status column
                    cell.fill = status_colors.get(req.status, PatternFill())
        
        # Schedules sheet
        ws_schedules = wb.create_sheet("Vacation Schedules")
        
        sch_headers = [
            'Schedule ID', 'Vacation Type', 'Start Date', 'End Date', 'Return Date',
            'Days', 'Status', 'Comment', 'Edit Count', 'Can Edit', 'Created At',
            'Last Edited', 'Timeline'
        ]
        
        for col, header in enumerate(sch_headers, 1):
            cell = ws_schedules.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, sch in enumerate(schedules, 2):
            timeline = f"Created: {sch.created_at.strftime('%Y-%m-%d')}"
            if sch.last_edited_at:
                timeline += f" → Last Edit: {sch.last_edited_at.strftime('%Y-%m-%d')}"
            
            data = [
                f'SCH{sch.id}',
                sch.vacation_type.name,
                sch.start_date.strftime('%Y-%m-%d'),
                sch.end_date.strftime('%Y-%m-%d'),
                sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                float(sch.number_of_days),
                sch.get_status_display(),
                sch.comment,
                sch.edit_count,
                'Yes' if sch.can_edit() else 'No',
                sch.created_at.strftime('%Y-%m-%d %H:%M'),
                sch.last_edited_at.strftime('%Y-%m-%d %H:%M') if sch.last_edited_at else '',
                timeline
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_schedules.cell(row=row, column=col, value=value)
                if col == 7:  # Status column
                    if sch.status == 'REGISTERED':
                        cell.fill = good_fill
                    else:
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Auto-adjust column widths
        for ws_current in [ws_summary, ws_requests, ws_schedules]:
            for column in ws_current.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws_current.column_dimensions[column_letter].width = adjusted_width
        
        filename = f'my_vacations_{emp.full_name.replace(" ", "_")}_{date.today().strftime("%Y%m%d")}.xlsx'
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        
        return response
        
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profili tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== SETTINGS VIEWSETS ====================

class VacationSettingViewSet(viewsets.ModelViewSet):
    """Vacation Settings CRUD"""
    queryset = VacationSetting.objects.filter(is_deleted=False)
    serializer_class = VacationSettingSerializer
    permission_classes = [IsAuthenticated]
    
    def get_swagger_tags(self):
        return ['Vacation - Settings']
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
class VacationTypeViewSet(viewsets.ModelViewSet):
    queryset = VacationType.objects.filter(is_deleted=False)
    serializer_class = VacationTypeSerializer
    permission_classes = [IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        has_perm, _ = check_vacation_permission(request.user, 'vacation.type.view')
        if not has_perm:
            return Response({
                'error': 'İcazə yoxdur',
                'required_permission': 'vacation.type.view'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().list(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        has_perm, _ = check_vacation_permission(request.user, 'vacation.type.create')
        if not has_perm:
            return Response({
                'error': 'İcazə yoxdur',
                'required_permission': 'vacation.type.create'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        has_perm, _ = check_vacation_permission(request.user, 'vacation.type.update')
        if not has_perm:
            return Response({
                'error': 'İcazə yoxdur',
                'required_permission': 'vacation.type.update'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        has_perm, _ = check_vacation_permission(request.user, 'vacation.type.update')
        if not has_perm:
            return Response({
                'error': 'İcazə yoxdur',
                'required_permission': 'vacation.type.update'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        has_perm, _ = check_vacation_permission(request.user, 'vacation.type.delete')
        if not has_perm:
            return Response({
                'error': 'İcazə yoxdur',
                'required_permission': 'vacation.type.delete'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

# ==================== FILE UPLOAD ENDPOINTS ====================

@swagger_auto_schema(
    method='get',
    operation_description="Get all attachments for a vacation request",
    operation_summary="List Vacation Request Attachments",
    tags=['Vacation - Files'],
    responses={
        200: openapi.Response(
            description='List of attachments',
            schema=VacationAttachmentSerializer(many=True)
        )
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_vacation_request_attachments(request, request_id):
    """Get all attachments for a vacation request"""
    try:
        vacation_request = get_object_or_404(
            VacationRequest, 
            request_id=request_id, 
            is_deleted=False
        )
        
        # Check access permission
        from .vacation_permissions import is_admin_user
        emp = None
        try:
            emp = Employee.objects.get(user=request.user, is_deleted=False)
        except Employee.DoesNotExist:
            pass
        
        can_view = False
        if is_admin_user(request.user):
            can_view = True
        elif emp and vacation_request.employee == emp:
            can_view = True
        elif vacation_request.requester == request.user:
            can_view = True
        elif emp and (vacation_request.line_manager == emp or vacation_request.hr_representative == emp):
            can_view = True
        
        if not can_view:
            return Response({
                'error': 'Permission denied'
            }, status=status.HTTP_403_FORBIDDEN)
        
        attachments = vacation_request.attachments.filter(is_deleted=False).order_by('-uploaded_at')
        
        return Response({
            'request_id': request_id,
            'count': attachments.count(),
            'attachments': VacationAttachmentSerializer(
                attachments, 
                many=True, 
                context={'request': request}
            ).data
        })
        
    except VacationRequest.DoesNotExist:
        return Response({
            'error': 'Vacation request not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)



@swagger_auto_schema(
    method='delete',
    operation_description="Delete a file attachment (Only uploader can delete)",
    operation_summary="Delete Vacation Attachment",
    tags=['Vacation - Files'],
    responses={
        200: 'File deleted successfully',
        403: 'Permission denied',
        404: 'Attachment not found'
    }
)
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_vacation_attachment(request, attachment_id):
    """Delete a file attachment"""
    try:
        from .vacation_permissions import is_admin_user
        from .vacation_models import VacationAttachment
        
        attachment = get_object_or_404(
            VacationAttachment, 
            id=attachment_id, 
            is_deleted=False
        )
        
        # Check permission - only uploader or admin can delete
        if attachment.uploaded_by != request.user and not is_admin_user(request.user):
            return Response({
                'error': 'You can only delete files you uploaded'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Soft delete
        attachment.is_deleted = True
        attachment.save()
        
        logger.info(f"🗑️ File deleted: {attachment.original_filename} by {request.user.username}")
        
        return Response({
            'message': 'File deleted successfully',
            'filename': attachment.original_filename
        })
        
    except Exception as e:
        logger.error(f"Error deleting file: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='post',
    operation_description="Upload multiple files at once to vacation request",
    operation_summary="Bulk Upload Vacation Attachments",
    tags=['Vacation - Files'],
    manual_parameters=[
        openapi.Parameter(
            'files',
            openapi.IN_FORM,
            type=openapi.TYPE_ARRAY,
            items=openapi.Items(type=openapi.TYPE_FILE),
            required=True,
            description='Multiple files to upload'
        )
    ],
    responses={
        201: 'Files uploaded successfully',
        400: 'Bad request'
    }
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def bulk_upload_vacation_attachments(request, request_id):
    """Upload multiple files at once"""
    try:
        vacation_request = get_object_or_404(
            VacationRequest, 
            request_id=request_id, 
            is_deleted=False
        )
        
        # Check permission
        emp = None
        try:
            emp = Employee.objects.get(user=request.user, is_deleted=False)
        except Employee.DoesNotExist:
            pass
        
        can_upload = False
        if emp and vacation_request.employee == emp:
            can_upload = True
        elif vacation_request.requester == request.user:
            can_upload = True
        
        if not can_upload:
            return Response({
                'error': 'Permission denied'
            }, status=status.HTTP_403_FORBIDDEN)
        
        files = request.FILES.getlist('files')
        if not files:
            return Response({
                'error': 'No files provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        uploaded_attachments = []
        errors = []
        
        for file in files:
            try:
                # Validate each file
                from .business_trip_serializers import TripAttachmentUploadSerializer
                upload_serializer = TripAttachmentUploadSerializer(data={'file': file})
                if not upload_serializer.is_valid():
                    errors.append({
                        'filename': file.name,
                        'errors': upload_serializer.errors
                    })
                    continue
                
                # Create attachment
                from .vacation_models import VacationAttachment
                attachment = VacationAttachment.objects.create(
                    vacation_request=vacation_request,
                    file=file,
                    original_filename=file.name,
                    file_size=file.size,
                    file_type=file.content_type,
                    uploaded_by=request.user
                )
                uploaded_attachments.append(attachment)
                
            except Exception as e:
                errors.append({
                    'filename': file.name,
                    'error': str(e)
                })
        
        logger.info(f"✅ Bulk upload: {len(uploaded_attachments)} files to {request_id}")
        
        return Response({
            'message': f'{len(uploaded_attachments)} files uploaded successfully',
            'uploaded': VacationAttachmentSerializer(
                uploaded_attachments, 
                many=True, 
                context={'request': request}
            ).data,
            'errors': errors,
            'success_count': len(uploaded_attachments),
            'error_count': len(errors)
        }, status=status.HTTP_201_CREATED)
        
    except VacationRequest.DoesNotExist:
        return Response({
            'error': 'Vacation request not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error in bulk upload: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_description="Get attachment details",
    operation_summary="Get Attachment Details",
    tags=['Vacation - Files'],
    responses={
        200: openapi.Response(
            description='Attachment details',
            schema=VacationAttachmentSerializer
        )
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_vacation_attachment_details(request, attachment_id):
    """Get details of a specific attachment"""
    try:
        from .vacation_models import VacationAttachment
        
        attachment = get_object_or_404(
            VacationAttachment, 
            id=attachment_id, 
            is_deleted=False
        )
        
        return Response({
            'attachment': VacationAttachmentSerializer(
                attachment, 
                context={'request': request}
            ).data
        })
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


# ==================== SCHEDULE DETAIL ====================

@swagger_auto_schema(
    method='get',
    operation_description="Get detailed information of a vacation schedule",
    operation_summary="Get Vacation Schedule Detail",
    tags=['Vacation'],
    responses={
        200: openapi.Response(
            description='Vacation schedule details',
            schema=VacationScheduleSerializer
        ),
        403: 'Permission denied',
        404: 'Schedule not found'
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_vacation_schedule_detail(request, pk):
    """
    Get detailed information of a vacation schedule
    
    Shows:
    - Basic schedule information
    - Employee details
    - Vacation configuration (type)
    - Edit history
    - Current status
    """
    try:
        from .vacation_permissions import is_admin_user
        
        # Get the vacation schedule
        schedule = VacationSchedule.objects.select_related(
            'employee', 
            'employee__department',
            'employee__business_function',
            'employee__unit',
            'employee__job_function',
            'vacation_type',
            'created_by',
            'last_edited_by'
        ).get(pk=pk, is_deleted=False)
        
        # Check access permission
        emp = None
        try:
            emp = Employee.objects.get(user=request.user, is_deleted=False)
        except Employee.DoesNotExist:
            pass
        
        # Determine if user can view this schedule
        can_view = False
        
        # Admin can view all
        if is_admin_user(request.user):
            can_view = True
        
        # Employee can view their own schedules
        elif emp and schedule.employee == emp:
            can_view = True
        
        # Creator can view schedules they created
        elif schedule.created_by == request.user:
            can_view = True
        
        # Line manager can view team schedules
        elif emp and schedule.employee.line_manager == emp:
            can_view = True
        
        # Check if user has view_all permission
        elif check_vacation_permission(request.user, 'vacation.schedule.view_all')[0]:
            can_view = True
        
        if not can_view:
            return Response({
                'error': 'Permission denied',
                'detail': 'You do not have permission to view this vacation schedule'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Serialize the data
        serializer = VacationScheduleSerializer(
            schedule, 
            context={'request': request}
        )
        
        # Add extra context information
        response_data = serializer.data
        
        # Add employee information
        response_data['employee_info'] = {
            'id': schedule.employee.id,
            'name': schedule.employee.full_name,
            'employee_id': getattr(schedule.employee, 'employee_id', ''),
            'department': schedule.employee.department.name if schedule.employee.department else None,
            'business_function': schedule.employee.business_function.name if schedule.employee.business_function else None,
            'phone': schedule.employee.phone
        }
        
        # Add edit history
        response_data['edit_history'] = {
            'edit_count': schedule.edit_count,
            'can_edit': schedule.can_edit(),
            'last_edited_by': schedule.last_edited_by.get_full_name() if schedule.last_edited_by else None,
            'last_edited_at': schedule.last_edited_at,
            'max_edits_allowed': VacationSetting.get_active().max_schedule_edits if VacationSetting.get_active() else 3
        }
        
        # Add creator information
        response_data['creator_info'] = {
            'name': schedule.created_by.get_full_name() if schedule.created_by else None,
            'email': schedule.created_by.email if schedule.created_by else None
        }
        
        # Add permission flags for frontend
        response_data['permissions'] = {
            'can_edit': (
                schedule.status == 'SCHEDULED' and
                emp and schedule.employee == emp and
                schedule.can_edit()
            ),
            'can_delete': (
                schedule.status == 'SCHEDULED' and
                (emp and schedule.employee == emp or 
                 (emp and schedule.employee.line_manager == emp))
            ),
            'can_register': (
                schedule.status == 'SCHEDULED' and
                check_vacation_permission(request.user, 'vacation.schedule.register')[0]
            ),
            'is_admin': is_admin_user(request.user)
        }
        
        return Response(response_data)
        
    except VacationSchedule.DoesNotExist:
        return Response({
            'error': 'Vacation schedule not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error fetching vacation schedule detail: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)
# ==================== UTILITIES ====================
@swagger_auto_schema(
    method='post',
    operation_description="İki tarix arasında iş günlərinin sayını hesabla",
    operation_summary="İş Günlərini Hesabla",
    tags=['Vacation'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['start_date', 'end_date'],
        properties={
            'start_date': openapi.Schema(
                type=openapi.TYPE_STRING,
                format='date',
                description='Başlama tarixi',
                example='2025-10-15'
            ),
            'end_date': openapi.Schema(
                type=openapi.TYPE_STRING,
                format='date',
                description='Bitmə tarixi',
                example='2025-10-20'
            ),
        }
    ),
    responses={
        200: openapi.Response(
            description='Hesablama nəticəsi',
            examples={
                'application/json': {
                    'working_days': 4,
                    'return_date': '2025-10-21'
                }
            }
        )
    }
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def calculate_working_days(request):
    """İş günlərini hesabla"""
    start = request.data.get('start_date')
    end = request.data.get('end_date')
    
    if not start or not end:
        return Response({'error': 'start_date və end_date mütləqdir'}, status=status.HTTP_400_BAD_REQUEST)
    
    settings = VacationSetting.get_active()
    if settings:
        try:
            start_dt = datetime.strptime(start, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end, '%Y-%m-%d').date()
            
            if start_dt > end_dt:
                return Response({'error': 'start_date end_date-dən kiçik olmalıdır'}, status=status.HTTP_400_BAD_REQUEST)
            
            days = settings.calculate_working_days(start_dt, end_dt)
            return_date = settings.calculate_return_date(end_dt)
            
            return Response({
                'working_days': days,
                'return_date': return_date.strftime('%Y-%m-%d'),
                'total_calendar_days': (end_dt - start_dt).days + 1
            })
            
        except ValueError:
            return Response({'error': 'Tarix formatı səhvdir. YYYY-MM-DD istifadə edin'}, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({'error': 'Settings tapılmadı'}, status=status.HTTP_404_NOT_FOUND)