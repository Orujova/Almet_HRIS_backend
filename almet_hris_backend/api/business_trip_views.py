# api/business_trip_views.py - Updated with Role-Based Permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status, viewsets
from django.db import transaction
from django.db.models import Q, Count, Sum
from datetime import date, datetime, timedelta
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .business_trip_models import *
from .business_trip_serializers import *
from .models import Employee
from .business_trip_permissions import (
    has_business_trip_permission,
    has_any_business_trip_permission,
    check_business_trip_permission,
    get_user_business_trip_permissions,
    is_admin_user
)
import logging


logger = logging.getLogger(__name__)
from .business_trip_notifications import notification_manager

from .business_trip_notifications import notification_manager
from .models import UserGraphToken  # ⭐ ADD THIS

# Helper function to get Microsoft Graph token
def get_graph_access_token(user):
    """
    Get Microsoft Graph access token for the authenticated user
    
    Args:
        user: Django User object
    
    Returns:
        str: Graph access token or None
    """
    token = UserGraphToken.get_valid_token(user)
    if not token:
        logger.warning(f"⚠️ No valid Graph token found for user {user.username}")
        logger.warning("Email notifications will be skipped")
    return token
# ==================== MY PERMISSIONS ====================
@swagger_auto_schema(
    method='get',
    operation_description="İstifadəçinin bütün Business Trip icazələrini əldə et",
    operation_summary="Get My Permissions",
    tags=['Business Trip'],
    responses={
        200: openapi.Response(
            description='User permissions',
            examples={
                'application/json': {
                    'is_admin': False,
                    'permissions': [
                        'business_trips.request.create',
                        'business_trips.request.view'
                    ],
                    'roles': ['Employee - Business Trips']
                }
            }
        )
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_business_trip_permissions(request):
    """İstifadəçinin Business Trip permissions-larını göstər"""
    from .role_models import EmployeeRole
    
    is_admin = is_admin_user(request.user)
    permissions = get_user_business_trip_permissions(request.user)
    
    # User roles
    try:
        emp = Employee.objects.get(user=request.user, is_deleted=False)
        roles = list(EmployeeRole.objects.filter(
            employee=emp,
            is_active=True
        ).values_list('role__name', flat=True))
    except Employee.DoesNotExist:
        roles = []
    
    return Response({
        'is_admin': is_admin,
        'permissions_count': len(permissions),
        'permissions': permissions,
        'roles_count': len(roles),
        'roles': roles
    })

# ==================== DASHBOARD ====================
@swagger_auto_schema(
    method='get',
    operation_description="Dashboard statistics",
    operation_summary="Dashboard",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='Dashboard data')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def trip_dashboard(request):
    """Dashboard - statistics (permissions yoxlanmır, hər kəs özünkünü görür)"""
    try:
        emp = Employee.objects.get(user=request.user)
        year = date.today().year
        
        # My trips
        my_trips = BusinessTripRequest.objects.filter(employee=emp)
        
        stats = {
            'pending_requests': my_trips.filter(
                status__in=['SUBMITTED', 'PENDING_LINE_MANAGER', 'PENDING_FINANCE', 'PENDING_HR']
            ).count(),
            'approved_trips': my_trips.filter(status='APPROVED').count(),
            'total_days_this_year': float(my_trips.filter(
                start_date__year=year,
                status='APPROVED'
            ).aggregate(total=Sum('number_of_days'))['total'] or 0),
            'upcoming_trips': my_trips.filter(
                status='APPROVED',
                start_date__gte=date.today()
            ).count()
        }
        
        return Response(stats)
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== SETTINGS ====================

@swagger_auto_schema(
    method='put',
    operation_description="Update default HR representative",
    operation_summary="Update HR Representative",
    tags=['Business Trip'],
    request_body=HRRepresentativeSerializer,
    responses={200: openapi.Response(description='HR updated')}
)
@api_view(['PUT'])
@has_business_trip_permission('business_trips.settings.update')
@permission_classes([IsAuthenticated])
def update_hr_representative(request):
    """Update default HR representative"""
    try:
        serializer = HRRepresentativeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        hr_id = serializer.validated_data['default_hr_representative_id']
        hr_employee = Employee.objects.get(id=hr_id, is_deleted=False)
        
        settings = TripSettings.get_active()
        if not settings:
            settings = TripSettings.objects.create(is_active=True, created_by=request.user)
        
        previous_hr = settings.default_hr_representative
        settings.default_hr_representative = hr_employee
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Default HR representative updated',
            'previous_hr': {
                'id': previous_hr.id,
                'name': previous_hr.full_name
            } if previous_hr else None,
            'current_hr': {
                'id': hr_employee.id,
                'name': hr_employee.full_name,
                'department': hr_employee.department.name if hr_employee.department else ''
            }
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='get',
    operation_description="Get HR representatives",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='HR list')}
)
@api_view(['GET'])
@has_business_trip_permission('business_trips.settings.view')
@permission_classes([IsAuthenticated])
def get_hr_representatives(request):
    """Get HR representatives list"""
    try:
        settings = TripSettings.get_active()
        current_default = settings.default_hr_representative if settings else None
        
        hr_employees = Employee.objects.filter(
            department__name__icontains='HR',
            is_deleted=False
        )
        
        hr_list = [{
            'id': emp.id,
            'name': emp.full_name,
            'email': emp.user.email if emp.user else '',
            'phone': emp.phone,
            'department': emp.department.name if emp.department else '',
            'is_default': current_default and current_default.id == emp.id
        } for emp in hr_employees]
        
        return Response({
            'current_default': {
                'id': current_default.id,
                'name': current_default.full_name
            } if current_default else None,
            'hr_representatives': hr_list
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='put',
    operation_description="Update default finance approver",
    operation_summary="Update Finance Approver",
    tags=['Business Trip'],
    request_body=FinanceApproverSerializer,
    responses={200: openapi.Response(description='Finance updated')}
)
@api_view(['PUT'])
@has_business_trip_permission('business_trips.settings.update')
@permission_classes([IsAuthenticated])
def update_finance_approver(request):
    """Update default finance approver"""
    try:
        serializer = FinanceApproverSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        finance_id = serializer.validated_data['default_finance_approver_id']
        finance_employee = Employee.objects.get(id=finance_id, is_deleted=False)
        
        settings = TripSettings.get_active()
        if not settings:
            settings = TripSettings.objects.create(is_active=True, created_by=request.user)
        
        previous_finance = settings.default_finance_approver
        settings.default_finance_approver = finance_employee
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Default Finance approver updated',
            'previous_finance': {
                'id': previous_finance.id,
                'name': previous_finance.full_name
            } if previous_finance else None,
            'current_finance': {
                'id': finance_employee.id,
                'name': finance_employee.full_name,
                'department': finance_employee.department.name if finance_employee.department else ''
            }
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='get',
    operation_description="Get finance approvers",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='Finance list')}
)
@api_view(['GET'])
@has_business_trip_permission('business_trips.settings.view')
@permission_classes([IsAuthenticated])
def get_finance_approvers(request):
    """Get finance approvers list"""
    try:
        settings = TripSettings.get_active()
        current_default = settings.default_finance_approver if settings else None
        
        finance_employees = Employee.objects.filter(
            Q(department__name__icontains='Finance') | Q(department__name__icontains='Payroll'),
            is_deleted=False
        )
        
        finance_list = [{
            'id': emp.id,
            'name': emp.full_name,
            'email': emp.user.email if emp.user else '',
            'phone': emp.phone,
            'department': emp.department.name if emp.department else '',
            'is_default': current_default and current_default.id == emp.id
        } for emp in finance_employees]
        
        return Response({
            'current_default': {
                'id': current_default.id,
                'name': current_default.full_name
            } if current_default else None,
            'finance_approvers': finance_list
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='put',
    operation_description="Update general trip settings (notification days)",
    tags=['Business Trip'],
    request_body=GeneralTripSettingsSerializer,
    responses={200: openapi.Response(description='Settings updated')}
)
@api_view(['PUT'])
@has_business_trip_permission('business_trips.settings.update')
@permission_classes([IsAuthenticated])
def update_general_settings(request):
    """Update general trip settings"""
    try:
        serializer = GeneralTripSettingsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        settings = TripSettings.get_active()
        if not settings:
            settings = TripSettings.objects.create(is_active=True, created_by=request.user)
        
        if 'notification_days_before' in data:
            settings.notification_days_before = data['notification_days_before']
        
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Trip settings updated',
            'settings': {
                'notification_days_before': settings.notification_days_before
            }
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='get',
    operation_description="Get general trip settings",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='Settings')}
)
@api_view(['GET'])
@has_business_trip_permission('business_trips.settings.view')
@permission_classes([IsAuthenticated])
def get_general_settings(request):
    """Get general trip settings"""
    try:
        settings = TripSettings.get_active()
        if not settings:
            return Response({
                'notification_days_before': 7
            })
        
        return Response({
            'notification_days_before': settings.notification_days_before
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== CREATE REQUEST ====================


@swagger_auto_schema(
    method='post',
    operation_description="Create business trip request",
    operation_summary="Create Trip Request",
    tags=['Business Trip'],
    request_body=BusinessTripRequestCreateSerializer,
    responses={201: openapi.Response(description='Request created')}
)
@api_view(['POST'])
@has_business_trip_permission('business_trips.request.create')
@permission_classes([IsAuthenticated])
def create_trip_request(request):
    """Create business trip request"""
    serializer = BusinessTripRequestCreateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    
    try:
        requester_emp = Employee.objects.get(user=request.user, is_deleted=False)
        
        # Employee selection
        if data['requester_type'] == 'for_me':
            employee = requester_emp
        else:
            if data.get('employee_id'):
                employee = Employee.objects.get(id=data['employee_id'])
                if employee.line_manager != requester_emp:
                    return Response({
                        'error': 'This employee is not in your team'
                    }, status=status.HTTP_403_FORBIDDEN)
            else:
                manual_data = data.get('employee_manual', {})
                if not manual_data.get('name'):
                    return Response({
                        'error': 'Employee name is required'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                employee = Employee.objects.create(
                    full_name=manual_data.get('name', ''),
                    phone=manual_data.get('phone', ''),
                    line_manager=requester_emp,
                    created_by=request.user
                )
        
        with transaction.atomic():
            trip_req = BusinessTripRequest.objects.create(
                employee=employee,
                requester=request.user,
                requester_type=data['requester_type'],
                travel_type_id=data['travel_type_id'],
                transport_type_id=data['transport_type_id'],
                purpose_id=data['purpose_id'],
                start_date=data['start_date'],
                end_date=data['end_date'],
                comment=data.get('comment', ''),
                finance_approver_id=data.get('finance_approver_id'),
                hr_representative_id=data.get('hr_representative_id')
            )
            
            # Create schedules
            for i, schedule_data in enumerate(data['schedules']):
                TripSchedule.objects.create(
                    trip_request=trip_req,
                    date=schedule_data['date'],
                    from_location=schedule_data['from_location'],
                    to_location=schedule_data['to_location'],
                    order=i + 1,
                    notes=schedule_data.get('notes', '')
                )
            
            # Create hotels
            for hotel_data in data.get('hotels', []):
                TripHotel.objects.create(
                    trip_request=trip_req,
                    hotel_name=hotel_data['hotel_name'],
                    check_in_date=hotel_data['check_in_date'],
                    check_out_date=hotel_data['check_out_date'],
                    location=hotel_data.get('location', ''),
                    notes=hotel_data.get('notes', '')
                )
            
            trip_req.submit_request(request.user)
            
            graph_token = get_graph_access_token(request.user)
            if graph_token:
                notification_manager.notify_request_created(trip_req, graph_token)
                logger.info("✅ Notification sent to Line Manager")
            else:
                logger.warning("⚠️ Graph token not available - notification skipped")
            
            return Response({
                'message': 'Trip request created and submitted.',
                'notification_sent': bool(graph_token),
                'request': BusinessTripRequestDetailSerializer(trip_req).data
            }, status=status.HTTP_201_CREATED)
            
            
    
    except Employee.DoesNotExist:
        return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
# ==================== MY REQUESTS ====================

@swagger_auto_schema(
    method='get',
    operation_description="Get all my trip requests",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='My requests')}
)
@api_view(['GET'])
@has_business_trip_permission('business_trips.request.view')
@permission_classes([IsAuthenticated])
def my_trip_requests(request):
    """Get all my trip requests"""
    try:
        emp = Employee.objects.get(user=request.user)
        requests = BusinessTripRequest.objects.filter(
            employee=emp, 
            is_deleted=False
        ).order_by('-created_at')
        
        return Response({
            'requests': BusinessTripRequestListSerializer(requests, many=True).data
        })
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== APPROVAL ====================

@swagger_auto_schema(
    method='get',
    operation_description="Get pending approvals",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='Pending approvals')}
)
@api_view(['GET'])
@has_business_trip_permission('business_trips.request.view_pending')
@permission_classes([IsAuthenticated])
def pending_approvals(request):
    """Get pending approvals"""
    try:
        emp = Employee.objects.get(user=request.user)
        
        # Admin görür bütün pending-ləri
        if is_admin_user(request.user):
            lm_requests = BusinessTripRequest.objects.filter(
                status='PENDING_LINE_MANAGER',
                is_deleted=False
            ).order_by('-created_at')
            
            finance_requests = BusinessTripRequest.objects.filter(
                status='PENDING_FINANCE',
                is_deleted=False
            ).order_by('-created_at')
            
            hr_requests = BusinessTripRequest.objects.filter(
                status='PENDING_HR',
                is_deleted=False
            ).order_by('-created_at')
        else:
            lm_requests = BusinessTripRequest.objects.filter(
                line_manager=emp,
                status='PENDING_LINE_MANAGER',
                is_deleted=False
            ).order_by('-created_at')
            
            finance_requests = BusinessTripRequest.objects.filter(
                finance_approver=emp,
                status='PENDING_FINANCE',
                is_deleted=False
            ).order_by('-created_at')
            
            hr_requests = BusinessTripRequest.objects.filter(
                hr_representative=emp,
                status='PENDING_HR',
                is_deleted=False
            ).order_by('-created_at')
        
        return Response({
            'line_manager_requests': BusinessTripRequestListSerializer(lm_requests, many=True).data,
            'finance_requests': BusinessTripRequestListSerializer(finance_requests, many=True).data,
            'hr_requests': BusinessTripRequestListSerializer(hr_requests, many=True).data,
            'total_pending': lm_requests.count() + finance_requests.count() + hr_requests.count(),
            'is_admin': is_admin_user(request.user)
        })
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='post',
    operation_description="Approve/Reject trip request",
    tags=['Business Trip'],
    request_body=TripApprovalSerializer,
    responses={200: openapi.Response(description='Action completed')}
)
@api_view(['POST'])
@has_business_trip_permission('business_trips.request.approve')
@permission_classes([IsAuthenticated])
def approve_reject_request(request, pk):
    """Approve/Reject trip request - admin və ya approver"""
    serializer = TripApprovalSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    graph_token = get_graph_access_token(request.user)

    try:
        trip_req = BusinessTripRequest.objects.get(pk=pk, is_deleted=False)
        
        if trip_req.status == 'PENDING_LINE_MANAGER':
            if data['action'] == 'approve':
                trip_req.approve_by_line_manager(request.user, data.get('comment', ''))
                msg = 'Approved by Line Manager'
                
                # ✅ SEND NOTIFICATION TO FINANCE
                if graph_token:
                    notification_manager.notify_line_manager_approved(trip_req, graph_token)
                    logger.info("✅ Notification sent to Finance")
            else:
                trip_req.reject_by_line_manager(request.user, data.get('reason', ''))
                msg = 'Rejected by Line Manager'
                
                # ✅ SEND REJECTION NOTIFICATION TO EMPLOYEE
                if graph_token:
                    notification_manager.notify_request_rejected(trip_req, graph_token)
                    logger.info("✅ Rejection notification sent to Employee")
        
        elif trip_req.status == 'PENDING_FINANCE':
            if data['action'] == 'approve':
                trip_req.approve_by_finance(request.user, data.get('amount'), data.get('comment', ''))
                msg = 'Approved by Finance'
                
                # ✅ SEND NOTIFICATION TO HR
                if graph_token:
                    notification_manager.notify_finance_approved(trip_req, graph_token)
                    logger.info("✅ Notification sent to HR")
            else:
                trip_req.reject_by_finance(request.user, data.get('reason', ''))
                msg = 'Rejected by Finance'
                
                # ✅ SEND REJECTION NOTIFICATION TO EMPLOYEE
                if graph_token:
                    notification_manager.notify_request_rejected(trip_req, graph_token)
                    logger.info("✅ Rejection notification sent to Employee")
        
        elif trip_req.status == 'PENDING_HR':
            if data['action'] == 'approve':
                trip_req.approve_by_hr(request.user, data.get('comment', ''))
                msg = 'Approved by HR - Request is now APPROVED'
                
                # ✅ SEND FINAL APPROVAL NOTIFICATION TO EMPLOYEE
                if graph_token:
                    notification_manager.notify_hr_approved(trip_req, graph_token)
                    logger.info("✅ Final approval notification sent to Employee")
            else:
                trip_req.reject_by_hr(request.user, data.get('reason', ''))
                msg = 'Rejected by HR'
                
                # ✅ SEND REJECTION NOTIFICATION TO EMPLOYEE
                if graph_token:
                    notification_manager.notify_request_rejected(trip_req, graph_token)
                    logger.info("✅ Rejection notification sent to Employee")
        else:
            return Response({'error': 'Request is not pending approval'}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'message': msg,
            'notification_sent': bool(graph_token),
            'request': BusinessTripRequestDetailSerializer(trip_req).data
        })
        
    
    
    except BusinessTripRequest.DoesNotExist:
        return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
# ==================== APPROVAL HISTORY ====================

@swagger_auto_schema(
    method='get',
    operation_description="Get approval history",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='Approval history')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def approval_history(request):
    """Get approval history"""
    try:
        lm_approved = BusinessTripRequest.objects.filter(
            line_manager_approved_by=request.user,
            is_deleted=False
        ).order_by('-line_manager_approved_at')[:20]
        
        finance_approved = BusinessTripRequest.objects.filter(
            finance_approved_by=request.user,
            is_deleted=False
        ).order_by('-finance_approved_at')[:20]
        
        hr_approved = BusinessTripRequest.objects.filter(
            hr_approved_by=request.user,
            is_deleted=False
        ).order_by('-hr_approved_at')[:20]
        
        rejected = BusinessTripRequest.objects.filter(
            rejected_by=request.user,
            is_deleted=False
        ).order_by('-rejected_at')[:20]
        
        history = []
        
        for req in lm_approved:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'travel_type': req.travel_type.name,
                'destination': f"{req.start_date} to {req.end_date}",
                'status': 'Approved (Line Manager)',
                'action': 'Approved',
                'comment': req.line_manager_comment,
                'date': req.line_manager_approved_at
            })
        
        for req in finance_approved:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'travel_type': req.travel_type.name,
                'destination': f"{req.start_date} to {req.end_date}",
                'amount': float(req.finance_amount) if req.finance_amount else None,
                'status': 'Approved (Finance)',
                'action': 'Approved',
                'comment': req.finance_comment,
                'date': req.finance_approved_at
            })
        
        for req in hr_approved:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'travel_type': req.travel_type.name,
                'destination': f"{req.start_date} to {req.end_date}",
                'status': 'Approved (HR)',
                'action': 'Approved',
                'comment': req.hr_comment,
                'date': req.hr_approved_at
            })
        
        for req in rejected:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'travel_type': req.travel_type.name,
                'destination': f"{req.start_date} to {req.end_date}",
                'status': req.get_status_display(),
                'action': 'Rejected',
                'comment': req.rejection_reason,
                'date': req.rejected_at
            })
        
        history.sort(key=lambda x: x['date'] if x['date'] else datetime.min, reverse=True)
        
        return Response({'history': history[:20]})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== CANCEL TRIP ====================


@swagger_auto_schema(
    method='post',
    operation_description="Cancel approved trip",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='Trip cancelled')}
)
@api_view(['POST'])
@has_business_trip_permission('business_trips.request.cancel')
@permission_classes([IsAuthenticated])
def cancel_trip(request, pk):
    """Cancel approved trip"""
    try:
        trip_req = BusinessTripRequest.objects.get(pk=pk, is_deleted=False)
        
        # Only approved trips can be cancelled
        if trip_req.status != 'APPROVED':
            return Response({'error': 'Only approved trips can be cancelled'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check permission - employee özünü və ya admin
        emp = Employee.objects.get(user=request.user)
        if trip_req.employee != emp and not is_admin_user(request.user):
            return Response({'error': 'You can only cancel your own trips'}, status=status.HTTP_403_FORBIDDEN)
        
      
        
        # ✅ SEND CANCELLATION NOTIFICATION TO ALL APPROVERS
        graph_token = get_graph_access_token(request.user)

        trip_req.status = 'CANCELLED'
        trip_req.cancelled_by = request.user
        trip_req.cancelled_at = timezone.now()
        trip_req.save()
        
        # ✅ SEND CANCELLATION NOTIFICATION TO ALL APPROVERS
        if graph_token:
            notification_manager.notify_trip_cancelled(trip_req, graph_token)
            logger.info("✅ Cancellation notifications sent to all approvers")
        
        return Response({
            'message': 'Trip cancelled successfully.',
            'notification_sent': bool(graph_token),
            'request': BusinessTripRequestDetailSerializer(trip_req).data
        })
        
    except BusinessTripRequest.DoesNotExist:
        return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== EXPORT ====================

@swagger_auto_schema(
    method='get',
    operation_description="Export my trips to Excel",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='Excel file')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_my_trips(request):
    """Export my trips to Excel (hər kəs özünkünü export edə bilər)"""
    try:
        emp = Employee.objects.get(user=request.user)
        requests = BusinessTripRequest.objects.filter(
            employee=emp, 
            is_deleted=False
        ).order_by('-created_at')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "My Business Trips"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        # Title
        ws['A1'] = f'BUSINESS TRIPS - {emp.full_name}'
        ws['A1'].font = Font(size=16, bold=True, color="2B4C7E")
        ws.merge_cells('A1:K1')
        
        ws['A2'] = f'Employee ID: {getattr(emp, "employee_id", "N/A")}'
        ws['A3'] = f'Department: {emp.department.name if emp.department else "N/A"}'
        ws['A4'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        
        # Headers
        headers = [
            'Request ID', 'Travel Type', 'Transport', 'Purpose', 
            'Start Date', 'End Date', 'Days', 'Status', 'Amount', 
            'Comment', 'Created At'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for row, req in enumerate(requests, 7):
            data = [
                req.request_id,
                req.travel_type.name,
                req.transport_type.name,
                req.purpose.name,
                req.start_date.strftime('%Y-%m-%d'),
                req.end_date.strftime('%Y-%m-%d'),
                float(req.number_of_days),
                req.get_status_display(),
                float(req.finance_amount) if req.finance_amount else '',
                req.comment,
                req.created_at.strftime('%Y-%m-%d %H:%M')
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        filename = f'my_trips_{emp.full_name.replace(" ", "_")}_{date.today().strftime("%Y%m%d")}.xlsx'
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        
        return response
    
    except Employee.DoesNotExist:
        return Response({'error': 'Employee profile not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== VIEWSETS ====================

class TravelTypeViewSet(viewsets.ModelViewSet):
    """Travel Type CRUD ViewSet"""
    queryset = TravelType.objects.filter(is_deleted=False, is_active=True)
    serializer_class = TravelTypeSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return TravelType.objects.filter(is_deleted=False, is_active=True).order_by('name')
    
    def list(self, request, *args, **kwargs):
        # Hər kəs list görə bilər
        return super().list(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({
                'error': 'İcazə yoxdur',
                'required_permission': 'business_trips.settings.update'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({
                'error': 'İcazə yoxdur',
                'required_permission': 'business_trips.settings.update'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({
                'error': 'İcazə yoxdur',
                'required_permission': 'business_trips.settings.update'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({
                'error': 'İcazə yoxdur',
                'required_permission': 'business_trips.settings.update'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

# ==================== CONFIGURATION OPTIONS ====================

@swagger_auto_schema(
    method='get',
    operation_description="Get all configuration options for form dropdowns",
    operation_summary="Get All Options",
    tags=['Business Trip'],
    responses={200: openapi.Response(description='All configuration options')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_options(request):
    """Get all configuration options for trip forms"""
    try:
        travel_types = TravelTypeSerializer(
            TravelType.objects.filter(is_active=True, is_deleted=False).order_by('name'), 
            many=True
        ).data
        
        transport_types = TransportTypeSerializer(
            TransportType.objects.filter(is_active=True, is_deleted=False).order_by('name'), 
            many=True
        ).data
        
        trip_purposes = TripPurposeSerializer(
            TripPurpose.objects.filter(is_active=True, is_deleted=False).order_by('name'), 
            many=True
        ).data
        
        # Get user's employee info for defaults
        try:
            emp = Employee.objects.get(user=request.user)
            user_defaults = {
                'employee_name': emp.full_name,
                'job_function': emp.job_function.name if emp.job_function else '',
                'department': emp.department.name if emp.department else '',
                'unit': emp.unit.name if emp.unit else '',
                'business_function': emp.business_function.name if emp.business_function else '',
                'phone_number': emp.phone or '',
                'line_manager': {
                    'id': emp.line_manager.id if emp.line_manager else None,
                    'name': emp.line_manager.full_name if emp.line_manager else '',
                }
            }
        except Employee.DoesNotExist:
            user_defaults = {}
        
        return Response({
            'travel_types': travel_types,
            'transport_types': transport_types,
            'trip_purposes': trip_purposes,
            'user_defaults': user_defaults
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class TransportTypeViewSet(viewsets.ModelViewSet):
    """Transport Type CRUD ViewSet"""
    queryset = TransportType.objects.filter(is_deleted=False, is_active=True)
    serializer_class = TransportTypeSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return TransportType.objects.filter(is_deleted=False, is_active=True).order_by('name')
    
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({'error': 'İcazə yoxdur'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({'error': 'İcazə yoxdur'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({'error': 'İcazə yoxdur'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({'error': 'İcazə yoxdur'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

class TripPurposeViewSet(viewsets.ModelViewSet):
    """Trip Purpose CRUD ViewSet"""
    queryset = TripPurpose.objects.filter(is_deleted=False, is_active=True)
    serializer_class = TripPurposeSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return TripPurpose.objects.filter(is_deleted=False, is_active=True).order_by('name')
    
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({'error': 'İcazə yoxdur'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({'error': 'İcazə yoxdur'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({'error': 'İcazə yoxdur'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        has_perm, _ = check_business_trip_permission(request.user, 'business_trips.settings.update')
        if not has_perm:
            return Response({'error': 'İcazə yoxdur'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

# ==================== ALL REQUESTS (ADMIN VIEW) ====================

@swagger_auto_schema(
    method='get',
    operation_description="Get all trip requests (admin view with filters)",
    tags=['Business Trip'],
    manual_parameters=[
        openapi.Parameter('status', openapi.IN_QUERY, type=openapi.TYPE_STRING),
        openapi.Parameter('travel_type_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
        openapi.Parameter('department_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
        openapi.Parameter('start_date', openapi.IN_QUERY, type=openapi.TYPE_STRING),
        openapi.Parameter('end_date', openapi.IN_QUERY, type=openapi.TYPE_STRING),
    ],
    responses={200: openapi.Response(description='All requests')}
)
@api_view(['GET'])
@has_business_trip_permission('business_trips.request.view_statistics')
@permission_classes([IsAuthenticated])
def all_trip_requests(request):
    """Get all trip requests with filters"""
    try:
        requests_qs = BusinessTripRequest.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'travel_type', 'transport_type', 'purpose'
        )
        
        # Apply filters
        status = request.GET.get('status')
        if status:
            requests_qs = requests_qs.filter(status=status)
        
        travel_type_id = request.GET.get('travel_type_id')
        if travel_type_id:
            requests_qs = requests_qs.filter(travel_type_id=travel_type_id)
        
        department_id = request.GET.get('department_id')
        if department_id:
            requests_qs = requests_qs.filter(employee__department_id=department_id)
        
        start_date = request.GET.get('start_date')
        if start_date:
            requests_qs = requests_qs.filter(start_date__gte=start_date)
        
        end_date = request.GET.get('end_date')
        if end_date:
            requests_qs = requests_qs.filter(end_date__lte=end_date)
        
        requests_qs = requests_qs.order_by('-created_at')
        
        return Response({
            'count': requests_qs.count(),
            'requests': BusinessTripRequestListSerializer(requests_qs, many=True).data
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='get',
    operation_description="Export all trips to Excel with enhanced formatting",
    tags=['Business Trip'],
    manual_parameters=[
        openapi.Parameter('status', openapi.IN_QUERY, type=openapi.TYPE_STRING),
        openapi.Parameter('year', openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
        openapi.Parameter('department_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
    ],
    responses={200: openapi.Response(description='Excel file')}
)
@api_view(['GET'])
@has_business_trip_permission('business_trips.request.view_statistics')
@permission_classes([IsAuthenticated])
def export_all_trips(request):
    """Export all trips to Excel with enhanced formatting"""
    try:
        requests_qs = BusinessTripRequest.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'employee__business_function',
            'travel_type', 'transport_type', 'purpose',
            'line_manager', 'finance_approver', 'hr_representative'
        )
        
        # Apply filters
        status_filter = request.GET.get('status')
        if status_filter:
            requests_qs = requests_qs.filter(status=status_filter)
        
        year = request.GET.get('year')
        if year:
            requests_qs = requests_qs.filter(start_date__year=year)
        
        department_id = request.GET.get('department_id')
        if department_id:
            requests_qs = requests_qs.filter(employee__department_id=department_id)
        
        requests_qs = requests_qs.order_by('-created_at')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Business Trips"
        
        # Styles
        title_font = Font(size=18, bold=True, color="1F4E79")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        # Title
        ws['A1'] = 'BUSINESS TRIPS EXPORT'
        ws['A1'].font = title_font
        ws.merge_cells('A1:N1')
        
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A2'].font = Font(size=10)
        
        # Applied filters
        filters = []
        if status_filter:
            filters.append(f"Status: {status_filter}")
        if year:
            filters.append(f"Year: {year}")
        if department_id:
            filters.append(f"Department ID: {department_id}")
        
        if filters:
            ws['A3'] = f'Filters: {", ".join(filters)}'
            ws['A3'].font = Font(size=10, italic=True)
        
        # Headers
        headers = [
            'Request ID', 'Employee Name', 'Employee ID', 'Department', 'Business Function',
            'Travel Type', 'Transport', 'Purpose', 'Start Date', 'End Date', 'Days',
            'Status', 'Amount', 'Created At'
        ]
        
        header_row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # Status colors
        status_colors = {
            'APPROVED': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'PENDING_LINE_MANAGER': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'PENDING_FINANCE': PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),
            'PENDING_HR': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
            'REJECTED_LINE_MANAGER': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'REJECTED_FINANCE': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'REJECTED_HR': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }
        
        # Data
        for row, req in enumerate(requests_qs, header_row + 1):
            data = [
                req.request_id,
                req.employee.full_name,
                getattr(req.employee, 'employee_id', ''),
                req.employee.department.name if req.employee.department else '',
                req.employee.business_function.name if req.employee.business_function else '',
                req.travel_type.name,
                req.transport_type.name,
                req.purpose.name,
                req.start_date.strftime('%Y-%m-%d'),
                req.end_date.strftime('%Y-%m-%d'),
                float(req.number_of_days),
                req.get_status_display(),
                float(req.finance_amount) if req.finance_amount else '',
                req.created_at.strftime('%Y-%m-%d %H:%M')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # Apply status color
                if col == 12:  # Status column
                    cell.fill = status_colors.get(req.status, PatternFill())
                
                # Alignment
                if col in [11, 13]:  # Numbers
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        filename = f'business_trips_export_{date.today().strftime("%Y%m%d")}.xlsx'
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        
        return response
    
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)