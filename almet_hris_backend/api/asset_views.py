# api/asset_views.py - COMPLETE REWRITE

from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Sum, Count
from django.db import transaction
from django.utils import timezone
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
import logging
import pandas as pd
import traceback

logger = logging.getLogger(__name__)

from .asset_models import (
    AssetCategory, AssetBatch, Asset, AssetAssignment, 
    AssetActivity, EmployeeOffboarding, AssetTransferRequest
)
from .asset_serializers import (
    AssetCategorySerializer, 
    AssetBatchListSerializer, AssetBatchDetailSerializer, AssetBatchCreateSerializer,
    AssetListSerializer, AssetDetailSerializer, AssetCreateSerializer, AssetCreateMultipleSerializer,
    AssetAssignmentSerializer, AssetAssignmentCreateSerializer,
    AssetActivitySerializer,
    AssetAcceptanceSerializer, AssetClarificationRequestSerializer,
    AssetCancellationSerializer, AssetClarificationProvisionSerializer,
    EmployeeOffboardingSerializer,
    AssetTransferRequestSerializer, AssetTransferRequestCreateSerializer,
    AssetBulkUploadSerializer
)
from .asset_permissions import (
    get_asset_access_level, filter_assets_by_access, filter_batches_by_access,
    require_asset_permission, can_user_manage_asset, get_access_summary
)
from .models import Employee
from .system_email_service import system_email_service


# ============================================
# CATEGORY VIEWSET
# ============================================
class AssetCategoryViewSet(viewsets.ModelViewSet):
    """
    Asset Kateqoriyaları
    - Laptop, Monitor, Phone, Keyboard və s.
    """
    
    queryset = AssetCategory.objects.all()
    serializer_class = AssetCategorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Kateqoriya üzrə statistika"""
        category = self.get_object()
        
        total_batches = category.batches.count()
        total_assets = Asset.objects.filter(category=category).count()
        
        # Status breakdown
        status_breakdown = {}
        for status_choice in Asset.STATUS_CHOICES:
            status_code = status_choice[0]
            count = Asset.objects.filter(category=category, status=status_code).count()
            if count > 0:
                status_breakdown[status_code] = {
                    'label': status_choice[1],
                    'count': count
                }
        
        return Response({
            'category': category.name,
            'total_batches': total_batches,
            'total_assets': total_assets,
            'status_breakdown': status_breakdown
        })


# ============================================
# BATCH VIEWSET
# ============================================
class AssetBatchViewSet(viewsets.ModelViewSet):
    """
    🎯 Asset Batch Management
    
    Batch = Partiya (Eyni növdən bir neçə asset)
    
    Əsas əməliyyatlar:
    1. Batch yaratma (SAY BURADAN QEYD EDİLİR)
    2. Batch-dən asset yaratma
    3. Quantity tracking
    """
    
    queryset = AssetBatch.objects.select_related('category', 'created_by').all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'status']
    search_fields = ['batch_number', 'asset_name', 'supplier']
    ordering_fields = ['created_at', 'asset_name', 'available_quantity']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return AssetBatchListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return AssetBatchCreateSerializer
        return AssetBatchDetailSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return filter_batches_by_access(self.request.user, queryset)
    
    @swagger_auto_schema(
        request_body=AssetBatchCreateSerializer,
        responses={
            201: openapi.Response(
                description="Batch yaradıldı",
                schema=AssetBatchDetailSerializer
            )
        }
    )
    @require_asset_permission('create')
    def create(self, request, *args, **kwargs):
        """
        🎯 Batch yaratma - SAY BURADAN QEYD EDİLİR
        
        Nümunə:
        {
            "asset_name": "Dell Latitude 5420",
            "category": 1,
            "initial_quantity": 10,  👈 BURADAN SAY
            "unit_price": 1500.00,
            "purchase_date": "2024-01-15",
            "useful_life_years": 5
        }
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        batch = serializer.save(created_by=request.user)
        
        logger.info(
            f"✅ Batch yaradıldı: {batch.batch_number} - {batch.asset_name} | "
            f"Quantity: {batch.initial_quantity} | "
            f"Value: {batch.total_value}"
        )
        
        return Response({
            'success': True,
            'message': f'Batch yaradıldı: {batch.batch_number}',
            'batch': AssetBatchDetailSerializer(batch, context={'request': request}).data
        }, status=status.HTTP_201_CREATED)
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetCreateMultipleSerializer,
        responses={
            200: openapi.Response(
                description="Asset-lər yaradıldı",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'created_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'assets': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT))
                    }
                )
            )
        }
    )
    @action(detail=True, methods=['post'], url_path='create-assets')
    @require_asset_permission('manage')
    def create_assets_from_batch(self, request, pk=None):
        """
        🎯 Batch-dən asset yaratma
        
        Nümunə:
        {
            "serial_numbers": ["SN001", "SN002", "SN003"]
        }
        
        Prosess:
        1. Batch-in available_quantity yoxlanılır
        2. Asset-lər yaradılır
        3. Batch-in available_quantity azalır
        """
        try:
            batch = self.get_object()
            
            # Serialize
            serializer = AssetCreateMultipleSerializer(
                data={**request.data, 'batch_id': batch.id},
                context={'request': request}
            )
            serializer.is_valid(raise_exception=True)
            
            # Create assets
            created_assets = serializer.save()
            
            return Response({
                'success': True,
                'message': f'{len(created_assets)} asset yaradıldı',
                'created_count': len(created_assets),
                'batch': {
                    'batch_number': batch.batch_number,
                    'available_quantity': batch.available_quantity,
                    'assigned_quantity': batch.assigned_quantity
                },
                'assets': AssetListSerializer(created_assets, many=True, context={'request': request}).data
            })
            
        except Exception as e:
            logger.error(f"❌ Batch-dən asset yaratma xətası: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def assets(self, request, pk=None):
        """Batch-dəki bütün asset-lər"""
        batch = self.get_object()
        assets = batch.assets.all()
        
        # Status breakdown
        status_summary = {}
        for choice in Asset.STATUS_CHOICES:
            count = assets.filter(status=choice[0]).count()
            if count > 0:
                status_summary[choice[0]] = {
                    'status': choice[1],
                    'count': count
                }
        
        return Response({
            'batch_number': batch.batch_number,
            'batch_name': batch.asset_name,
            'quantity_summary': batch.get_quantity_summary(),
            'total_assets': assets.count(),
            'status_summary': status_summary,
            'assets': AssetListSerializer(assets, many=True, context={'request': request}).data
        })
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Batch statistikası"""
        queryset = self.get_queryset()
        
        total_batches = queryset.count()
        active_batches = queryset.filter(status='ACTIVE').count()
        
        # Quantity summary
        quantity_summary = queryset.aggregate(
            total_initial=Sum('initial_quantity'),
            total_available=Sum('available_quantity'),
            total_assigned=Sum('assigned_quantity'),
            total_out_of_stock=Sum('out_of_stock_quantity')
        )
        
        # Financial summary
        financial_summary = queryset.aggregate(
            total_value=Sum('total_value')
        )
        
        return Response({
            'total_batches': total_batches,
            'active_batches': active_batches,
            'quantity_summary': {
                'total_initial': quantity_summary['total_initial'] or 0,
                'total_available': quantity_summary['total_available'] or 0,
                'total_assigned': quantity_summary['total_assigned'] or 0,
                'total_out_of_stock': quantity_summary['total_out_of_stock'] or 0
            },
            'total_value': float(financial_summary['total_value'] or 0)
        })


# ============================================
# ASSET VIEWSET - MAIN
# ============================================
class AssetViewSet(viewsets.ModelViewSet):
    """
    🎯 Asset Management - Əsas sistem
    
    Əməliyyatlar:
    1. Asset yaratma (batch-dən)
    2. Asset təyin etmə (assign)
    3. İşçi qəbul etmə (accept)
    4. Aydınlaşdırma
    5. Geri qaytarma
    """
    
    queryset = Asset.objects.select_related(
        'batch', 'category', 'assigned_to', 'created_by', 'updated_by'
    ).all()
    
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'category', 'batch', 'assigned_to']
    search_fields = ['asset_number', 'serial_number', 'asset_name', 'batch__batch_number']
    ordering_fields = ['created_at', 'asset_name', 'status', 'updated_at']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return AssetListSerializer
        elif self.action == 'create':
            return AssetCreateSerializer
        return AssetDetailSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return filter_assets_by_access(self.request.user, queryset)
    
    @swagger_auto_schema(
        request_body=AssetCreateSerializer,
        responses={201: AssetDetailSerializer}
    )
    @require_asset_permission('create')
    def create(self, request, *args, **kwargs):
        """
        🎯 Asset yaratma
        
        Nümunə:
        {
            "batch_id": 5,
            "serial_number": "SN123456"
        }
        
        ⚠️ Asset yaradanda batch-in available_quantity avtomatik azalır
        """
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            asset = serializer.save()
            
            # Log activity
            AssetActivity.objects.create(
                asset=asset,
                activity_type='CREATED',
                description=f"Asset batch-dən yaradıldı: {asset.batch.batch_number}",
                performed_by=request.user,
                metadata={
                    'batch_number': asset.batch.batch_number,
                    'batch_id': asset.batch.id,
                    'creation_method': 'manual'
                }
            )
            
            logger.info(f"✅ Asset yaradıldı: {asset.asset_number} from {asset.batch.batch_number}")
            
            return Response({
                'success': True,
                'message': f'Asset yaradıldı: {asset.asset_number}',
                'asset': AssetDetailSerializer(asset, context={'request': request}).data
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"❌ Asset yaratma xətası: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetAssignmentCreateSerializer,
        responses={200: openapi.Response(description="Assets assigned")}
    )
    @action(detail=False, methods=['post'], url_path='assign-to-employee')
    @require_asset_permission('manage')
    def assign_to_employee(self, request):
        """
        🎯 Asset-ləri işçiyə təyin etmə
        
        Nümunə:
        {
            "asset_ids": ["uuid1", "uuid2"],
            "employee_id": 123,
            "check_out_date": "2024-01-15",
            "check_out_notes": "Yeni laptop",
            "condition_on_checkout": "GOOD"
        }
        
        Prosess:
        1. Asset status: IN_STOCK → ASSIGNED
        2. Asset assigned_to = employee
        3. Assignment record yaradılır
        4. Email göndərilir
        
        ⚠️ Batch quantity burada dəyişmir (artıq create-də azalıb)
        """
        try:
            serializer = AssetAssignmentCreateSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            employee = serializer.validated_data['employee']
            assets = serializer.validated_data['assets']
            check_out_date = serializer.validated_data['check_out_date']
            check_out_notes = serializer.validated_data.get('check_out_notes', '')
            condition = serializer.validated_data['condition_on_checkout']
            
            assignments_created = []
            
            with transaction.atomic():
                for asset in assets:
                    # Create assignment
                    assignment = AssetAssignment.objects.create(
                        asset=asset,
                        employee=employee,
                        check_out_date=check_out_date,
                        check_out_notes=check_out_notes,
                        condition_on_checkout=condition,
                        assigned_by=request.user
                    )
                    
                    # Update asset
                    asset.status = 'ASSIGNED'
                    asset.assigned_to = employee
                    asset.updated_by = request.user
                    asset.save()
                    
                    # Log activity
                    AssetActivity.objects.create(
                        asset=asset,
                        activity_type='ASSIGNED',
                        description=f"İşçiyə təyin edildi: {employee.full_name} - təsdiq gözlənilir",
                        performed_by=request.user,
                        metadata={
                            'employee_id': employee.employee_id,
                            'employee_name': employee.full_name,
                            'check_out_date': check_out_date.isoformat(),
                            'condition': condition
                        }
                    )
                    
                    assignments_created.append(assignment)
            
            # Send email notification
            self._send_assignment_email(employee, assets, request.user)
            
            logger.info(f"✅ {len(assets)} asset təyin edildi → {employee.full_name}")
            
            return Response({
                'success': True,
                'message': f'{len(assets)} asset təyin edildi: {employee.full_name}',
                'employee': {
                    'id': employee.id,
                    'name': employee.full_name,
                    'employee_id': employee.employee_id
                },
                'assignments': AssetAssignmentSerializer(assignments_created, many=True, context={'request': request}).data
            })
            
        except Exception as e:
            logger.error(f"❌ Assignment xətası: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _send_assignment_email(self, employee, assets, assigned_by):
        """Email notification"""
        try:
            if not employee.user or not employee.user.email:
                logger.warning(f"⚠️ Employee {employee.full_name} - email yoxdur")
                return
            
            asset_list = '<ul>' + ''.join([
                f'<li><strong>{asset.asset_name}</strong> - {asset.serial_number}</li>'
                for asset in assets
            ]) + '</ul>'
            
            html_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2 style="color: #2563eb;">Asset Təyinatı</h2>
                <p>Hörmətli {employee.full_name},</p>
                <p>Sizə aşağıdaki asset-lər təyin edilmişdir:</p>
                {asset_list}
                <p><strong>Zəhmət olmasa təsdiq edin:</strong></p>
                <ul>
                    <li>✅ Qəbul et - Hər şey düzgündürsə</li>
                    <li>❓ Aydınlaşdırma sorğusu - Sualınız varsa</li>
                </ul>
                <p>Təyin edən: <strong>{assigned_by.get_full_name() or assigned_by.username}</strong></p>
                <p>HRIS sisteminə daxil olub təsdiq edə bilərsiniz.</p>
                <hr>
                <p style="color: #6b7280; font-size: 12px;">Bu avtomatik mesajdır</p>
            </body>
            </html>
            """
            
            system_email_service.send_email_as_system(
                from_email='myalmet@almettrading.com',
                to_email=employee.user.email,
                subject=f'Asset Təyinatı - {len(assets)} əşya',
                body_html=html_body
            )
            
            logger.info(f"✅ Email göndərildi → {employee.user.email}")
            
        except Exception as e:
            logger.error(f"❌ Email xətası: {str(e)}")
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetAcceptanceSerializer,
        responses={200: openapi.Response(description="Asset accepted")}
    )
    @action(detail=False, methods=['post'], url_path='accept-assignment')
    def accept_assignment(self, request):
        """
        🎯 İşçi asset-i qəbul edir
        
        Prosess:
        1. Asset status: ASSIGNED → IN_USE
        2. Activity log
        """
        try:
            asset_id = request.data.get('asset_id')
            comments = request.data.get('comments', '')
            
            if not asset_id:
                return Response(
                    {'error': 'asset_id tələb olunur'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            asset = Asset.objects.get(id=asset_id)
            
            # Check permission
            access = get_asset_access_level(request.user)
            if not access['employee'] or asset.assigned_to != access['employee']:
                return Response(
                    {'error': 'Bu asset-i yalnız sizə təyin edilmiş olsa qəbul edə bilərsiniz'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            if not asset.can_be_approved():
                return Response(
                    {'error': f'Asset qəbul edilə bilməz. Status: {asset.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            with transaction.atomic():
                asset.status = 'IN_USE'
                asset.updated_by = request.user
                asset.save()
                
                AssetActivity.objects.create(
                    asset=asset,
                    activity_type='ACCEPTED',
                    description=f"İşçi tərəfindən qəbul edildi: {access['employee'].full_name}",
                    performed_by=request.user,
                    metadata={
                        'comments': comments,
                        'accepted_at': timezone.now().isoformat()
                    }
                )
            
            logger.info(f"✅ Asset qəbul edildi: {asset.asset_number} by {access['employee'].full_name}")
            
            return Response({
                'success': True,
                'message': 'Asset uğurla qəbul edildi',
                'asset_id': str(asset.id),
                'asset_number': asset.asset_number
            })
            
        except Asset.DoesNotExist:
            return Response({'error': 'Asset tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"❌ Accept xətası: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


    @action(detail=False, methods=['get'], url_path='assignments')
    def assignment_history_list(self, request):
        """
        🎯 Assignment History - Bütün assignment-lər
        
        Filters:
        - employee_id: İşçiyə görə
        - asset_id: Asset-ə görə
        - date_from: Başlanğıc tarixi
        - date_to: Bitmə tarixi
        - is_active: Aktiv/completed
        """
        try:
            # Get all assignments
            queryset = AssetAssignment.objects.select_related(
                'asset', 'asset__category', 'asset__batch',
                'employee', 'assigned_by', 'checked_in_by'
            ).all()
            
            # Apply filters
            employee_id = request.query_params.get('employee_id')
            if employee_id:
                queryset = queryset.filter(employee_id=employee_id)
            
            asset_id = request.query_params.get('asset_id')
            if asset_id:
                queryset = queryset.filter(asset_id=asset_id)
            
            date_from = request.query_params.get('date_from')
            if date_from:
                queryset = queryset.filter(check_out_date__gte=date_from)
            
            date_to = request.query_params.get('date_to')
            if date_to:
                queryset = queryset.filter(check_out_date__lte=date_to)
            
            is_active = request.query_params.get('is_active')
            if is_active == 'true':
                queryset = queryset.filter(check_in_date__isnull=True)
            elif is_active == 'false':
                queryset = queryset.filter(check_in_date__isnull=False)
            
            # Search
            search = request.query_params.get('search')
            if search:
                queryset = queryset.filter(
                    Q(asset__asset_name__icontains=search) |
                    Q(asset__serial_number__icontains=search) |
                    Q(employee__full_name__icontains=search) |
                    Q(employee__employee_id__icontains=search)
                )
            
            # Ordering
            queryset = queryset.order_by('-check_out_date')
            
            # Pagination
            page_size = int(request.query_params.get('page_size', 15))
            page = int(request.query_params.get('page', 1))
            
            from django.core.paginator import Paginator
            paginator = Paginator(queryset, page_size)
            page_obj = paginator.get_page(page)
            
            return Response({
                'count': paginator.count,
                'total_pages': paginator.num_pages,
                'current_page': page,
                'page_size': page_size,
                'results': AssetAssignmentSerializer(
                    page_obj, 
                    many=True, 
                    context={'request': request}
                ).data
            })
            
        except Exception as e:
            logger.error(f"❌ Assignment history error: {str(e)}")
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='assignments/export')
    def export_assignments(self, request):
        """
        🎯 Export Assignments to Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.http import HttpResponse
            
            # Get filtered queryset (same filters as assignment_history_list)
            queryset = AssetAssignment.objects.select_related(
                'asset', 'employee', 'assigned_by', 'checked_in_by'
            ).all()
            
            # Apply filters from request body
            employee_id = request.data.get('employee_id')
            if employee_id:
                queryset = queryset.filter(employee_id=employee_id)
            
            date_from = request.data.get('date_from')
            if date_from:
                queryset = queryset.filter(check_out_date__gte=date_from)
            
            date_to = request.data.get('date_to')
            if date_to:
                queryset = queryset.filter(check_out_date__lte=date_to)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Assignments"
            
            # Header style
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Headers
            headers = [
                'Asset Name', 'Serial Number', 'Category',
                'Employee', 'Employee ID', 'Department',
                'Check Out Date', 'Check In Date', 'Duration (days)',
                'Condition Out', 'Condition In', 'Status',
                'Assigned By', 'Checked In By'
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_idx, assignment in enumerate(queryset, start=2):
                ws.cell(row=row_idx, column=1, value=assignment.asset.asset_name)
                ws.cell(row=row_idx, column=2, value=assignment.asset.serial_number)
                ws.cell(row=row_idx, column=3, value=assignment.asset.category.name)
                ws.cell(row=row_idx, column=4, value=assignment.employee.full_name)
                ws.cell(row=row_idx, column=5, value=assignment.employee.employee_id)
                ws.cell(row=row_idx, column=6, value=assignment.employee.department or 'N/A')
                ws.cell(row=row_idx, column=7, value=assignment.check_out_date.strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=8, value=assignment.check_in_date.strftime('%Y-%m-%d') if assignment.check_in_date else 'Active')
                ws.cell(row=row_idx, column=9, value=assignment.duration_days)
                ws.cell(row=row_idx, column=10, value=assignment.condition_on_checkout)
                ws.cell(row=row_idx, column=11, value=assignment.condition_on_checkin or 'N/A')
                ws.cell(row=row_idx, column=12, value='Active' if assignment.is_active else 'Completed')
                ws.cell(row=row_idx, column=13, value=assignment.assigned_by.get_full_name())
                ws.cell(row=row_idx, column=14, value=assignment.checked_in_by.get_full_name() if assignment.checked_in_by else 'N/A')
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=assignments_{timezone.now().strftime("%Y%m%d")}.xlsx'
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"❌ Export assignments error: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='activities')
    def activities(self, request, pk=None):
        """
        🎯 Asset Activities Log
        """
        asset = self.get_object()
        activities = asset.activities.select_related(
            'performed_by'
        ).order_by('-performed_at')
        
        return Response({
            'asset': {
                'id': str(asset.id),
                'asset_number': asset.asset_number,
                'asset_name': asset.asset_name
            },
            'activities': AssetActivitySerializer(
                activities, 
                many=True, 
                context={'request': request}
            ).data
        })
        
    @swagger_auto_schema(
        method='post',
        request_body=AssetClarificationRequestSerializer,
        responses={200: openapi.Response(description="Clarification requested")}
    )
    @action(detail=False, methods=['post'], url_path='request-clarification')
    def request_clarification(self, request):
        """
        🎯 İşçi aydınlaşdırma sorğusu göndərir
        
        Prosess:
        1. Asset status: ASSIGNED → NEED_CLARIFICATION
        2. Clarification məlumatları saxlanılır
        """
        try:
            asset_id = request.data.get('asset_id')
            reason = request.data.get('clarification_reason')
            
            if not asset_id or not reason:
                return Response(
                    {'error': 'asset_id və clarification_reason tələb olunur'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            asset = Asset.objects.get(id=asset_id)
            
            access = get_asset_access_level(request.user)
            if not access['employee'] or asset.assigned_to != access['employee']:
                return Response(
                    {'error': 'Bu asset-i yalnız sizə təyin edilmiş olsa aydınlaşdırma sorğusu göndərə bilərsiniz'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            if asset.status not in ['ASSIGNED', 'NEED_CLARIFICATION']:
                return Response(
                    {'error': f'Aydınlaşdırma sorğusu göndərilə bilməz. Status: {asset.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            with transaction.atomic():
                asset.status = 'NEED_CLARIFICATION'
                asset.clarification_requested_reason = reason
                asset.clarification_requested_at = timezone.now()
                asset.clarification_requested_by = request.user
                asset.clarification_response = None
                asset.clarification_provided_at = None
                asset.clarification_provided_by = None
                asset.updated_by = request.user
                asset.save()
                
                AssetActivity.objects.create(
                    asset=asset,
                    activity_type='CLARIFICATION_REQUESTED',
                    description=f"Aydınlaşdırma sorğusu: {access['employee'].full_name}",
                    performed_by=request.user,
                    metadata={'reason': reason}
                )
            
            logger.info(f"✅ Aydınlaşdırma sorğusu: {asset.asset_number}")
            
            return Response({
                'success': True,
                'message': 'Aydınlaşdırma sorğusu göndərildi',
                'asset_id': str(asset.id)
            })
            
        except Asset.DoesNotExist:
            return Response({'error': 'Asset tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"❌ Clarification xətası: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetClarificationProvisionSerializer,
        responses={200: openapi.Response(description="Clarification provided")}
    )
    @action(detail=False, methods=['post'], url_path='provide-clarification')
    @require_asset_permission('manage')
    def provide_clarification(self, request):
        """
        🎯 Admin/Manager aydınlaşdırma cavabı verir
        
        Prosess:
        1. Asset status: NEED_CLARIFICATION → ASSIGNED
        2. Cavab saxlanılır
        """
        try:
            asset_id = request.data.get('asset_id')
            response_text = request.data.get('clarification_response')
            
            if not asset_id or not response_text:
                return Response(
                    {'error': 'asset_id və clarification_response tələb olunur'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            asset = Asset.objects.get(id=asset_id)
            
            if asset.status != 'NEED_CLARIFICATION':
                return Response(
                    {'error': f'Asset aydınlaşdırma gözləmir. Status: {asset.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            with transaction.atomic():
                asset.status = 'ASSIGNED'
                asset.clarification_response = response_text
                asset.clarification_provided_at = timezone.now()
                asset.clarification_provided_by = request.user
                asset.updated_by = request.user
                asset.save()
                
                AssetActivity.objects.create(
                    asset=asset,
                    activity_type='CLARIFICATION_PROVIDED',
                    description=f"Aydınlaşdırma cavabı verildi",
                    performed_by=request.user,
                    metadata={'response': response_text}
                )
            
            logger.info(f"✅ Aydınlaşdırma cavabı: {asset.asset_number}")
            
            return Response({
                'success': True,
                'message': 'Aydınlaşdırma cavabı verildi',
                'asset_id': str(asset.id)
            })
            
        except Asset.DoesNotExist:
            return Response({'error': 'Asset tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"❌ Provide clarification xətası: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def assignment_history(self, request, pk=None):
        """Asset-in assignment tarixçəsi"""
        asset = self.get_object()
        assignments = asset.assignments.all().order_by('-check_out_date')
        
        return Response({
            'asset': {
                'id': str(asset.id),
                'asset_number': asset.asset_number,
                'asset_name': asset.asset_name,
                'serial_number': asset.serial_number
            },
            'assignments': AssetAssignmentSerializer(assignments, many=True, context={'request': request}).data
        })
    
    @action(detail=False, methods=['get'])
    def my_assets(self, request):
        """İstifadəçinin öz asset-ləri"""
        access = get_asset_access_level(request.user)
        
        if not access['employee']:
            return Response({'assets': [], 'message': 'Sizin işçi profiliniz yoxdur'})
        
        assets = Asset.objects.filter(
            assigned_to=access['employee']
        ).select_related('batch', 'category')
        
        return Response({
            'employee': {
                'id': access['employee'].id,
                'name': access['employee'].full_name,
                'employee_id': access['employee'].employee_id
            },
            'total_assets': assets.count(),
            'assets': AssetListSerializer(assets, many=True, context={'request': request}).data
        })
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetBulkUploadSerializer,
        manual_parameters=[
            openapi.Parameter(
                'file',
                openapi.IN_FORM,
                description='Excel/CSV file',
                type=openapi.TYPE_FILE,
                required=True
            )
        ],
        consumes=['multipart/form-data']
    )
    @action(detail=False, methods=['post'], url_path='bulk-upload', parser_classes=[MultiPartParser, FormParser])
    @require_asset_permission('create')
    def bulk_upload(self, request):
        """
        🎯 Excel/CSV-dən bulk upload
        
        Excel format:
        | asset_name | category | quantity | serial_numbers | unit_price | purchase_date | supplier |
        """
        try:
            serializer = AssetBulkUploadSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            file = serializer.validated_data['file']
            
            # Read file
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            required_columns = ['asset_name', 'category', 'quantity', 'unit_price', 'purchase_date']
            missing = set(required_columns) - set(df.columns)
            
            if missing:
                return Response(
                    {'error': f'Lazımi sütunlar yoxdur: {", ".join(missing)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            results = {'success': 0, 'failed': 0, 'errors': []}
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # Get or create category
                        category, _ = AssetCategory.objects.get_or_create(
                            name=row['category'],
                            defaults={'created_by': request.user}
                        )
                        
                        # Create batch
                        batch = AssetBatch.objects.create(
                            asset_name=row['asset_name'],
                            category=category,
                            initial_quantity=int(row['quantity']),
                            available_quantity=int(row['quantity']),
                            unit_price=float(row['unit_price']),
                            purchase_date=pd.to_datetime(row['purchase_date']).date(),
                            useful_life_years=int(row.get('useful_life_years', 5)),
                            supplier=row.get('supplier', ''),
                            created_by=request.user
                        )
                        
                        results['success'] += 1
                        
                    except Exception as e:
                        results['failed'] += 1
                        results['errors'].append(f"Sətir {index + 2}: {str(e)}")
            
            logger.info(f"✅ Bulk upload: {results['success']} uğurlu, {results['failed']} uğursuz")
            
            return Response({
                'success': True,
                'imported': results['success'],
                'failed': results['failed'],
                'errors': results['errors'][:10]
            })
            
        except Exception as e:
            logger.error(f"❌ Bulk upload xətası: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def access_info(self, request):
        """İstifadəçinin icazə məlumatları"""
        return Response(get_access_summary(request.user))
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Asset statistikası"""
        queryset = self.get_queryset()
        
        total_assets = queryset.count()
        
        # Status breakdown
        status_breakdown = {}
        for choice in Asset.STATUS_CHOICES:
            count = queryset.filter(status=choice[0]).count()
            if count > 0:
                status_breakdown[choice[0]] = {
                    'label': choice[1],
                    'count': count,
                    'percentage': round((count / total_assets * 100), 1) if total_assets > 0 else 0
                }
        
        # Category breakdown
        category_breakdown = {}
        categories = queryset.values('category__name').annotate(count=Count('id'))
        for cat in categories:
            if cat['category__name']:
                category_breakdown[cat['category__name']] = cat['count']
        
        # Assignment breakdown
        assigned_count = queryset.filter(assigned_to__isnull=False).count()
        unassigned_count = total_assets - assigned_count
        
        return Response({
            'total_assets': total_assets,
            'status_breakdown': status_breakdown,
            'category_breakdown': category_breakdown,
            'assignment_summary': {
                'assigned': assigned_count,
                'unassigned': unassigned_count,
                'assignment_rate': round((assigned_count / total_assets * 100), 1) if total_assets > 0 else 0
            }
        })


# ============================================
# OFFBOARDING VIEWSET
# ============================================
class EmployeeOffboardingViewSet(viewsets.ModelViewSet):
    """
    🎯 Employee Offboarding
    İşdən çıxan işçinin asset-lərinin transferi
    """
    
    queryset = EmployeeOffboarding.objects.select_related('employee', 'created_by', 'approved_by').all()
    serializer_class = EmployeeOffboardingSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['-created_at']
    
    def get_queryset(self):
        access = get_asset_access_level(self.request.user)
        queryset = super().get_queryset()
        
        if access['can_view_all_assets']:
            return queryset
        
        if access['accessible_employee_ids']:
            return queryset.filter(employee_id__in=access['accessible_employee_ids'])
        
        return queryset.none()
    
    @swagger_auto_schema(
        method='post',
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER),
                'last_working_day': openapi.Schema(type=openapi.TYPE_STRING, format='date'),
                'notes': openapi.Schema(type=openapi.TYPE_STRING)
            }
        )
    )
    @action(detail=False, methods=['post'], url_path='initiate')
    @require_asset_permission('manage')
    def initiate_offboarding(self, request):
        """Offboarding prosesini başlat"""
        try:
            employee_id = request.data.get('employee_id')
            last_working_day = request.data.get('last_working_day')
            
            employee = Employee.objects.get(id=employee_id, is_deleted=False)
            
            # Count assets
            assets = Asset.objects.filter(assigned_to=employee, status__in=['ASSIGNED', 'IN_USE'])
            total_assets = assets.count()
            
            offboarding = EmployeeOffboarding.objects.create(
                employee=employee,
                last_working_day=last_working_day,
                total_assets=total_assets,
                notes=request.data.get('notes', ''),
                created_by=request.user
            )
            
            logger.info(f"✅ Offboarding başladıldı: {employee.full_name} - {total_assets} asset")
            
            return Response({
                'success': True,
                'offboarding_id': offboarding.id,
                'employee': employee.full_name,
                'total_assets': total_assets
            })
            
        except Employee.DoesNotExist:
            return Response({'error': 'İşçi tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"❌ Offboarding xətası: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def assets(self, request, pk=None):
        """Offboarding üçün asset-lər"""
        offboarding = self.get_object()
        assets = Asset.objects.filter(
            assigned_to=offboarding.employee,
            status__in=['ASSIGNED', 'IN_USE']
        )
        
        return Response({
            'employee': offboarding.employee.full_name,
            'total_assets': assets.count(),
            'assets': AssetListSerializer(assets, many=True, context={'request': request}).data
        })


# ============================================
# TRANSFER REQUEST VIEWSET
# ============================================
class AssetTransferRequestViewSet(viewsets.ModelViewSet):
    """
    🎯 Asset Transfer Requests (Offboarding)
    """
    
    queryset = AssetTransferRequest.objects.select_related(
        'asset', 'from_employee', 'to_employee', 'requested_by', 'approved_by'
    ).all()
    serializer_class = AssetTransferRequestSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['-requested_at']
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetTransferRequestCreateSerializer
    )
    @action(detail=False, methods=['post'], url_path='create')
    def create_transfer(self, request):
        """Transfer sorğusu yarat"""
        try:
            serializer = AssetTransferRequestCreateSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            asset = serializer.validated_data['asset']
            from_employee = serializer.validated_data['from_employee']
            to_employee = serializer.validated_data['to_employee']
            
            # Get or create offboarding
            offboarding, _ = EmployeeOffboarding.objects.get_or_create(
                employee=from_employee,
                status='IN_PROGRESS',
                defaults={
                    'last_working_day': timezone.now().date(),
                    'total_assets': Asset.objects.filter(assigned_to=from_employee).count(),
                    'created_by': request.user
                }
            )
            
            transfer = AssetTransferRequest.objects.create(
                offboarding=offboarding,
                asset=asset,
                from_employee=from_employee,
                to_employee=to_employee,
                transfer_notes=serializer.validated_data.get('transfer_notes', ''),
                requested_by=request.user
            )
            
            # Send notification
            self._send_transfer_notification(transfer)
            
            logger.info(f"✅ Transfer sorğusu: {asset.asset_number} → {to_employee.full_name}")
            
            return Response({
                'success': True,
                'transfer_id': transfer.id,
                'message': f'Transfer sorğusu yaradıldı'
            })
            
        except Exception as e:
            logger.error(f"❌ Transfer sorğusu xətası: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _send_transfer_notification(self, transfer):
        """Send transfer notification email to involved employees (EN)"""
        try:
            from_employee_email = 'n.nanda@almettrading.co.uk'
            # to_employee_email = transfer.to_employee.user.email
    
            emails = list(filter(None, [from_employee_email]))
    
            if not emails:
                logger.warning("⚠️ Transfer email recipients not found")
                return
    
            html_body = f"""
            <html>
            <body>
                <h2>Asset Transfer Request</h2>
                <p>An asset transfer has been initiated with the following details:</p>
    
                <ul>
                    <li><strong>Asset:</strong> {transfer.asset.asset_name} ({transfer.asset.serial_number})</li>
                    <li><strong>From:</strong> {transfer.from_employee.full_name}</li>
                    <li><strong>To:</strong> {transfer.to_employee.full_name}</li>
                    <li><strong>Requested by:</strong> {transfer.requested_by.get_full_name()}</li>
                    <li><strong>Reason:</strong> Offboarding</li>
                </ul>
    
                <p>The transfer is currently pending approval.</p>
                <p>You will be notified once the transfer is approved or rejected.</p>
    
                <br>
                <p>Best regards,<br>
                HRIS System</p>
            </body>
            </html>
            """
    
            system_email_service.send_email_as_system(
                from_email='myalmet@almettrading.com',
                to_email=emails,
                subject='Asset Transfer Request Notification',
                body_html=html_body
            )
    
            logger.info("✅ Transfer notification email sent to employees")
    
        except Exception as e:
            logger.error(f"❌ Transfer email error: {str(e)}")

    @swagger_auto_schema(
        method='post',
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'approved': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                'rejection_reason': openapi.Schema(type=openapi.TYPE_STRING)
            }
        )
    )
    @action(detail=True, methods=['post'], url_path='approve')
    @require_asset_permission('approve')
    def approve_transfer(self, request, pk=None):
        """Transfer-i təsdiq və ya rədd et"""
        try:
            transfer = self.get_object()
            approved = request.data.get('approved', False)
            
            if approved:
                with transaction.atomic():
                    asset = transfer.asset
                    old_employee = transfer.from_employee
                    new_employee = transfer.to_employee
                    
                    # Check in from old employee
                    active_assignment = asset.assignments.filter(check_in_date__isnull=True).first()
                    if active_assignment:
                        active_assignment.check_in_date = timezone.now().date()
                        active_assignment.checked_in_by = request.user
                        active_assignment.save()
                    
                    # Create new assignment
                    AssetAssignment.objects.create(
                        asset=asset,
                        employee=new_employee,
                        check_out_date=timezone.now().date(),
                        assigned_by=request.user,
                        check_out_notes=f'Transfer: {old_employee.full_name} → {new_employee.full_name}'
                    )
                    
                    # Update asset
                    asset.assigned_to = new_employee
                    asset.status = 'ASSIGNED'
                    asset.updated_by = request.user
                    asset.save()
                    
                    # Update transfer
                    transfer.status = 'COMPLETED'
                    transfer.approved_by = request.user
                    transfer.approved_at = timezone.now()
                    transfer.completed_at = timezone.now()
                    transfer.save()
                    
                    # Update offboarding
                    transfer.offboarding.assets_transferred += 1
                    transfer.offboarding.save()
                    
                    # Log activity
                    AssetActivity.objects.create(
                        asset=asset,
                        activity_type='TRANSFERRED',
                        description=f'Transfer: {old_employee.full_name} → {new_employee.full_name}',
                        performed_by=request.user,
                        metadata={
                            'from_employee': old_employee.full_name,
                            'to_employee': new_employee.full_name,
                            'transfer_id': transfer.id
                        }
                    )
                
                logger.info(f"✅ Transfer təsdiqləndi: {asset.asset_number}")
                
                return Response({
                    'success': True,
                    'message': 'Transfer təsdiqləndi və tamamlandı'
                })
            else:
                # Reject
                transfer.status = 'REJECTED'
                transfer.rejection_reason = request.data.get('rejection_reason', '')
                transfer.approved_by = request.user
                transfer.approved_at = timezone.now()
                transfer.save()
                
                logger.info(f"✅ Transfer rədd edildi: {transfer.asset.asset_number}")
                
                return Response({
                    'success': True,
                    'message': 'Transfer rədd edildi'
                })
            
        except Exception as e:
            logger.error(f"❌ Approve transfer xətası: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)