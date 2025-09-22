# api/vacation_admin.py - Updated Vacation Management System Admin

from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.db.models import Count, Sum
from django.http import HttpResponse
from .vacation_models import (
    VacationSetting, VacationType, EmployeeVacationBalance,
    VacationRequest, VacationActivity, VacationSchedule
)

@admin.register(VacationSetting)
class VacationSettingAdmin(admin.ModelAdmin):
    list_display = [
        'id', 'is_active', 'default_hr_representative', 'allow_negative_balance',
        'max_schedule_edits', 'notification_days_before', 'total_non_working_days',
        'created_at'
    ]
    list_filter = ['is_active', 'allow_negative_balance', 'created_at']
    search_fields = ['default_hr_representative__full_name']
    readonly_fields = ['created_at', 'updated_at']
    
    fieldsets = (
        ('Basic Settings', {
            'fields': ('is_active', 'default_hr_representative')
        }),
        ('Production Calendar Settings', {
            'fields': ('non_working_days',),
            'description': 'Enter non-working days in YYYY-MM-DD format, one per line. These are holidays and special non-working days.'
        }),
        ('Balance Settings', {
            'fields': ('allow_negative_balance',),
            'description': 'Allow employees to request vacation when remaining balance is zero'
        }),
        ('Schedule Settings', {
            'fields': ('max_schedule_edits',),
            'description': 'Maximum number of times a scheduled vacation can be edited'
        }),
        ('Notification Settings', {
            'fields': ('notification_days_before', 'notification_frequency'),
            'description': 'Notification settings for upcoming vacations'
        }),
        ('Metadata', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def total_non_working_days(self, obj):
        return len(obj.non_working_days) if obj.non_working_days else 0
    total_non_working_days.short_description = 'Non-Working Days Count'
    
    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    actions = ['activate_settings']
    
    def activate_settings(self, request, queryset):
        """Activate selected settings (only one can be active)"""
        if queryset.count() != 1:
            self.message_user(request, "Please select exactly one setting to activate.", level='ERROR')
            return
        
        # Deactivate all settings first
        VacationSetting.objects.update(is_active=False)
        
        # Activate selected setting
        selected = queryset.first()
        selected.is_active = True
        selected.save()
        
        self.message_user(request, f"Settings ID {selected.id} activated successfully.")
    
    activate_settings.short_description = "Activate selected settings"

@admin.register(VacationType)
class VacationTypeAdmin(admin.ModelAdmin):
    list_display = [
        'name', 'code', 'requires_approval', 'affects_balance',
        'max_consecutive_days', 'color_display', 'is_active',
        'requests_count', 'created_at'
    ]
    list_filter = ['requires_approval', 'affects_balance', 'is_active', 'created_at']
    search_fields = ['name', 'code', 'description']
    readonly_fields = ['created_at', 'updated_at', 'requests_count']
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'code', 'description')
        }),
        ('Settings', {
            'fields': ('requires_approval', 'affects_balance', 'max_consecutive_days')
        }),
        ('Display', {
            'fields': ('color', 'is_active')
        }),
        ('Statistics', {
            'fields': ('requests_count',),
            'classes': ('collapse',)
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def color_display(self, obj):
        return format_html(
            '<span style="display: inline-block; width: 20px; height: 20px; '
            'background-color: {}; border: 1px solid #ccc; border-radius: 3px;"></span> {}',
            obj.color,
            obj.color
        )
    color_display.short_description = 'Color'
    
    def requests_count(self, obj):
        return obj.vacationrequest_set.count()
    requests_count.short_description = 'Total Requests'

@admin.register(EmployeeVacationBalance)
class EmployeeVacationBalanceAdmin(admin.ModelAdmin):
    list_display = [
        'employee_link', 'year', 'start_balance', 'yearly_balance', 'total_balance_display', 
        'used_days', 'scheduled_days', 'remaining_balance_display', 'should_be_planned_display',
        'balance_status', 'updated_at'
    ]
    list_filter = ['year', 'employee__department', 'employee__business_function']
    search_fields = [
        'employee__full_name', 'employee__employee_id',
        'employee__user__email'
    ]
    readonly_fields = [
        'total_balance_display', 'remaining_balance_display', 'should_be_planned_display',
        'created_at', 'updated_at'
    ]
    
    fieldsets = (
        ('Employee Information', {
            'fields': ('employee', 'year')
        }),
        ('Balance Details', {
            'fields': ('start_balance', 'yearly_balance', 'total_balance_display')
        }),
        ('Usage', {
            'fields': ('used_days', 'scheduled_days', 'remaining_balance_display', 'should_be_planned_display')
        }),
        ('Metadata', {
            'fields': ('updated_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def employee_link(self, obj):
        url = reverse('admin:api_employee_change', args=[obj.employee.pk])
        return format_html('<a href="{}">{}</a>', url, obj.employee.full_name)
    employee_link.short_description = 'Employee'
    employee_link.admin_order_field = 'employee__full_name'
    
    def total_balance_display(self, obj):
        return f"{obj.total_balance} days"
    total_balance_display.short_description = 'Total Balance'
    
    def remaining_balance_display(self, obj):
        remaining = obj.remaining_balance
        if remaining < 0:
            return format_html('<span style="color: red; font-weight: bold;">{} days</span>', remaining)
        return f"{remaining} days"
    remaining_balance_display.short_description = 'Remaining Balance'
    
    def should_be_planned_display(self, obj):
        should_plan = obj.should_be_planned
        if should_plan > 0:
            return format_html('<span style="color: orange; font-weight: bold;">{} days</span>', should_plan)
        return f"{should_plan} days"
    should_be_planned_display.short_description = 'Should Be Planned'
    
    def balance_status(self, obj):
        remaining = obj.remaining_balance
        if remaining < 0:
            color = 'red'
            status = 'Negative'
        elif remaining < 5:
            color = 'orange'
            status = 'Low'
        elif remaining > obj.yearly_balance * 0.8:
            color = 'blue'
            status = 'High'
        else:
            color = 'green'
            status = 'Normal'
        
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, status
        )
    balance_status.short_description = 'Status'
    
    def save_model(self, request, obj, form, change):
        obj.updated_by = request.user
        super().save_model(request, obj, form, change)
    
    actions = ['export_to_excel']
    
    def export_to_excel(self, request, queryset):
        """Export selected balances to Excel"""
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="vacation_balances.xlsx"'
        
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vacation Balances"
        
        # Headers
        headers = [
            'Employee ID', 'Employee Name', 'Year', 'Start Balance', 
            'Yearly Balance', 'Total Balance', 'Used Days', 'Scheduled Days', 
            'Remaining Balance', 'Should Be Planned'
        ]
        ws.append(headers)
        
        # Data
        for balance in queryset:
            ws.append([
                balance.employee.employee_id,
                balance.employee.full_name,
                balance.year,
                float(balance.start_balance),
                float(balance.yearly_balance),
                float(balance.total_balance),
                float(balance.used_days),
                float(balance.scheduled_days),
                float(balance.remaining_balance),
                float(balance.should_be_planned)
            ])
        
        wb.save(response)
        return response
    
    export_to_excel.short_description = "Export selected balances to Excel"

@admin.register(VacationRequest)
class VacationRequestAdmin(admin.ModelAdmin):
    list_display = [
        'request_id', 'employee_link', 'vacation_type', 'request_type',
        'start_date', 'end_date', 'number_of_days', 'status_display',
        'line_manager_link', 'hr_representative_link', 'created_at'
    ]
    list_filter = [
        'status', 'request_type', 'vacation_type', 'start_date',
        'employee__department', 'created_at'
    ]
    search_fields = [
        'request_id', 'employee__full_name', 'employee__employee_id',
        'comment', 'line_manager__full_name', 'hr_representative__full_name'
    ]
    readonly_fields = [
        'request_id', 'return_date', 'number_of_days',
        'line_manager_approved_at', 'hr_approved_at', 'rejected_at',
        'edit_count', 'created_at', 'updated_at', 'can_edit_display'
    ]
    date_hierarchy = 'start_date'
    
    fieldsets = (
        ('Request Information', {
            'fields': ('request_id', 'employee', 'requester', 'request_type', 'can_edit_display')
        }),
        ('Vacation Details', {
            'fields': (
                'vacation_type', 'start_date', 'end_date', 'return_date',
                'number_of_days', 'comment'
            )
        }),
        ('Approval Workflow', {
            'fields': ('line_manager', 'hr_representative', 'status')
        }),
        ('Line Manager Approval', {
            'fields': (
                'line_manager_approved_at', 'line_manager_approved_by',
                'line_manager_comment'
            ),
            'classes': ('collapse',)
        }),
        ('HR Approval', {
            'fields': (
                'hr_approved_at', 'hr_approved_by', 'hr_comment'
            ),
            'classes': ('collapse',)
        }),
        ('Rejection Details', {
            'fields': ('rejected_at', 'rejected_by', 'rejection_reason'),
            'classes': ('collapse',)
        }),
        ('Edit Tracking', {
            'fields': ('edit_count', 'last_edited_at', 'last_edited_by'),
            'classes': ('collapse',)
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def employee_link(self, obj):
        if obj.employee:
            url = reverse('admin:api_employee_change', args=[obj.employee.pk])
            return format_html('<a href="{}">{}</a>', url, obj.employee.full_name)
        return '-'
    employee_link.short_description = 'Employee'
    employee_link.admin_order_field = 'employee__full_name'
    
    def line_manager_link(self, obj):
        if obj.line_manager:
            url = reverse('admin:api_employee_change', args=[obj.line_manager.pk])
            return format_html('<a href="{}">{}</a>', url, obj.line_manager.full_name)
        return '-'
    line_manager_link.short_description = 'Line Manager'
    
    def hr_representative_link(self, obj):
        if obj.hr_representative:
            url = reverse('admin:api_employee_change', args=[obj.hr_representative.pk])
            return format_html('<a href="{}">{}</a>', url, obj.hr_representative.full_name)
        return '-'
    hr_representative_link.short_description = 'HR Representative'
    
    def status_display(self, obj):
        status_colors = {
            'DRAFT': 'gray',
            'IN_PROGRESS': 'blue',
            'PENDING_LINE_MANAGER': 'orange',
            'PENDING_HR': 'purple',
            'APPROVED': 'green',
            'REJECTED_LINE_MANAGER': 'red',
            'REJECTED_HR': 'red',
            'CANCELLED': 'darkred',
            'REGISTERED': 'darkgreen',
            'COMPLETED': 'darkgreen',
        }
        
        color = status_colors.get(obj.status, 'black')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_status_display()
        )
    status_display.short_description = 'Status'
    status_display.admin_order_field = 'status'
    
    def can_edit_display(self, obj):
        can_edit = obj.can_be_edited()
        color = 'green' if can_edit else 'red'
        text = 'Yes' if can_edit else 'No'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, text
        )
    can_edit_display.short_description = 'Can Edit'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'employee', 'requester', 'vacation_type',
            'line_manager', 'hr_representative'
        )
    
    actions = ['export_requests', 'approve_selected', 'reject_selected']
    
    def export_requests(self, request, queryset):
        """Export selected requests to Excel"""
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="vacation_requests.xlsx"'
        
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vacation Requests"
        
        # Headers
        headers = [
            'Request ID', 'Employee Name', 'Employee ID', 'Request Type',
            'Vacation Type', 'Start Date', 'End Date', 'Days', 'Status',
            'Line Manager', 'HR Representative', 'Created At'
        ]
        ws.append(headers)
        
        # Data
        for req in queryset:
            ws.append([
                req.request_id,
                req.employee.full_name,
                req.employee.employee_id,
                req.get_request_type_display(),
                req.vacation_type.name,
                req.start_date.strftime('%Y-%m-%d') if req.start_date else '',
                req.end_date.strftime('%Y-%m-%d') if req.end_date else '',
                float(req.number_of_days),
                req.get_status_display(),
                req.line_manager.full_name if req.line_manager else '',
                req.hr_representative.full_name if req.hr_representative else '',
                req.created_at.strftime('%Y-%m-%d %H:%M') if req.created_at else ''
            ])
        
        wb.save(response)
        return response
    
    export_requests.short_description = "Export selected requests to Excel"

class VacationActivityInline(admin.TabularInline):
    model = VacationActivity
    extra = 0
    readonly_fields = ['activity_type', 'description', 'performed_by', 'created_at']
    can_delete = False
    
    def has_add_permission(self, request, obj=None):
        return False

@admin.register(VacationActivity)
class VacationActivityAdmin(admin.ModelAdmin):
    list_display = [
        'vacation_request_link', 'vacation_schedule_link', 'activity_type', 
        'performed_by', 'short_description', 'created_at'
    ]
    list_filter = ['activity_type', 'created_at', 'performed_by']
    search_fields = [
        'vacation_request__request_id', 'description',
        'performed_by__username'
    ]
    readonly_fields = ['created_at']
    date_hierarchy = 'created_at'
    
    def vacation_request_link(self, obj):
        if obj.vacation_request:
            url = reverse('admin:api_vacationrequest_change', args=[obj.vacation_request.pk])
            return format_html('<a href="{}">{}</a>', url, obj.vacation_request.request_id)
        return '-'
    vacation_request_link.short_description = 'Request'
    vacation_request_link.admin_order_field = 'vacation_request__request_id'
    
    def vacation_schedule_link(self, obj):
        if obj.vacation_schedule:
            url = reverse('admin:api_vacationschedule_change', args=[obj.vacation_schedule.pk])
            return format_html('<a href="{}">Schedule #{}</a>', url, obj.vacation_schedule.id)
        return '-'
    vacation_schedule_link.short_description = 'Schedule'
    
    def short_description(self, obj):
        return obj.description[:50] + '...' if len(obj.description) > 50 else obj.description
    short_description.short_description = 'Description'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'vacation_request', 'vacation_schedule', 'performed_by'
        )

@admin.register(VacationSchedule)
class VacationScheduleAdmin(admin.ModelAdmin):
    list_display = [
        'id', 'employee_link', 'vacation_type', 'start_date', 'end_date',
        'number_of_days', 'status_display', 'edit_count',
        'can_edit_display', 'created_at'
    ]
    list_filter = [
        'status', 'vacation_type', 'start_date',
        'employee__department', 'created_at'
    ]
    search_fields = [
        'employee__full_name', 'employee__employee_id',
        'notes', 'comment'
    ]
    readonly_fields = ['return_date', 'number_of_days', 'created_at', 'updated_at', 'can_edit_display']
    date_hierarchy = 'start_date'
    
    fieldsets = (
        ('Employee Information', {
            'fields': ('employee', 'vacation_type')
        }),
        ('Schedule Details', {
            'fields': ('start_date', 'end_date', 'return_date', 'number_of_days', 'status')
        }),
        ('Edit Tracking', {
            'fields': ('edit_count', 'last_edited_at', 'last_edited_by', 'can_edit_display')
        }),
        ('Additional Information', {
            'fields': ('comment', 'notes')
        }),
        ('Metadata', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def employee_link(self, obj):
        if obj.employee:
            url = reverse('admin:api_employee_change', args=[obj.employee.pk])
            return format_html('<a href="{}">{}</a>', url, obj.employee.full_name)
        return '-'
    employee_link.short_description = 'Employee'
    employee_link.admin_order_field = 'employee__full_name'
    
    def status_display(self, obj):
        color = 'green' if obj.status == 'REGISTERED' else 'blue'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_status_display()
        )
    status_display.short_description = 'Status'
    status_display.admin_order_field = 'status'
    
    def can_edit_display(self, obj):
        can_edit = obj.can_edit()
        color = 'green' if can_edit else 'red'
        text = 'Yes' if can_edit else 'No'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, text
        )
    can_edit_display.short_description = 'Can Edit'
    
    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'employee', 'vacation_type', 'created_by'
        )
    
    actions = ['register_as_taken', 'export_schedules']
    
    def register_as_taken(self, request, queryset):
        """Register selected schedules as taken"""
        count = 0
        for schedule in queryset:
            if schedule.status == 'SCHEDULED':
                try:
                    balance = schedule.get_employee_balance()
                    if balance:
                        balance.use_days(schedule.number_of_days)
                    
                    schedule.status = 'REGISTERED'
                    schedule.save()
                    
                    # Log activity
                    VacationActivity.objects.create(
                        vacation_schedule=schedule,
                        activity_type='REGISTERED',
                        description=f"Schedule registered as taken by admin {request.user.get_full_name()}",
                        performed_by=request.user
                    )
                    count += 1
                except Exception as e:
                    self.message_user(request, f"Error processing schedule {schedule.id}: {str(e)}", level='ERROR')
        
        self.message_user(request, f"{count} schedules registered as taken successfully.")
    
    register_as_taken.short_description = "Register selected schedules as taken"
    
    def export_schedules(self, request, queryset):
        """Export selected schedules to Excel"""
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="vacation_schedules.xlsx"'
        
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vacation Schedules"
        
        # Headers
        headers = [
            'ID', 'Employee Name', 'Employee ID', 'Vacation Type',
            'Start Date', 'End Date', 'Return Date', 'Days', 'Status',
            'Edit Count', 'Comment', 'Created At'
        ]
        ws.append(headers)
        
        # Data
        for schedule in queryset:
            ws.append([
                schedule.id,
                schedule.employee.full_name,
                schedule.employee.employee_id,
                schedule.vacation_type.name,
                schedule.start_date.strftime('%Y-%m-%d') if schedule.start_date else '',
                schedule.end_date.strftime('%Y-%m-%d') if schedule.end_date else '',
                schedule.return_date.strftime('%Y-%m-%d') if schedule.return_date else '',
                float(schedule.number_of_days),
                schedule.get_status_display(),
                schedule.edit_count,
                schedule.comment,
                schedule.created_at.strftime('%Y-%m-%d %H:%M') if schedule.created_at else ''
            ])
        
        wb.save(response)
        return response
    
    export_schedules.short_description = "Export selected schedules to Excel"

# Add vacation activity inline to VacationRequest admin
VacationRequestAdmin.inlines = [VacationActivityInline]